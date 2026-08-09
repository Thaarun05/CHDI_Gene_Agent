"""Offline tests for Section 5a STRING functional network."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from gene_dossier.config import Settings
from gene_dossier.models import AssertionType, EvidenceRecord, ToolResult
from gene_dossier.report_schema import ReportContentBlock, ReportSlot, resolve_report_slot
from gene_dossier.section_5a import (
    DIRECT_QUERY_ASSOCIATION_DEF,
    SECTION_PPI,
    SUBSECTION_5A,
    SUPPLEMENTARY_SCOPE,
    Section5aConfig,
    accept_section_5a_report,
    evaluate_section_5a_complete,
    node_generate_section_5a_derived_artifacts,
    write_string_workbook,
)
from gene_dossier.section_5a_sources import paths_for, sha256_file, write_json_atomic
from gene_dossier.section_bundle import (
    DEFAULT_SECTION_BUNDLE_KEYS,
    SECTION_SOURCE_DEPENDENCIES,
    SUPPORTED_SECTION_BUNDLE_KEYS,
    sources_for_sections,
    validate_section_keys,
)
from gene_dossier.tools import string_db as sd

FIXTURES = Path(__file__).parent / "fixtures" / "string_db"
QUERY_ID = "9606.ENSP00000354476"


def _load_json(name: str):
    return json.loads((FIXTURES / name).read_text(encoding="utf-8"))


def _resolve_tr() -> ToolResult:
    return ToolResult(
        source_name=sd.SOURCE_NAME,
        endpoint_name="resolve_string_identifier",
        gene_symbol="SREBF2",
        request_url="https://version-12-0.string-db.org/api/json/get_string_ids",
        request_params={"identifiers": "SREBF2", "species": "9606"},
        success=True,
        status_code=200,
        data={
            "gene_symbol": "SREBF2",
            "string_id": QUERY_ID,
            "preferred_name": "SREBF2",
            "taxon_name": "Homo sapiens",
            "ncbi_taxon_id": 9606,
            "identifier_status": "resolved",
            "raw": _load_json("get_string_ids_srebf2.json"),
        },
    )


def _network_tr() -> ToolResult:
    return ToolResult(
        source_name=sd.SOURCE_NAME,
        endpoint_name="network",
        gene_symbol="SREBF2",
        request_url="https://version-12-0.string-db.org/api/json/network",
        request_params={
            "identifiers": QUERY_ID,
            "species": "9606",
            "add_nodes": "30",
            "required_score": "400",
            "network_type": "functional",
        },
        success=True,
        status_code=200,
        data=_load_json("network_srebf2.json"),
    )


def _link_tr() -> ToolResult:
    return ToolResult(
        source_name=sd.SOURCE_NAME,
        endpoint_name="get_link",
        gene_symbol="SREBF2",
        request_url="https://version-12-0.string-db.org/api/json/get_link",
        request_params={},
        success=True,
        status_code=200,
        data=_load_json("get_link_srebf2.json"),
    )


def test_5a_in_supported_not_in_default() -> None:
    assert "5a" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert "5a" not in DEFAULT_SECTION_BUNDLE_KEYS
    assert DEFAULT_SECTION_BUNDLE_KEYS[-1] == "4a"
    assert SUPPORTED_SECTION_BUNDLE_KEYS[-1] == "6a"
    assert "5b" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert DEFAULT_SECTION_BUNDLE_KEYS == (
        "1a",
        "1b",
        "1c",
        "1d",
        "1e",
        "2a",
        "2b",
        "2c",
        "3a",
        "4a",
    )
    assert validate_section_keys(["5.a"]) == ["5a"]
    assert SECTION_SOURCE_DEPENDENCIES["5a"] == set()
    assert "STRING" not in sources_for_sections(["5a"])


def test_workbook_five_sheets_scope_and_definition(tmp_path: Path) -> None:
    from gene_dossier.section_5a import canonicalize_network

    rows = _load_json("network_srebf2.json")
    canon = canonicalize_network(
        rows,
        query_string_id=QUERY_ID,
        species_taxon_id=9606,
        required_score=400,
    )
    stats = canon["stats"]
    xlsx = tmp_path / "SREBF2_STRING.xlsx"
    write_string_workbook(
        xlsx,
        gene_symbol="SREBF2",
        summary={
            "string_version": sd.STRING_VERSION,
            "taxon_name": "Homo sapiens",
            "species_taxon_id": 9606,
            "network_type": "functional",
            "required_score": 400,
            "add_nodes": 30,
            "unique_node_count": stats["unique_node_count"],
            "unique_edge_count": stats["unique_edge_count"],
            "direct_query_edge_count": stats["direct_query_edge_count"],
            "neighbor_neighbor_edge_count": stats["neighbor_neighbor_edge_count"],
            "min_combined_score": stats["min_combined_score"],
            "max_combined_score": stats["max_combined_score"],
            "string_network_url": "https://string-db.org/network/x",
        },
        canonical_edges=canon["canonical_edges"],
        direct_partners=canon["direct_query_partners"],
        nodes=canon["nodes"],
        audit_meta={
            "generation_timestamp": "2026-01-01T00:00:00+00:00",
            "network_raw_artifact_id": "raw-1",
            "network_api_run_id": "api-1",
            "network_response_sha256": "deadbeef",
            "warnings": [],
        },
        query_string_id=QUERY_ID,
    )
    wb = load_workbook(xlsx)
    assert wb.sheetnames == [
        "Summary",
        "Network Associations",
        "Direct Query Partners",
        "Network Nodes",
        "Audit",
    ]
    summary_text = " ".join(
        "" if v is None else str(v)
        for row in wb["Summary"].iter_rows(values_only=True)
        for v in row
    )
    audit_text = " ".join(
        "" if v is None else str(v)
        for row in wb["Audit"].iter_rows(values_only=True)
        for v in row
    )
    assert SUPPLEMENTARY_SCOPE in summary_text
    assert DIRECT_QUERY_ASSOCIATION_DEF in summary_text
    assert DIRECT_QUERY_ASSOCIATION_DEF in audit_text
    assert SUPPLEMENTARY_SCOPE in audit_text


def test_ppi_assertion_maps_to_slot_5a_not_2a_or_4a(tmp_path: Path) -> None:
    settings = Settings(output_path=tmp_path)
    state = {
        "gene_symbol": "SREBF2",
        "dossier_run_id": "run-5a-slot",
        "run_type": "section_bundle",
        "selected_section_keys": ["5a"],
        "evidence_records": [],
        "api_runs": [],
        "raw_artifacts": [],
        "tool_results": [],
        "coverage": [],
        "errors": [],
    }
    network_api = MagicMock()
    network_api.id = "api-network-shared"
    resolve_api = MagicMock()
    resolve_api.id = "api-resolve"
    link_api = MagicMock()
    link_api.id = "api-link"

    def _persist(**kwargs):
        tr = kwargs["tool_result"]
        role = kwargs.get("artifact_role") or ""
        if tr.endpoint_name == "network" or role == "network_raw_bytes":
            return (
                network_api,
                {"id": "raw-network", "sha256": "netdigest"},
                "netdigest",
            )
        if tr.endpoint_name == "resolve_string_identifier":
            return (resolve_api, {"id": "raw-resolve", "sha256": "r"}, "r")
        return (link_api, {"id": "raw-link", "sha256": "l"}, "l")

    with (
        patch.object(sd, "resolve_string_identifier", return_value=_resolve_tr()),
        patch.object(sd, "fetch_network", return_value=_network_tr()),
        patch.object(sd, "fetch_network_link", return_value=_link_tr()),
        patch.object(sd, "interaction_partners") as partners,
        patch("gene_dossier.section_5a._persist_string_raw", side_effect=_persist),
    ):
        out = node_generate_section_5a_derived_artifacts(
            state,
            settings=settings,
            persist_db=False,
            config=Section5aConfig(output_root=tmp_path, attempt_network_figure=False),
        )
    partners.assert_not_called()
    records = [
        r
        for r in out["evidence_records"]
        if isinstance(r, EvidenceRecord) and str(r.fact_type or "").startswith("section_5a_")
    ]
    assert records
    network_api_ids = {
        r.api_run_id for r in records if r.fact_type != "section_5a_network_figure"
    }
    # One ApiRun for network partners (shared across summary/partners/workbook)
    assert "api-network-shared" in network_api_ids
    assert network_api_ids == {"api-network-shared"}
    for rec in records:
        assert rec.assertion_type is AssertionType.ppi
        assert rec.section == SECTION_PPI
        assert rec.subsection == SUBSECTION_5A
        slot = resolve_report_slot(rec)
        assert slot == ReportSlot("5", "a")
        assert slot != ReportSlot("2", "a")
        assert slot != ReportSlot("4", "a")


def test_scientific_success_visual_unavailable_when_figure_disabled(tmp_path: Path) -> None:
    settings = Settings(output_path=tmp_path)
    state = {
        "gene_symbol": "SREBF2",
        "dossier_run_id": "run-5a-nofig",
        "run_type": "section_bundle",
        "selected_section_keys": ["5a"],
        "evidence_records": [],
        "api_runs": [],
        "raw_artifacts": [],
        "tool_results": [],
        "coverage": [],
        "errors": [],
    }
    with (
        patch.object(sd, "resolve_string_identifier", return_value=_resolve_tr()),
        patch.object(sd, "fetch_network", return_value=_network_tr()),
        patch.object(sd, "fetch_network_link", return_value=_link_tr()),
        patch.object(sd, "fetch_network_image") as image,
        patch(
            "gene_dossier.section_5a._persist_string_raw",
            return_value=(
                MagicMock(id="api-1"),
                {"id": "raw-1", "sha256": "digest"},
                "digest",
            ),
        ),
    ):
        out = node_generate_section_5a_derived_artifacts(
            state,
            settings=settings,
            persist_db=False,
            config=Section5aConfig(output_root=tmp_path, attempt_network_figure=False),
        )
    image.assert_not_called()
    status = out["section_5a_status"]["rendering_status"]
    assert status["scientific_status"] == "success"
    assert status["visual_status"] == "not_attempted_optional"
    assert status["presentation_status"] == "success"
    summary = out["section_5a_status"]["summary"]
    assert summary["unique_node_count"] == 31
    assert summary["unique_edge_count"] == 238
    assert summary["direct_query_edge_count"] == 30


def test_accept_promote_matrix(tmp_path: Path) -> None:
    paths = paths_for(tmp_path)
    attempt1 = paths.new_gene_attempt("SREBF2", run_id="run1")
    attempt2 = paths.new_gene_attempt("SREBF2", run_id="run2")
    write_json_atomic(attempt1 / "summary.json", {"ok": 1})
    write_json_atomic(attempt2 / "summary.json", {"ok": 2})

    first = accept_section_5a_report(
        paths,
        gene_symbol="SREBF2",
        attempt_dir=attempt1,
        acceptance={"section_5a_complete": True},
        promote_existing=False,
    )
    assert first is not None
    kept = accept_section_5a_report(
        paths,
        gene_symbol="SREBF2",
        attempt_dir=attempt2,
        acceptance={"section_5a_complete": True},
        promote_existing=False,
    )
    assert kept is None
    promoted = accept_section_5a_report(
        paths,
        gene_symbol="SREBF2",
        attempt_dir=attempt2,
        acceptance={"section_5a_complete": True},
        promote_existing=True,
    )
    assert promoted is not None
    pointer = json.loads(paths.accepted_gene_pointer("SREBF2").read_text(encoding="utf-8"))
    assert pointer["attempt_dir"] == str(attempt2)


def test_evaluate_section_5a_complete_checks(tmp_path: Path) -> None:
    from gene_dossier.section_5a import canonicalize_network

    rows = _load_json("network_srebf2.json")
    canon = canonicalize_network(
        rows,
        query_string_id=QUERY_ID,
        species_taxon_id=9606,
        required_score=400,
    )
    attempt = tmp_path / "attempt"
    (attempt / "supplementary").mkdir(parents=True)
    (attempt / "figures").mkdir(parents=True)
    write_json_atomic(
        attempt / "network_associations.json",
        {"edges": canon["canonical_edges"]},
    )
    write_json_atomic(attempt / "network_nodes.json", {"nodes": canon["nodes"]})
    xlsx = attempt / "supplementary" / "SREBF2_STRING.xlsx"
    write_string_workbook(
        xlsx,
        gene_symbol="SREBF2",
        summary={
            "string_version": sd.STRING_VERSION,
            "taxon_name": "Homo sapiens",
            "species_taxon_id": 9606,
            "network_type": "functional",
            "required_score": 400,
            "add_nodes": 30,
            **canon["stats"],
            "string_network_url": "https://string-db.org/network/x",
        },
        canonical_edges=canon["canonical_edges"],
        direct_partners=canon["direct_query_partners"],
        nodes=canon["nodes"],
        audit_meta={"generation_timestamp": "t", "warnings": []},
        query_string_id=QUERY_ID,
    )
    digest = sha256_file(xlsx)
    png = attempt / "figures" / "SREBF2_STRING_network.png"
    png.write_bytes((FIXTURES / "tiny_valid.png").read_bytes())
    html = tmp_path / "out.html"
    pdf = tmp_path / "out.pdf"
    html.write_text(
        '<h2 class="major-heading">5. Protein-protein interaction (PPI) partners</h2>'
        '<h3 class="sub-heading">a. STRING</h3>'
        "<p>intro</p>",
        encoding="utf-8",
    )
    pdf.write_bytes(b"%PDF-1.4")
    blocks = [
        ReportContentBlock(kind="narrative", text="i", presentation_role="section_5a_intro"),
        ReportContentBlock(
            kind="narrative", text="s", presentation_role="section_5a_supplementary_note"
        ),
        ReportContentBlock(
            kind="figure", text="f", presentation_role="section_5a_network_figure"
        ),
        ReportContentBlock(
            kind="narrative", text="l", presentation_role="section_5a_network_legend"
        ),
    ]
    status = {
        "rendering_status": {
            "scientific_status": "success",
            "presentation_status": "success",
            "visual_status": "success",
        },
        "summary": {
            "resolved_string_id": QUERY_ID,
            "unique_edge_count": canon["stats"]["unique_edge_count"],
            "unique_node_count": canon["stats"]["unique_node_count"],
            "direct_query_edge_count": canon["stats"]["direct_query_edge_count"],
            "neighbor_neighbor_edge_count": canon["stats"]["neighbor_neighbor_edge_count"],
            "network_response_sha256": "netsha",
            "supplementary_xlsx": "SREBF2_STRING.xlsx",
            "supplementary_xlsx_sha256": digest,
            "network_figure_sha256": sha256_file(png),
            "network_figure_relative_path": "figures/SREBF2_STRING_network.png",
        },
        "audit": {
            "gene_attempt_dir": str(attempt),
            "artifacts": {"network_figure": "figures/SREBF2_STRING_network.png"},
            "network_response_sha256": "netsha",
        },
    }
    result = evaluate_section_5a_complete(
        status=status,
        attempt_dir=attempt,
        html_path=html,
        pdf_path=pdf,
        presentation_blocks=blocks,
    )
    assert result["complete"] is True
    assert result["section_5a_complete"] is True
    assert result["checks"]
    assert all(v.get("passed") for v in result["checks"].values())

    status["rendering_status"]["scientific_status"] = "no_interactions"
    failed = evaluate_section_5a_complete(
        status=status,
        attempt_dir=attempt,
        html_path=html,
        pdf_path=pdf,
        presentation_blocks=blocks,
    )
    assert failed["complete"] is False
    assert failed["checks"]["scientific_status_success"]["passed"] is False


def test_config_rejects_add_nodes_not_30() -> None:
    Section5aConfig()
    import pytest

    with pytest.raises(ValueError, match="add_nodes"):
        Section5aConfig(add_nodes=10)
    with pytest.raises(ValueError, match="add_nodes"):
        Section5aConfig(add_nodes=31)
    with pytest.raises(ValueError, match="network_type"):
        Section5aConfig(network_type="physical")
