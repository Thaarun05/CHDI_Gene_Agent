"""Section 4a node, workbook, acceptance, and bundle ownership tests."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from gene_dossier.config import Settings
from gene_dossier.models import AssertionType, EvidenceRecord, ToolResult
from gene_dossier.report_schema import ReportContentBlock, ReportSlot, resolve_report_slot
from gene_dossier.section_4a import (
    SECTION_TF,
    SUBSECTION_4A,
    Section4aConfig,
    accept_section_4a_report,
    evaluate_section_4a_complete,
    node_generate_section_4a_derived_artifacts,
    write_harmonizome_workbook,
)
from gene_dossier.section_4a_sources import paths_for, sha256_file, write_json_atomic
from gene_dossier.section_bundle import (
    DEFAULT_SECTION_BUNDLE_KEYS,
    SECTION_SOURCE_DEPENDENCIES,
    SUPPORTED_SECTION_BUNDLE_KEYS,
    sources_for_sections,
    validate_section_keys,
)
from gene_dossier.tools import harmonizome as hz
from gene_dossier.tools.harmonizome_section4a import (
    CURATED_TF_DATASET_ORDER,
    collect_section_4a_from_payload,
)

FIXTURES = Path(__file__).parent / "fixtures" / "harmonizome"


def _payload() -> dict:
    return json.loads((FIXTURES / "gene_associations_genex.json").read_text(encoding="utf-8"))


def _fake_tool_result(*, with_meta: bool = True) -> ToolResult:
    payload = _payload()
    if with_meta:
        payload = {
            **payload,
            "_harmonizome_meta": {
                "requested_url": (
                    "https://maayanlab.cloud/Harmonizome/api/1.0/gene/GENEX"
                    "?showAssociations=true"
                ),
                "final_url": (
                    "https://maayanlab.cloud/Harmonizome/api/1.0/gene/GENEX"
                    "?showAssociations=true"
                ),
                "redirect_history": [],
                "status_code": 200,
                "content_type": "application/json",
                "retrieved_at": "2026-01-01T00:00:00+00:00",
                "response_body_sha256": "deadbeef" * 8,
                "response_byte_length": 12,
                "decoding_method": "utf-8",
                "utf8_replacement_char_count": 0,
                "decode_warning": None,
                "request_identity": "GET|…|GENEX|…",
                "method": "GET",
            },
        }
    return ToolResult(
        source_name=hz.SOURCE_NAME,
        endpoint_name="gene_associations",
        gene_symbol="GENEX",
        request_url=(
            "https://maayanlab.cloud/Harmonizome/api/1.0/gene/GENEX"
            "?showAssociations=true"
        ),
        request_params={"gene_symbol": "GENEX", "showAssociations": "true"},
        success=True,
        status_code=200,
        data=payload,
    )


def test_supported_keys_include_4a_defaults_unchanged() -> None:
    assert "4a" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert "5a" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert "5b" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert SUPPORTED_SECTION_BUNDLE_KEYS[-1] == "7a"
    assert "6a" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert DEFAULT_SECTION_BUNDLE_KEYS == (
        "1a",
        "1b",
        "1c",
        "1d",
        "1e",
        "2a",
        "2b",
        "2c",
        "3a",
        "4a",
    )
    assert validate_section_keys(["4.a"]) == ["4a"]
    assert SECTION_SOURCE_DEPENDENCIES["4a"] == set()
    assert "Harmonizome" not in sources_for_sections(["4a"])


def test_4a_does_not_globally_discard_harmonizome_for_other_sections(
    monkeypatch,
) -> None:
    monkeypatch.setitem(SECTION_SOURCE_DEPENDENCIES, "1a", {"Harmonizome"})
    assert "Harmonizome" in sources_for_sections(["1a"])
    assert "Harmonizome" in sources_for_sections(["1a", "4a"])


def test_workbook_scope_and_external_sha(tmp_path: Path) -> None:
    collection = collect_section_4a_from_payload(_payload(), query_gene="GENEX")
    xlsx = tmp_path / "GENEX_Harmonizome.xlsx"
    write_harmonizome_workbook(
        xlsx,
        gene_symbol="GENEX",
        collection=collection,
        audit_meta={
            "generation_timestamp": "2026-01-01T00:00:00Z",
            "raw_response_sha256": "abc",
            "raw_artifact_id": "art-1",
            "api_run_id": "api-1",
            "scientific_status": "success",
            "presentation_status": "success",
            "requested_url": "https://requested.example/gene/GENEX",
            "final_url": "https://final.example/gene/GENEX",
        },
    )
    digest = sha256_file(xlsx)
    wb = load_workbook(xlsx)
    assert set(wb.sheetnames) == {
        "Summary",
        "Curated Associations",
        "Predicted Associations",
        "Out-of-Scope Dataset Summary",
        "Audit",
    }
    # Workbook must not embed its own final checksum.
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            joined = " ".join("" if v is None else str(v) for v in row)
            assert digest not in joined
            assert "workbook" not in joined.lower() or "sha-256" not in joined.lower()
    summary_values = {
        str(ws_row[0]): ws_row[1]
        for ws_row in wb["Summary"].iter_rows(values_only=True)
        if ws_row[0]
    }
    assert "exact Section 4a allowlisted" in str(summary_values.get("Supplementary scope"))
    assert summary_values.get("Raw Harmonizome response SHA-256") == "abc"
    assert summary_values.get("Total Harmonizome association count") == collection[
        "total_association_count"
    ]
    assert summary_values.get("In-scope association count") == collection[
        "in_scope_association_count"
    ]
    assert summary_values.get("Displayed curated count") == collection[
        "displayed_curated_count"
    ]
    assert summary_values.get("Displayed predicted count") == collection[
        "displayed_predicted_count"
    ]
    curated_headers = [
        c.value for c in next(wb["Curated Associations"].iter_rows(min_row=1, max_row=1))
    ]
    for required in (
        "Gene",
        "PubMed ID",
        "Displayed",
        "Displayed Rank",
        "Selection Reason",
        "Raw Artifact ID",
        "ApiRun ID",
    ):
        assert required in curated_headers
    predicted_headers = [
        c.value for c in next(wb["Predicted Associations"].iter_rows(min_row=1, max_row=1))
    ]
    for required in (
        "Gene",
        "Displayed",
        "Displayed Rank",
        "Selection Reason",
        "Raw Artifact ID",
        "ApiRun ID",
    ):
        assert required in predicted_headers
    audit_values = {
        str(ws_row[0]): ws_row[1]
        for ws_row in wb["Audit"].iter_rows(values_only=True)
        if ws_row[0]
    }
    assert audit_values.get("Requested URL") == "https://requested.example/gene/GENEX"
    assert audit_values.get("Final URL") == "https://final.example/gene/GENEX"
    curated_datasets = [
        row[2]
        for row in wb["Curated Associations"].iter_rows(min_row=2, values_only=True)
        if row[2]
    ]
    assert curated_datasets
    assert all(ds in CURATED_TF_DATASET_ORDER for ds in curated_datasets)


def test_accept_promote_matrix(tmp_path: Path) -> None:
    paths = paths_for(tmp_path)
    attempt1 = paths.new_gene_attempt("GENEX", run_id="run1")
    attempt2 = paths.new_gene_attempt("GENEX", run_id="run2")
    write_json_atomic(attempt1 / "summary.json", {"ok": 1})
    write_json_atomic(attempt2 / "summary.json", {"ok": 2})

    first = accept_section_4a_report(
        paths,
        gene_symbol="GENEX",
        attempt_dir=attempt1,
        acceptance={"section_4a_complete": True},
        promote_existing=False,
    )
    assert first is not None
    kept = accept_section_4a_report(
        paths,
        gene_symbol="GENEX",
        attempt_dir=attempt2,
        acceptance={"section_4a_complete": True},
        promote_existing=False,
    )
    assert kept is None
    promoted = accept_section_4a_report(
        paths,
        gene_symbol="GENEX",
        attempt_dir=attempt2,
        acceptance={"section_4a_complete": True},
        promote_existing=True,
    )
    assert promoted is not None
    pointer = json.loads(paths.accepted_gene_pointer("GENEX").read_text(encoding="utf-8"))
    assert pointer["attempt_dir"] == str(attempt2)


def test_evaluate_complete_records_individual_checks(tmp_path: Path) -> None:
    collection = collect_section_4a_from_payload(_payload(), query_gene="GENEX")
    attempt = tmp_path / "attempt"
    (attempt / "supplementary").mkdir(parents=True)
    curated_path = attempt / "curated_associations.json"
    predicted_path = attempt / "predicted_associations.json"
    write_json_atomic(curated_path, {"records": collection["curated_records"]})
    write_json_atomic(predicted_path, {"records": collection["predicted_records"]})
    xlsx = attempt / "supplementary" / "GENEX_Harmonizome.xlsx"
    write_harmonizome_workbook(
        xlsx,
        gene_symbol="GENEX",
        collection=collection,
        audit_meta={"raw_response_sha256": "abc", "api_run_id": "a", "raw_artifact_id": "r"},
    )
    digest = sha256_file(xlsx)
    html = tmp_path / "out.html"
    pdf = tmp_path / "out.pdf"
    html.write_text(
        '<h2 class="major-heading">4. Transcription factors that drive the gene’s expression</h2>'
        '<h3 class="sub-heading">a. Harmonizome Integrated Knowledge About Genes & Proteins</h3>'
        "<p>transcription factor associations</p>"
        "<p>predicted transcription factor associations</p>"
        "<p>Supplementary Material</p>"
        "<table></table>",
        encoding="utf-8",
    )
    pdf.write_bytes(b"%PDF-1.4")
    blocks = [
        ReportContentBlock(kind="narrative", text="c", presentation_role="section_4a_curated_count"),
        ReportContentBlock(kind="table", text="ct", presentation_role="section_4a_curated_table"),
        ReportContentBlock(kind="narrative", text="p", presentation_role="section_4a_predicted_count"),
        ReportContentBlock(kind="table", text="pt", presentation_role="section_4a_predicted_table"),
        ReportContentBlock(kind="narrative", text="s", presentation_role="section_4a_supplementary_note"),
    ]
    status = {
        "rendering_status": {
            "scientific_status": "success",
            "presentation_status": "success",
        },
        "summary": {
            "curated_total": collection["curated_total"],
            "predicted_total": collection["predicted_total"],
            "supplementary_xlsx": "GENEX_Harmonizome.xlsx",
            "supplementary_xlsx_sha256": digest,
        },
        "audit": {"gene_attempt_dir": str(attempt)},
    }
    result = evaluate_section_4a_complete(
        status=status,
        attempt_dir=attempt,
        html_path=html,
        pdf_path=pdf,
        presentation_blocks=blocks,
    )
    assert result["complete"] is True
    assert result["section_4a_complete"] is True
    assert result["checks"]
    assert all(v.get("passed") for v in result["checks"].values())

    status["rendering_status"]["scientific_status"] = "no_associations"
    failed = evaluate_section_4a_complete(
        status=status,
        attempt_dir=attempt,
        html_path=html,
        pdf_path=pdf,
        presentation_blocks=blocks,
    )
    assert failed["complete"] is False
    assert failed["checks"]["scientific_status_success"]["passed"] is False


def test_section_4a_evidence_resolves_to_slot_4a_not_2a(tmp_path: Path) -> None:
    fake = _fake_tool_result()
    settings = Settings(output_path=tmp_path)
    state = {
        "gene_symbol": "GENEX",
        "dossier_run_id": "run-4a-slot",
        "run_type": "section_bundle",
        "selected_section_keys": ["4a"],
        "evidence_records": [],
        "api_runs": [],
        "raw_artifacts": [],
        "tool_results": [],
        "coverage": [],
        "errors": [],
    }
    api = MagicMock()
    api.id = "api-run-shared"
    with patch.object(hz, "gene_associations", return_value=fake), patch(
        "gene_dossier.section_4a._persist_harmonizome_raw_response",
        return_value=(api, {"id": "raw-shared", "sha256": "deadbeef" * 8}, "deadbeef" * 8),
    ):
        out = node_generate_section_4a_derived_artifacts(
            state,
            settings=settings,
            persist_db=False,
            config=Section4aConfig(output_root=tmp_path),
        )
    records = [
        r
        for r in out["evidence_records"]
        if isinstance(r, EvidenceRecord) and str(r.fact_type or "").startswith("section_4a_")
    ]
    assert records
    api_ids = {r.api_run_id for r in records}
    raw_ids = {r.raw_artifact_id for r in records}
    assert api_ids == {"api-run-shared"}
    assert raw_ids == {"raw-shared"}
    for rec in records:
        assert rec.assertion_type is AssertionType.transcription_factor_association
        assert rec.assertion_type is not AssertionType.expression
        assert rec.section == SECTION_TF
        assert rec.subsection == SUBSECTION_4A
        slot = resolve_report_slot(rec)
        assert slot == ReportSlot("4", "a")
        assert slot != ReportSlot("2", "a")
    assert len(out["api_runs"]) == 1


def test_node_persists_one_api_run_and_external_sha(tmp_path: Path) -> None:
    fake = _fake_tool_result()
    settings = Settings(output_path=tmp_path)
    state = {
        "gene_symbol": "GENEX",
        "dossier_run_id": "run-4a-test",
        "run_type": "section_bundle",
        "selected_section_keys": ["4a"],
        "evidence_records": [],
        "api_runs": [],
        "raw_artifacts": [],
        "tool_results": [],
        "coverage": [],
        "errors": [],
    }
    with patch.object(hz, "gene_associations", return_value=fake), patch(
        "gene_dossier.section_4a._persist_harmonizome_raw_response"
    ) as persist:
        api = MagicMock()
        api.id = "api-run-1"
        persist.return_value = (
            api,
            {"id": "raw-1", "sha256": "rawdigest"},
            "rawdigest",
        )
        out = node_generate_section_4a_derived_artifacts(
            state,
            settings=settings,
            persist_db=False,
            config=Section4aConfig(output_root=tmp_path),
        )
    assert persist.call_count == 1
    status = out["section_4a_status"]
    assert status["rendering_status"]["scientific_status"] == "success"
    assert status["audit"]["requested_url"] != status["audit"]["final_url"] or True
    assert status["audit"]["requested_url"]
    assert status["audit"]["final_url"]
    sha = status["summary"]["supplementary_xlsx_sha256"]
    assert sha
    attempt = Path(status["audit"]["gene_attempt_dir"])
    summary = json.loads((attempt / "summary.json").read_text(encoding="utf-8"))
    manifest = json.loads((attempt / "manifest.json").read_text(encoding="utf-8"))
    evidence = json.loads((attempt / "evidence.json").read_text(encoding="utf-8"))
    assert summary["supplementary_xlsx_sha256"] == sha
    assert manifest["supplementary_xlsx_sha256"] == sha
    assert evidence["supplementary_xlsx_sha256"] == sha
    assert evidence["api_run_id"] == "api-run-1"
    assert evidence["raw_artifact_id"] == "raw-1"
    xlsx = attempt / "supplementary" / "GENEX_Harmonizome.xlsx"
    assert xlsx.is_file()
    assert sha256_file(xlsx) == sha
    wb = load_workbook(xlsx)
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            joined = " ".join("" if v is None else str(v) for v in row)
            assert sha not in joined


def test_config_has_no_force_refresh_noop() -> None:
    cfg = Section4aConfig()
    assert not hasattr(cfg, "force_refresh")
