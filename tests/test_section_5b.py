"""Section 5b BioGRID counting / reconcile tests."""

from __future__ import annotations

from pathlib import Path

from gene_dossier.section_5b import (
    STATUS_NO_INTERACTIONS,
    STATUS_TARGET_MISMATCH,
    build_nonredundant_pairs,
    canonical_pair_key,
    reconcile_target,
    write_biogrid_workbook,
)
from gene_dossier.section_bundle import DEFAULT_SECTION_BUNDLE_KEYS, SUPPORTED_SECTION_BUNDLE_KEYS
from openpyxl import load_workbook


def _row(
    iid: int,
    *,
    a: int,
    b: int,
    symbol_a: str = "SREBF2",
    symbol_b: str = "PARTNER",
    entrez_a: int = 6721,
    entrez_b: int = 100,
    org_a: int = 9606,
    org_b: int = 9606,
    system_type: str = "physical",
    page: int = 0,
    idx: int = 0,
) -> dict:
    return {
        "BIOGRID_INTERACTION_ID": iid,
        "BIOGRID_ID_A": a,
        "BIOGRID_ID_B": b,
        "OFFICIAL_SYMBOL_A": symbol_a,
        "OFFICIAL_SYMBOL_B": symbol_b,
        "ENTREZ_GENE_A": str(entrez_a),
        "ENTREZ_GENE_B": str(entrez_b),
        "ORGANISM_A": org_a,
        "ORGANISM_B": org_b,
        "EXPERIMENTAL_SYSTEM": "Two-hybrid",
        "EXPERIMENTAL_SYSTEM_TYPE": system_type,
        "PUBMED_ID": str(1000 + iid),
        "_page_start": page,
        "_within_page_index": idx,
    }


def test_supported_includes_5b_default_through_4a() -> None:
    assert SUPPORTED_SECTION_BUNDLE_KEYS[-3:] == ("5a", "5b", "6a")
    assert "5b" not in DEFAULT_SECTION_BUNDLE_KEYS
    assert DEFAULT_SECTION_BUNDLE_KEYS[-1] == "4a"


def test_srebf2_shaped_378_to_325() -> None:
    target_bg = 112599
    rows = []
    iid = 1
    # 325 unique partners
    for i in range(325):
        rows.append(
            _row(
                iid,
                a=target_bg,
                b=2000 + i,
                symbol_b=f"P{i}",
                entrez_b=10000 + i,
                page=0,
                idx=i,
            )
        )
        iid += 1
    # 53 duplicate evidence on first 53 pairs (same unordered pair, different pubmed/system)
    for i in range(53):
        rows.append(
            _row(
                iid,
                a=2000 + i,  # reversed orientation
                b=target_bg,
                symbol_a=f"P{i}",
                symbol_b="SREBF2",
                entrez_a=10000 + i,
                entrez_b=6721,
                system_type="genetic",
                page=0,
                idx=325 + i,
            )
        )
        iid += 1
    assert len(rows) == 378
    target, err, viol = reconcile_target(rows, gene_symbol="SREBF2", expected_entrez_id=6721)
    assert err is None and not viol
    assert target["target_biogrid_id"] == target_bg
    built = build_nonredundant_pairs(rows, target)
    assert built["counts"]["raw_evidence_record_count"] == 378
    assert built["counts"]["nonredundant_pair_count"] == 325
    assert built["counts"]["physical_evidence_record_count"] == 325
    assert built["counts"]["genetic_evidence_record_count"] == 53


def test_cdh10_shaped_11_to_8() -> None:
    target_bg = 107443
    partners = [10, 11, 12, 13, 14, 15, 16, 17]
    rows = []
    iid = 1
    for i, p in enumerate(partners):
        rows.append(
            _row(
                iid,
                a=target_bg,
                b=p,
                symbol_a="CDH10",
                symbol_b=f"C{i}",
                entrez_a=1008,
                entrez_b=2000 + i,
                page=0,
                idx=i,
            )
        )
        iid += 1
    # 3 duplicate evidence rows
    for i in range(3):
        rows.append(
            _row(
                iid,
                a=partners[i],
                b=target_bg,
                symbol_a=f"C{i}",
                symbol_b="CDH10",
                entrez_a=2000 + i,
                entrez_b=1008,
                page=0,
                idx=8 + i,
            )
        )
        iid += 1
    assert len(rows) == 11
    target, err, _ = reconcile_target(rows, gene_symbol="CDH10", expected_entrez_id=1008)
    assert err is None
    built = build_nonredundant_pairs(rows, target)
    assert built["counts"]["raw_evidence_record_count"] == 11
    assert built["counts"]["nonredundant_pair_count"] == 8


def test_self_and_cross_species_kept() -> None:
    target_bg = 112599
    rows = [
        _row(1, a=target_bg, b=target_bg, symbol_b="SREBF2", entrez_b=6721, idx=0),
        _row(
            2,
            a=target_bg,
            b=999,
            symbol_b="MOUSEG",
            entrez_b=50,
            org_b=10090,
            idx=1,
        ),
    ]
    target, err, _ = reconcile_target(rows, gene_symbol="SREBF2", expected_entrez_id=6721)
    assert err is None
    built = build_nonredundant_pairs(rows, target)
    assert built["counts"]["nonredundant_pair_count"] == 2
    assert built["counts"]["self_interaction_evidence_record_count"] == 1
    assert built["counts"]["cross_species_evidence_record_count"] == 1
    self_pair = next(p for p in built["pairs"] if p["self_interaction"])
    assert self_pair["partner_biogrid_id"] == target_bg


def test_target_mismatch_conflicting_biogrid_id() -> None:
    rows = [
        _row(1, a=112599, b=10, idx=0),
        _row(
            2,
            a=555555,
            b=11,
            symbol_a="SREBF2",
            entrez_a=6721,
            idx=1,
        ),
    ]
    target, err, _ = reconcile_target(rows, gene_symbol="SREBF2", expected_entrez_id=6721)
    assert err == STATUS_TARGET_MISMATCH or target is None or err == STATUS_TARGET_MISMATCH


def test_stable_primary_uses_page_then_index() -> None:
    target_bg = 1
    rows = [
        _row(99, a=target_bg, b=2, page=1, idx=0, symbol_a="G", entrez_a=10),
        _row(5, a=2, b=target_bg, page=0, idx=1, symbol_a="P", symbol_b="G", entrez_a=20, entrez_b=10),
        _row(7, a=target_bg, b=2, page=0, idx=0, symbol_a="G", entrez_a=10),
    ]
    target, err, _ = reconcile_target(rows, gene_symbol="G", expected_entrez_id=10)
    assert err is None
    built = build_nonredundant_pairs(
        rows,
        target,
        page_api_runs={0: "api0", 1: "api1"},
        page_raw_ids={0: "raw0", 1: "raw1"},
    )
    pair = built["pairs"][0]
    assert pair["primary_api_run_id"] == "api0"
    assert pair["primary_raw_artifact_id"] == "raw0"
    assert pair["contributing_api_run_ids"] == ["api0", "api1"]
    assert pair["evidence_record_count"] == 3


def test_stable_primary_biogrid_id_tiebreaker() -> None:
    """Same page_start and within_page_index → BIOGRID_INTERACTION_ID ascending."""
    target_bg = 1
    rows = [
        _row(50, a=target_bg, b=2, page=0, idx=0, symbol_a="G", entrez_a=10),
        _row(10, a=target_bg, b=2, page=0, idx=0, symbol_a="G", entrez_a=10),
        _row(30, a=target_bg, b=2, page=0, idx=0, symbol_a="G", entrez_a=10),
    ]
    target, err, _ = reconcile_target(rows, gene_symbol="G", expected_entrez_id=10)
    assert err is None
    built = build_nonredundant_pairs(
        rows, target, page_api_runs={0: "api0"}, page_raw_ids={0: "raw0"}
    )
    assert built["pairs"][0]["primary_biogrid_interaction_id"] == 10
    assert [r["BIOGRID_INTERACTION_ID"] for r in built["ordered_rows"]] == [10, 30, 50]


def test_zero_interactions_workbook_empty_with_provenance(tmp_path: Path) -> None:
    path = tmp_path / "GENEX_BIOGRID.xlsx"
    write_biogrid_workbook(
        path,
        gene_symbol="GENEX",
        target=None,
        pairs=[],
        evidence_rows=[],
        provenance=[
            ("Gene", "GENEX"),
            ("raw_evidence_record_count", 0),
            ("nonredundant_pair_count", 0),
            ("BioGRID version", "4.4.242"),
        ],
    )
    wb = load_workbook(path, read_only=True)
    assert "Nonredundant_Interactions" in wb.sheetnames
    assert "Evidence_Records" in wb.sheetnames
    assert "Provenance" in wb.sheetnames
    assert (wb["Evidence_Records"].max_row or 1) - 1 == 0
    assert (wb["Nonredundant_Interactions"].max_row or 1) - 1 == 0
    assert (wb["Provenance"].max_row or 1) - 1 >= 1
    wb.close()
    assert STATUS_NO_INTERACTIONS == "no_interactions"


def test_workbook_raw_rows_not_polished_count(tmp_path: Path) -> None:
    target = {
        "target_symbol": "G",
        "target_entrez_id": 10,
        "target_biogrid_id": 1,
        "target_taxon_id": 9606,
    }
    rows = [
        _row(1, a=1, b=2, symbol_a="G", entrez_a=10, idx=0),
        _row(2, a=2, b=1, symbol_a="P", symbol_b="G", entrez_a=20, entrez_b=10, idx=1),
    ]
    built = build_nonredundant_pairs(rows, target)
    path = tmp_path / "G_BIOGRID.xlsx"
    write_biogrid_workbook(
        path,
        gene_symbol="G",
        target=target,
        pairs=built["pairs"],
        evidence_rows=built["ordered_rows"],
        provenance=[("raw_evidence_record_count", 2), ("nonredundant_pair_count", 1)],
    )
    wb = load_workbook(path, read_only=True)
    assert (wb["Evidence_Records"].max_row or 1) - 1 == 2
    assert (wb["Nonredundant_Interactions"].max_row or 1) - 1 == 1
    wb.close()
    # Polished evidence would be summary + 1 pair + workbook (+ optional figure) ~= 3
    # and must NOT equal raw 2 for acceptance confusion — documented by this split.


def test_canonical_pair_unordered() -> None:
    assert canonical_pair_key(5, 2) == (2, 5)
    assert canonical_pair_key(2, 5) == (2, 5)
