"""Section 5b BioGRID Network Viewer capture / acceptance tests."""

from __future__ import annotations

import base64
import io
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest
from PIL import Image

from gene_dossier.section_5b import (
    EXPORT_PNG_SELECTOR,
    INTERACTION_BODY_SELECTOR,
    STATUS_NO_INTERACTIONS,
    STATUS_NOT_ATTEMPTED,
    STATUS_NOT_ATTEMPTED_NO_NETWORK,
    STATUS_SUCCESS,
    STATUS_UNAVAILABLE,
    _single_canvas_to_data_url,
    allowed_biogrid_hostname,
    capture_biogrid_network_figure,
    evaluate_section_5b_complete,
    validate_network_png,
)
from gene_dossier.section_5b_sources import sha256_bytes


def _png_bytes(*, width: int = 500, height: int = 400, color=(20, 40, 200), second=None) -> bytes:
    img = Image.new("RGB", (width, height), color)
    if second is not None:
        for x in range(0, width // 2):
            for y in range(0, height // 2):
                img.putpixel((x, y), second)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def test_does_not_require_window_cy_or_graph_in_source() -> None:
    src = Path("src/gene_dossier/section_5b.py").read_text(encoding="utf-8")
    assert "window.cy" not in src
    assert "window.graph" not in src
    assert INTERACTION_BODY_SELECTOR in src
    assert EXPORT_PNG_SELECTOR in src
    assert "interaction_body_screenshot" in src
    assert "biogrid_export_png_button" in src


def test_allowed_biogrid_hostname_exact() -> None:
    assert allowed_biogrid_hostname("https://thebiogrid.org/112599/network/homo-sapiens/srebf2.html")
    assert allowed_biogrid_hostname("https://www.thebiogrid.org/112599/network/homo-sapiens/srebf2.html")
    assert not allowed_biogrid_hostname("https://evil-thebiogrid.org/x")
    assert not allowed_biogrid_hostname("https://example.com/thebiogrid.org/x")
    assert not allowed_biogrid_hostname("https://thebiogrid.org.evil.com/x")


def test_sparse_valid_png_accepted() -> None:
    # Compresses well but is large enough and non-uniform.
    raw = _png_bytes(width=420, height=320, color=(255, 255, 255), second=(0, 0, 0))
    assert len(raw) < 20_000
    check = validate_network_png(raw)
    assert check["ok"] is True
    assert check["width"] == 420
    assert check["height"] == 320
    assert check["sha256"] == sha256_bytes(raw)


def test_uniform_blank_png_rejected() -> None:
    raw = _png_bytes(width=500, height=400, color=(255, 255, 255))
    check = validate_network_png(raw)
    assert check["ok"] is False
    assert check["reason"] == "blank_or_uniform"


def test_tiny_png_rejected() -> None:
    raw = _png_bytes(width=100, height=100, color=(1, 2, 3), second=(4, 5, 6))
    check = validate_network_png(raw)
    assert check["ok"] is False
    assert check["reason"] == "dimensions"


def _make_fake_page(*, metrics: dict[str, Any], export_visible: bool = True) -> MagicMock:
    page = MagicMock()
    page.url = "https://thebiogrid.org/112599/network/homo-sapiens/srebf2.html"
    page.title.return_value = "SREBF2 | BioGRID"
    page.content.return_value = "BioGRID ID 112599 SREBF2 network"
    page.goto.return_value = None
    page.wait_for_timeout.return_value = None
    page.wait_for_function.return_value = None

    body = MagicMock()
    body.wait_for.return_value = None
    body.is_visible.return_value = True
    body.bounding_box.return_value = {"x": 0, "y": 0, "width": 900, "height": 700}
    body.scroll_into_view_if_needed.return_value = None
    body.screenshot.return_value = _png_bytes(color=(10, 20, 30), second=(200, 100, 50))

    export = MagicMock()
    export.count.return_value = 1 if export_visible else 0
    export.first.is_visible.return_value = export_visible

    def _locator(sel: str) -> MagicMock:
        if sel == INTERACTION_BODY_SELECTOR:
            return body
        if sel == EXPORT_PNG_SELECTOR:
            return export
        return MagicMock()

    page.locator.side_effect = _locator

    def _evaluate(script, *args, **kwargs):
        text = script if isinstance(script, str) else ""
        if "interaction_body_found" in text or "largest_canvas_width" in text:
            return metrics
        if "__biogridExportHookInstalled" in text or "__biogridPngCapture" in text and "async" not in text:
            return None
        if "async ()" in text or "blobUrlToDataUrl" in text:
            return None
        if "meaningful.length" in text or "toDataURL" in text:
            return {"error": "not_single_meaningful_canvas", "count": metrics.get("meaningful_canvas_count", 0)}
        return metrics

    page.evaluate.side_effect = _evaluate
    return page


def test_capture_prefers_download_export(monkeypatch: pytest.MonkeyPatch) -> None:
    png = _png_bytes(color=(11, 22, 33), second=(44, 55, 66))
    metrics = {
        "interaction_body_found": True,
        "interaction_body_visible": True,
        "canvas_count": 3,
        "visible_canvas_count": 3,
        "meaningful_canvas_count": 3,
        "largest_canvas_width": 800,
        "largest_canvas_height": 600,
    }
    page = _make_fake_page(metrics=metrics)

    class _Download:
        suggested_filename = "network.png"

        def path(self) -> str:
            return ""

        def save_as(self, dest: str) -> None:
            Path(dest).write_bytes(png)

    class _DownloadCtx:
        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        @property
        def value(self):
            return _Download()

    page.expect_download.return_value = _DownloadCtx()

    browser = MagicMock()
    browser.new_page.return_value = page
    pw = MagicMock()
    pw.chromium.launch.return_value = browser

    class _PWCtx:
        def __enter__(self):
            return pw

        def __exit__(self, *args):
            return False

    monkeypatch.setattr(
        "playwright.sync_api.sync_playwright",
        lambda: _PWCtx(),
    )

    raw, audit = capture_biogrid_network_figure(biogrid_id=112599, gene_symbol="SREBF2")
    assert raw == png
    assert audit["capture_method"] == "biogrid_export_png_button"
    assert audit["requires_window_cy"] is False
    assert audit["canvas_count"] == 3
    assert audit["interaction_body_found"] is True


def test_capture_client_side_data_url_export(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    png = _png_bytes(color=(9, 8, 7), second=(1, 2, 3))
    data_url = "data:image/png;base64," + base64.b64encode(png).decode("ascii")
    metrics = {
        "interaction_body_found": True,
        "interaction_body_visible": True,
        "canvas_count": 2,
        "visible_canvas_count": 2,
        "meaningful_canvas_count": 2,
        "largest_canvas_width": 700,
        "largest_canvas_height": 500,
    }
    page = _make_fake_page(metrics=metrics)

    class _DownloadCtx:
        def __enter__(self):
            raise TimeoutError("no download event")

        def __exit__(self, *args):
            return False

    page.expect_download.return_value = _DownloadCtx()

    def _evaluate(script, *args, **kwargs):
        text = script if isinstance(script, str) else ""
        if "largest_canvas_width" in text:
            return metrics
        if "blobUrlToDataUrl" in text or "async ()" in text:
            return {"kind": "data_url", "href": data_url, "download": "export.png"}
        return None

    page.evaluate.side_effect = _evaluate

    browser = MagicMock()
    browser.new_page.return_value = page
    pw = MagicMock()
    pw.chromium.launch.return_value = browser

    class _PWCtx:
        def __enter__(self):
            return pw

        def __exit__(self, *args):
            return False

    monkeypatch.setattr("playwright.sync_api.sync_playwright", lambda: _PWCtx())
    raw, audit = capture_biogrid_network_figure(
        biogrid_id=112599, gene_symbol="SREBF2", debug_dir=tmp_path
    )
    assert raw == png
    assert audit["capture_method"] == "biogrid_export_png_button"
    assert audit.get("export_path") == "client_side_data_or_blob"


def test_capture_interaction_body_screenshot_fallback(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    metrics = {
        "interaction_body_found": True,
        "interaction_body_visible": True,
        "canvas_count": 3,
        "visible_canvas_count": 3,
        "meaningful_canvas_count": 3,
        "largest_canvas_width": 900,
        "largest_canvas_height": 700,
    }
    page = _make_fake_page(metrics=metrics, export_visible=False)

    browser = MagicMock()
    browser.new_page.return_value = page
    pw = MagicMock()
    pw.chromium.launch.return_value = browser

    class _PWCtx:
        def __enter__(self):
            return pw

        def __exit__(self, *args):
            return False

    monkeypatch.setattr("playwright.sync_api.sync_playwright", lambda: _PWCtx())
    raw, audit = capture_biogrid_network_figure(biogrid_id=112599, gene_symbol="SREBF2")
    assert raw is not None
    assert audit["capture_method"] == "interaction_body_screenshot"
    assert audit["meaningful_canvas_count"] == 3
    assert audit["capture_method"] != "single_canvas_toDataURL"
    assert "window.cy" not in str(audit)


def test_multi_canvas_skips_blind_todataurl() -> None:
    page = MagicMock()
    page.evaluate.return_value = {"error": "not_single_meaningful_canvas", "count": 3}
    raw, meta = _single_canvas_to_data_url(page)
    assert raw is None
    assert meta.get("error") == "not_single_meaningful_canvas"


def test_single_canvas_fallback_may_be_used(monkeypatch: pytest.MonkeyPatch) -> None:
    png = _png_bytes(color=(50, 60, 70), second=(80, 90, 100))
    data_url = "data:image/png;base64," + base64.b64encode(png).decode("ascii")
    metrics = {
        "interaction_body_found": True,
        "interaction_body_visible": True,
        "canvas_count": 1,
        "visible_canvas_count": 1,
        "meaningful_canvas_count": 1,
        "largest_canvas_width": 640,
        "largest_canvas_height": 480,
    }
    page = _make_fake_page(metrics=metrics, export_visible=False)
    # Force screenshot validation failure by returning a blank image.
    blank = _png_bytes(width=500, height=400, color=(255, 255, 255))
    page.locator(INTERACTION_BODY_SELECTOR).screenshot.return_value = blank

    def _evaluate(script, *args, **kwargs):
        text = script if isinstance(script, str) else ""
        if "largest_canvas_width" in text:
            return metrics
        if "toDataURL" in text:
            return {"data_url": data_url}
        return None

    page.evaluate.side_effect = _evaluate

    browser = MagicMock()
    browser.new_page.return_value = page
    pw = MagicMock()
    pw.chromium.launch.return_value = browser

    class _PWCtx:
        def __enter__(self):
            return pw

        def __exit__(self, *args):
            return False

    monkeypatch.setattr("playwright.sync_api.sync_playwright", lambda: _PWCtx())
    raw, audit = capture_biogrid_network_figure(biogrid_id=112599, gene_symbol="SREBF2")
    assert raw == png
    assert audit["capture_method"] == "single_canvas_toDataURL"


def test_unexpected_host_rejected(monkeypatch: pytest.MonkeyPatch) -> None:
    page = _make_fake_page(
        metrics={
            "interaction_body_found": True,
            "interaction_body_visible": True,
            "canvas_count": 1,
            "visible_canvas_count": 1,
            "meaningful_canvas_count": 1,
            "largest_canvas_width": 500,
            "largest_canvas_height": 400,
        }
    )
    page.url = "https://evil.example.com/network"

    browser = MagicMock()
    browser.new_page.return_value = page
    pw = MagicMock()
    pw.chromium.launch.return_value = browser

    class _PWCtx:
        def __enter__(self):
            return pw

        def __exit__(self, *args):
            return False

    monkeypatch.setattr("playwright.sync_api.sync_playwright", lambda: _PWCtx())
    raw, audit = capture_biogrid_network_figure(biogrid_id=1, gene_symbol="X")
    assert raw is None
    assert audit["failure_stage"] == "identity_validation_failed"


def _acceptance_status(
    *,
    scientific: str = STATUS_SUCCESS,
    visual: str = STATUS_SUCCESS,
    presentation: str = STATUS_SUCCESS,
    figure_requested: bool = True,
    attempt_dir: Path,
    gene: str = "SREBF2",
    nr: int = 325,
    raw: int = 378,
    fig_sha: str | None = None,
    capture: dict | None = None,
) -> dict[str, Any]:
    from openpyxl import Workbook

    xlsx = attempt_dir / "supplementary" / f"{gene}_BIOGRID.xlsx"
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Nonredundant_Interactions"
    ws.append(["h"])
    for _ in range(nr):
        ws.append(["r"])
    ws2 = wb.create_sheet("Evidence_Records")
    ws2.append(["h"])
    for _ in range(raw):
        ws2.append(["r"])
    wb.create_sheet("Provenance").append(["k", "v"])
    wb.save(xlsx)

    (attempt_dir / "evidence_records.json").write_text(
        "[" + ",".join(["{}"] * raw) + "]", encoding="utf-8"
    )
    fig_path = attempt_dir / "figures" / f"{gene}_BIOGRID_network.png"
    if fig_sha is not None:
        fig_path.parent.mkdir(parents=True, exist_ok=True)
        # Write matching bytes if sha provided via helper outside
    return {
        "summary": {
            "official_symbol": gene,
            "gene_symbol": gene,
            "scientific_status": scientific,
            "visual_status": visual,
            "presentation_status": presentation,
            "network_figure_requested": figure_requested,
            "raw_evidence_record_count": raw,
            "nonredundant_pair_count": nr,
            "supplementary_xlsx": f"{gene}_BIOGRID.xlsx",
            "supplementary_xlsx_sha256": __import__(
                "gene_dossier.section_5b_sources", fromlist=["sha256_file"]
            ).sha256_file(xlsx),
            "network_figure_relative_path": (
                f"figures/{gene}_BIOGRID_network.png" if fig_sha else None
            ),
            "network_figure_sha256": fig_sha,
            "figure_capture": capture or {},
        },
        "rendering_status": {
            "scientific_status": scientific,
            "visual_status": visual,
            "presentation_status": presentation,
        },
        "audit": {"network_figure_requested": figure_requested, "figure_capture": capture or {}},
    }


def test_requested_figure_unavailable_fails_acceptance(tmp_path: Path) -> None:
    pdf = tmp_path / "out.pdf"
    pdf.write_bytes(b"%PDF")
    status = _acceptance_status(
        visual=STATUS_UNAVAILABLE,
        figure_requested=True,
        attempt_dir=tmp_path,
        fig_sha=None,
    )
    html = (
        "5. Protein-protein interaction (PPI) partners"
        "<h3>b. BioGRID</h3>"
        "SREBF2 has 325 unique interactions."
    )
    result = evaluate_section_5b_complete(
        status=status, html_text=html, pdf_path=pdf, attempt_dir=tmp_path
    )
    assert result["complete"] is False
    assert result["checks"]["visual_status_success"]["passed"] is False


def test_optional_disabled_figure_acceptance(tmp_path: Path) -> None:
    pdf = tmp_path / "out.pdf"
    pdf.write_bytes(b"%PDF")
    status = _acceptance_status(
        visual=STATUS_NOT_ATTEMPTED,
        figure_requested=False,
        attempt_dir=tmp_path,
    )
    html = (
        "5. Protein-protein interaction (PPI) partners"
        "<h3>b. BioGRID</h3>"
        "SREBF2 has 325 unique interactions."
    )
    result = evaluate_section_5b_complete(
        status=status, html_text=html, pdf_path=pdf, attempt_dir=tmp_path
    )
    assert result["checks"]["visual_status_not_attempted_optional"]["passed"] is True


def test_no_interactions_allows_not_attempted_no_network(tmp_path: Path) -> None:
    pdf = tmp_path / "out.pdf"
    pdf.write_bytes(b"%PDF")
    status = _acceptance_status(
        scientific=STATUS_NO_INTERACTIONS,
        visual=STATUS_NOT_ATTEMPTED_NO_NETWORK,
        presentation="partial",
        figure_requested=True,
        attempt_dir=tmp_path,
        nr=0,
        raw=0,
    )
    html = "5. Protein-protein interaction (PPI) partners<h3>b. BioGRID</h3>"
    result = evaluate_section_5b_complete(
        status=status, html_text=html, pdf_path=pdf, attempt_dir=tmp_path
    )
    assert result["checks"]["visual_status_not_attempted_no_network"]["passed"] is True
    assert result["complete"] is False  # not a scientific success inventory
