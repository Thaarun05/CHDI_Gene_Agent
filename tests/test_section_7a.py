"""Offline tests for Section 7a chemical tools (opt-in)."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import patch

from gene_dossier.models import (
    AssertionType,
    EvidenceGrade,
    EvidenceRecord,
    SourceType,
    ToolResult,
)
from gene_dossier.report_schema import ReportSlot, resolve_report_slot
from gene_dossier.section_7a import (
    PARSER_VERSION,
    POLISHED_PUBCHEM_CAP,
    Section7aConfig,
    build_literature_effect_prose,
    build_pubmed_perturbation_term,
    classify_chembl_activity,
    classify_evidence_span,
    classify_tool_eligibility,
    derive_literature_status,
    evaluate_section_7a_complete,
    ncats_display_text,
    node_generate_section_7a_derived_artifacts,
    rank_literature_candidates,
    resolve_aliases_from_state,
    select_polished_pubchem_rows,
    write_chembl_workbook,
)
from gene_dossier.section_7a_sources import paths_for, sha256_file, write_json_atomic
from gene_dossier.section_bundle import (
    DEFAULT_SECTION_BUNDLE_KEYS,
    SECTION_SOURCE_DEPENDENCIES,
    SUPPORTED_SECTION_BUNDLE_KEYS,
)
from gene_dossier.source_ids import make_source_id
from gene_dossier.tools import chembl as chembl_client
from gene_dossier.tools import pubchem as pubchem_client


def test_supported_opt_in_not_default():
    assert "7a" in SUPPORTED_SECTION_BUNDLE_KEYS
    assert SUPPORTED_SECTION_BUNDLE_KEYS[-1] == "7a"
    assert "7a" not in DEFAULT_SECTION_BUNDLE_KEYS
    assert SECTION_SOURCE_DEPENDENCIES["7a"] == set()
    assert DEFAULT_SECTION_BUNDLE_KEYS == (
        "1a",
        "1b",
        "1c",
        "1d",
        "1e",
        "2a",
        "2b",
        "2c",
        "3a",
        "4a",
    )


def test_chemical_tool_routes_to_slot_7a():
    rec = EvidenceRecord(
        source_id=make_source_id("ChEMBL", "SREBF2", AssertionType.chemical_tool, "x"),
        dossier_run_id="run",
        gene_symbol="SREBF2",
        section="Chemical tools",
        subsection="Queried",
        source_name="ChEMBL",
        source_type=SourceType.chemical_database,
        assertion_type=AssertionType.chemical_tool,
        fact_type="section_7a_summary",
        evidence_grade=EvidenceGrade.C,
        value={},
        display_text="x",
    )
    assert resolve_report_slot(rec) == ReportSlot("7", "a")


def test_no_production_hardcoded_srebf2_aliases():
    src = Path("src/gene_dossier/section_7a.py").read_text(encoding="utf-8")
    assert "SREBP2" not in src
    assert "SREBP-2" not in src
    assert "GENE_SPECIFIC" not in src
    # No gene-specific production alias tables.
    assert "ALIAS_MAP" not in src
    assert "HARDCODED_ALIASES" not in src
    # No reagent stoplist hardcodes.
    assert "HEPES" not in src
    assert '"Tris"' not in src


def test_alias_resolution_from_identity_evidence_only():
    rec = EvidenceRecord(
        source_id=make_source_id("NCBI Gene", "SREBF2", AssertionType.gene_identity, "6721"),
        dossier_run_id="run",
        gene_symbol="SREBF2",
        section="General",
        source_name="NCBI Gene",
        source_type=SourceType.curated_database,
        assertion_type=AssertionType.gene_identity,
        fact_type="entrez_gene_id",
        evidence_grade=EvidenceGrade.C,
        taxon_id=9606,
        value={
            "entrez_gene_id": "6721",
            "otheraliases": "SREBP2; SREBP-2",
            "taxon_id": 9606,
        },
        display_text="id",
    )
    payload = resolve_aliases_from_state(
        {
            "gene_symbol": "SREBF2",
            "gene_ids": {"official_symbol": "SREBF2", "entrez_gene_id": 6721},
            "evidence_records": [rec],
        }
    )
    assert payload["alias_resolution_status"] == "resolved"
    aliases = {a["alias"] for a in payload["aliases"]}
    assert "SREBP2" in aliases
    assert "SREBP-2" in aliases
    for row in payload["aliases"]:
        assert row["authoritative_source"]
        assert row["source_identifier"]
        assert row["normalized_canonical_symbol"] == "SREBF2"


def test_alias_resolution_canonical_only_when_missing():
    payload = resolve_aliases_from_state(
        {"gene_symbol": "SREBF2", "gene_ids": {"official_symbol": "SREBF2"}, "evidence_records": []}
    )
    assert payload["alias_resolution_status"] == "canonical_only"
    assert payload["aliases"] == []
    assert payload["limitation"] == "authoritative_aliases_unavailable"
    term = build_pubmed_perturbation_term(payload)
    assert "SREBF2[Title/Abstract]" in term
    assert "SREBP2" not in term


def test_negative_correlate_not_direct_inhibitor_in_classifier():
    result = classify_evidence_span(
        title="Cholesterol correlates with SREBF2",
        abstract="Cholesterol and SREBF2 are negatively correlated in expression datasets.",
        chemical_name="Cholesterol",
        gene_symbol="SREBF2",
    )
    assert result["evidence_class"] != "direct_target_evidence"
    assert result["tool_eligibility"] == "endogenous_or_contextual_chemical"


def test_endogenous_does_not_outrank_explicit_tool():
    endogenous = {
        "chemical_name": "Cholesterol",
        "evidence_class": "literature_interaction",
        "tool_eligibility": "endogenous_or_contextual_chemical",
        "support_count": 999,
        "pmids": ["1"],
    }
    tool = {
        "chemical_name": "Fatostatin",
        "evidence_class": "literature_negative_effect",
        "tool_eligibility": "explicit_chemical_tool",
        "support_count": 2,
        "pmids": ["2"],
    }
    ranked = rank_literature_candidates([endogenous, tool])
    assert ranked[0]["chemical_name"] == "Fatostatin"


def test_indirect_pathway_cdh10_style_span():
    abstract = (
        "UNC0642 inhibits G9a/EHMT2 histone methyltransferase activity. "
        "Pharmacologic G9a inhibition rescued CDH10 expression in hypoxia."
    )
    result = classify_evidence_span(
        title="Epigenetic repression of CDH10",
        abstract=abstract,
        chemical_name="UNC0642",
        gene_symbol="CDH10",
    )
    assert result["evidence_class"] == "indirect_pathway_effect"
    assert result["supporting_span_text"]
    assert result["supporting_sentence_indices"]
    assert result["evidence_scope"] == "local_window"
    assert classify_tool_eligibility(abstract, "UNC0642") in {
        "explicit_chemical_tool",
        "perturbational_chemical",
    }
    prose = build_literature_effect_prose({**result, "chemical_name": "UNC0642"})
    assert "UNC0642" in prose
    assert "indirect pathway effect" in prose
    assert "G9a-mediated repression of CDH10" not in prose  # not title substitute


def test_hepes_does_not_inherit_unc0642_evidence():
    """Document-wide co-occurrence must not promote methods reagents."""
    abstract = (
        "Cells were washed in HEPES buffer and Tris-HCl. "
        "UNC0642 inhibits G9a/EHMT2. "
        "Pharmacologic G9a inhibition rescued CDH10 expression in hypoxia."
    )
    hepes = classify_evidence_span(
        title="G9a-mediated repression of CDH10",
        abstract=abstract,
        chemical_name="HEPES",
        gene_symbol="CDH10",
    )
    tris = classify_evidence_span(
        title="G9a-mediated repression of CDH10",
        abstract=abstract,
        chemical_name="Tris",
        gene_symbol="CDH10",
    )
    unc = classify_evidence_span(
        title="G9a-mediated repression of CDH10",
        abstract=abstract,
        chemical_name="UNC0642",
        gene_symbol="CDH10",
    )
    assert unc["evidence_class"] == "indirect_pathway_effect"
    assert hepes["evidence_class"] == "insufficient_effect_detail"
    assert tris["evidence_class"] == "insufficient_effect_detail"
    assert not hepes.get("supporting_span_text")
    assert not tris.get("supporting_span_text")


def test_sterols_not_promoted_by_remote_treatment_language():
    text = (
        "Sterols regulate SREBF2 transcription in hepatocytes. "
        "Cells were treated with fatostatin, a small molecule inhibitor."
    )
    assert (
        classify_tool_eligibility(text, "Sterols")
        == "endogenous_or_contextual_chemical"
    )
    # "treatment with" a different agent in the same sentence is not enough.
    depletion = (
        "Depletion of sterols by treatment with a bile acid-binding resin "
        "(colestipol) and a cholesterol synthesis inhibitor (mevinolin) led "
        "to increased nuclear SREBF2."
    )
    assert (
        classify_tool_eligibility(depletion, "Sterols")
        == "endogenous_or_contextual_chemical"
    )
    # Local experimental use of sterols remains eligible.
    local = (
        "Cells were treated with sterols as a small molecule probe of SREBF2. "
        "Sterols reduced SREBF2 expression."
    )
    assert classify_tool_eligibility(local, "Sterols") in {
        "explicit_chemical_tool",
        "perturbational_chemical",
    }


def test_fulltext_interaction_not_display_eligible_without_perturbation():
    from gene_dossier.section_7a import _display_eligible

    entry = {
        "chemical_name": "Melatonin",
        "evidence_class": "literature_interaction",
        "tool_eligibility": "explicit_chemical_tool",
        "supporting_span_text": "CDH10 and Melatonin appear in the same table.",
        "evidence_scope": "local_window",
        "grounding_source": "pubtator_fulltext_local_window",
    }
    assert _display_eligible(entry) is False
    entry["evidence_class"] = "indirect_pathway_effect"
    assert _display_eligible(entry) is True
    assert _display_eligible({**entry, "chemical_name": "BF"}) is False



def test_chembl_direct_requires_relationship_metadata():
    activity = {"standard_type": "IC50", "standard_value": 10}
    assert classify_chembl_activity(activity, None) == "target_linked_activity"
    assert classify_chembl_activity(activity, {"relationship_type": "H"}) == (
        "target_linked_activity"
    )
    assert classify_chembl_activity(activity, {"relationship_type": "D"}) == (
        "direct_target_evidence"
    )


def test_chembl_resolve_authoritative_srebf2_fixture_shape():
    targets = [
        {
            "target_chembl_id": "CHEMBL1795166",
            "organism": "Homo sapiens",
            "target_type": "SINGLE PROTEIN",
            "target_components": [{"accession": "Q12772"}],
        }
    ]
    tid, method, _detail = chembl_client.resolve_authoritative_target(
        targets, uniprot_accession="Q12772", gene_symbol="SREBF2"
    )
    assert tid == "CHEMBL1795166"
    assert method == "uniprot_single_protein"


def test_chembl_cdh10_no_authoritative_target():
    targets = [
        {
            "target_chembl_id": "CHEMBL999",
            "organism": "Homo sapiens",
            "target_type": "SINGLE PROTEIN",
            "target_components": [{"accession": "P99999"}],
        }
    ]
    tid, method, _ = chembl_client.resolve_authoritative_target(
        targets, uniprot_accession="Q9Y6N8", gene_symbol="CDH10"
    )
    assert tid is None
    assert method == "not_found"


def test_pubchem_broad_screen_excluded():
    targets_payload = {
        "InformationList": {
            "Information": [
                {
                    "AID": 1,
                    "GeneID": list(range(1, 20)),
                    "GeneSymbol": [f"G{i}" for i in range(1, 20)],
                }
            ]
        }
    }
    focused, reason = pubchem_client.classify_focused_assay(
        None,
        targets_payload,
        uniprot="Q12772",
        entrez=6721,
        symbol="SREBF2",
    )
    assert focused is False
    assert reason == "broad_screen_membership_only"


def test_pubchem_polished_representatives_capped():
    rows = []
    for i in range(1, 12):
        rows.append(
            {
                "aid": i,
                "focused": True,
                "reason": "direct_target",
                "pmid": "100" if i <= 4 else str(200 + i),
            }
        )
    polished = select_polished_pubchem_rows(rows)
    assert len(polished) <= POLISHED_PUBCHEM_CAP
    assert all(r["aid"] in {row["aid"] for row in rows} for r in polished)
    # One representative for shared PMID group.
    assert sum(1 for r in polished if r.get("pmid") == "100") == 1


def test_ncats_display_names_candidate():
    text = ncats_display_text(
        [
            {
                "name": "NICODICOSAPENT",
                "evidence_class": "facet_target_match_unconfirmed",
            }
        ]
    )
    assert "NICODICOSAPENT" in text
    assert "not independently confirmed" in text
    assert ncats_display_text([]) == "NCATS Inxight: Drugs – No results"


def test_literature_status_composite():
    assert derive_literature_status("no_results", "success") == "success"
    assert (
        derive_literature_status("source_error", "success")
        == "success_with_source_limitations"
    )
    assert derive_literature_status("no_results", "no_results") == "no_results"


def test_audit_sha_external_only(tmp_path: Path):
    paths = paths_for(tmp_path)
    attempt = paths.new_gene_attempt("SREBF2", run_id="testrun")
    audit = {"parser_version": PARSER_VERSION, "gene_symbol": "SREBF2", "sources": {}}
    audit_path = attempt / "section_7a_audit.json"
    write_json_atomic(audit_path, audit)
    sha = sha256_file(audit_path)
    write_json_atomic(
        attempt / "manifest.json",
        {"section_7a_audit_sha256": sha, "section_key": "7a"},
    )
    payload = json.loads(audit_path.read_text(encoding="utf-8"))
    assert "sha256" not in payload
    assert "section_7a_audit_sha256" not in payload
    assert sha256_file(audit_path) == sha


def test_workbook_hashed_after_close(tmp_path: Path):
    path = tmp_path / "SREBF2_Chembl_Inhibitor.xlsx"
    write_chembl_workbook(
        path,
        gene="SREBF2",
        target={"target_chembl_id": "CHEMBL1795166"},
        activities=[
            {
                "molecule_chembl_id": "CHEMBL1",
                "canonical_smiles": "C",
                "standard_type": "IC50",
                "standard_relation": "=",
                "standard_value": 1,
                "standard_units": "nM",
                "pchembl_value": 9,
                "assay_chembl_id": "CHEMBLA",
                "document_chembl_id": "CHEMBLD",
                "evidence_class": "target_linked_activity",
                "relationship_type": "H",
                "confidence_score": 8,
            }
        ],
        assays=[
            {
                "assay_chembl_id": "CHEMBLA",
                "description": "Binding assay",
                "assay_type": "B",
                "target_chembl_id": "CHEMBL1795166",
                "relationship_type": "H",
                "confidence_score": 8,
                "confidence_description": "Homologous single protein target",
                "document_chembl_id": "CHEMBLD",
            }
        ],
        provenance={"parser_version": PARSER_VERSION},
    )
    sha = sha256_file(path)
    assert len(sha) == 64
    assert sha256_file(path) == sha
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "Assays" in wb.sheetnames
    assert wb["Assays"]["A1"].value == "assay_chembl_id"
    assert wb["Assays"]["B2"].value == "Binding assay"
    assert wb["Assays"]["E2"].value == "H"


def test_evaluate_accepts_external_sha(tmp_path: Path):
    attempt = tmp_path / "attempt"
    attempt.mkdir()
    audit = {"parser_version": PARSER_VERSION, "gene_symbol": "SREBF2", "sources": {}}
    audit_path = attempt / "section_7a_audit.json"
    write_json_atomic(audit_path, audit)
    sha = sha256_file(audit_path)
    html = (
        '<section id="section-7"></section>'
        '<section class="subsection-7a"></section>'
        "ChEMBL DrugBank PubMed PubChem NCATS"
    )
    pdf = tmp_path / "out.pdf"
    pdf.write_bytes(b"%PDF-1.4")
    status = {
        "summary": {
            "gene_symbol": "SREBF2",
            "section_7a_audit_path": str(audit_path),
            "section_7a_audit_sha256": sha,
            "source_blocks": {
                "chembl": {
                    "source_status": "no_authoritative_target",
                    "activity_count": 0,
                    "activity_page_api_run_ids": [],
                },
                "drugbank": {"source_status": "unavailable_not_configured"},
                "pubmed": {"source_status": "no_results"},
                "pubchem": {
                    "source_status": "no_results",
                    "focused": [],
                    "polished_focused": [],
                    "excluded": [],
                },
                "ncats": {"source_status": "no_results", "candidates": []},
            },
            "polished_literature": [],
        }
    }
    result = evaluate_section_7a_complete(
        status=status,
        html_text=html,
        pdf_path=pdf,
        attempt_dir=attempt,
    )
    assert result["checks"]["audit_sha_matches"]["passed"] is True
    assert result["checks"]["audit_has_no_self_sha"]["passed"] is True
    assert result["checks"]["drugbank_unavailable_not_no_results"]["passed"] is True
    assert result["checks"]["bundle_7a_supported"]["passed"] is True
    assert result["checks"]["bundle_7a_not_default"]["passed"] is True
    assert result["complete"] is True


def test_chembl_pagination_persists_separate_api_runs(tmp_path: Path):
    """Two activity pages must yield two HTTP ApiRuns; aggregation is local only."""
    page_calls = {"n": 0}

    def _page_result(offset: int, batch: list[dict], *, next_url: str | None):
        return ToolResult(
            source_name="ChEMBL",
            endpoint_name="activities_by_target",
            success=True,
            gene_symbol="SREBF2",
            request_url=f"https://www.ebi.ac.uk/chembl/api/data/activity.json?offset={offset}",
            request_params={"target_chembl_id": "CHEMBL1795166", "offset": str(offset)},
            status_code=200,
            data={
                "activities": batch,
                "page_meta": {
                    "total_count": 2,
                    "offset": offset,
                    "limit": 1,
                    "next": next_url,
                },
            },
        )

    def fake_activities_by_target(tid, gene_symbol="", limit=1000, offset=0, settings=None):
        page_calls["n"] += 1
        if offset == 0:
            return _page_result(
                0,
                [
                    {
                        "molecule_chembl_id": "CHEMBL1",
                        "standard_type": "IC50",
                        "assay_chembl_id": "CHEMBLA1",
                        "target_chembl_id": tid,
                    }
                ],
                next_url="next",
            )
        return _page_result(
            1,
            [
                {
                    "molecule_chembl_id": "CHEMBL2",
                    "standard_type": "Ki",
                    "assay_chembl_id": "CHEMBLA2",
                    "target_chembl_id": tid,
                }
            ],
            next_url=None,
        )

    def fake_assays_by_target(tid, gene_symbol="", limit=100, offset=0, settings=None):
        return ToolResult(
            source_name="ChEMBL",
            endpoint_name="assays_by_target",
            success=True,
            gene_symbol="SREBF2",
            request_url="https://www.ebi.ac.uk/chembl/api/data/assay.json",
            request_params={"target_chembl_id": tid, "offset": str(offset)},
            status_code=200,
            data={
                "assays": [
                    {
                        "assay_chembl_id": "CHEMBLA1",
                        "description": "assay 1",
                        "relationship_type": "D",
                        "confidence_score": 9,
                        "target_chembl_id": tid,
                    },
                    {
                        "assay_chembl_id": "CHEMBLA2",
                        "description": "assay 2",
                        "relationship_type": "H",
                        "confidence_score": 8,
                        "target_chembl_id": tid,
                    },
                ],
                "page_meta": {"total_count": 2, "next": None},
            },
        )

    def fake_target_search(term, settings=None):
        return ToolResult(
            source_name="ChEMBL",
            endpoint_name="target_search",
            success=True,
            gene_symbol="SREBF2",
            request_url="https://www.ebi.ac.uk/chembl/api/data/target/search.json",
            request_params={"q": term},
            status_code=200,
            data={
                "targets": [
                    {
                        "target_chembl_id": "CHEMBL1795166",
                        "organism": "Homo sapiens",
                        "target_type": "SINGLE PROTEIN",
                        "target_components": [{"accession": "Q12772"}],
                    }
                ]
            },
        )

    with (
        patch("gene_dossier.section_7a.chembl_client.target_search", side_effect=fake_target_search),
        patch(
            "gene_dossier.section_7a.chembl_client.activities_by_target",
            side_effect=fake_activities_by_target,
        ),
        patch(
            "gene_dossier.section_7a.chembl_client.assays_by_target",
            side_effect=fake_assays_by_target,
        ),
        patch("gene_dossier.section_7a.drugbank_client.fetch_status") as db,
        patch("gene_dossier.section_7a.pubchem_client.aids_by_geneid") as aids,
        patch("gene_dossier.section_7a.ncats_client.search_substances") as ncats,
        patch("gene_dossier.section_7a.pubtator_client.entity_autocomplete") as ac,
        patch("gene_dossier.section_7a.pubtator_client.relations") as rel,
        patch("gene_dossier.section_7a.pubmed_client.esearch_custom") as es,
    ):
        db.return_value = ToolResult(
            source_name="DrugBank",
            endpoint_name="status",
            success=False,
            gene_symbol="SREBF2",
            request_url="",
            request_params={},
            error_type="unavailable_not_configured",
        )
        aids.return_value = ToolResult(
            source_name="PubChem",
            endpoint_name="aids_by_geneid",
            success=True,
            gene_symbol="SREBF2",
            request_url="",
            request_params={},
            status_code=200,
            data={"IdentifierList": {"AID": []}},
        )
        ncats.return_value = ToolResult(
            source_name="NCATS",
            endpoint_name="search",
            success=True,
            gene_symbol="SREBF2",
            request_url="",
            request_params={},
            status_code=200,
            data={"total": 0, "content": []},
        )
        ac.return_value = ToolResult(
            source_name="PubTator3",
            endpoint_name="entity_autocomplete",
            success=True,
            gene_symbol="SREBF2",
            request_url="",
            request_params={},
            status_code=200,
            data=[{"_id": "@GENE_SREBF2", "db_id": "6721"}],
        )
        rel.return_value = ToolResult(
            source_name="PubTator3",
            endpoint_name="relations",
            success=True,
            gene_symbol="SREBF2",
            request_url="",
            request_params={},
            status_code=200,
            data=[],
        )
        es.return_value = ToolResult(
            source_name="PubMed",
            endpoint_name="esearch_custom",
            success=True,
            gene_symbol="SREBF2",
            request_url="",
            request_params={},
            status_code=200,
            data={"esearchresult": {"idlist": []}},
        )
        out = node_generate_section_7a_derived_artifacts(
            {
                "run_type": "section_bundle",
                "selected_section_keys": ["7a"],
                "gene_symbol": "SREBF2",
                "dossier_run_id": "run-chembl-pages",
                "gene_ids": {
                    "official_symbol": "SREBF2",
                    "entrez_gene_id": 6721,
                    "uniprot_accession": "Q12772",
                },
                "evidence_records": [],
                "api_runs": [],
                "raw_artifacts": [],
                "tool_results": [],
            },
            persist_db=False,
            config=Section7aConfig(output_root=tmp_path),
        )

    chembl = out["section_7a_status"]["summary"]["source_blocks"]["chembl"]
    assert page_calls["n"] == 2
    assert chembl["activity_pages_fetched"] == 2
    assert len(chembl["activity_page_api_run_ids"]) == 2
    assert len(set(chembl["activity_page_api_run_ids"])) == 2
    assert len(chembl["activity_page_raw_artifact_ids"]) == 2
    assert chembl["activity_count"] == 2
    chembl_activity_runs = [
        tr
        for tr in out["tool_results"]
        if getattr(tr, "endpoint_name", None) == "activities_by_target"
    ]
    assert len(chembl_activity_runs) == 2


def test_node_skips_unless_selected(tmp_path: Path):
    state = {
        "run_type": "section_bundle",
        "selected_section_keys": ["6a"],
        "gene_symbol": "SREBF2",
    }
    out = node_generate_section_7a_derived_artifacts(
        state,
        persist_db=False,
        config=Section7aConfig(output_root=tmp_path),
    )
    assert "section_7a_status" not in out
