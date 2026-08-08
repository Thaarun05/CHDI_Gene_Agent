"""Section 5b: Protein-protein interaction (PPI) partners — BioGRID."""

from __future__ import annotations

import base64
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook

from gene_dossier.config import Settings, get_settings
from gene_dossier.source_ids import make_source_id
from gene_dossier.models import (
    AssertionType,
    EvidenceGrade,
    EvidenceRecord,
    SourceType,
)
from gene_dossier.section_5b_sources import (
    MANIFEST_FILENAME,
    accept_gene_report,
    paths_for,
    sha256_bytes,
    sha256_file,
    write_json_atomic,
)
from gene_dossier.tools import biogrid as bg
from gene_dossier.workflow import WorkflowTransientContext

SECTION_PPI = "Protein-protein interaction (PPI) partners"
SUBSECTION_5B = "BioGRID"
PARSER_VERSION = "section_5b_biogrid_v1"

STATUS_SUCCESS = "success"
STATUS_SOURCE_UNAVAILABLE = "source_unavailable"
STATUS_PARTIAL = "partial"
STATUS_FAILED = "failed"
STATUS_NO_INTERACTIONS = "no_interactions"
STATUS_TARGET_MISMATCH = "target_identity_mismatch"
STATUS_UNAVAILABLE = "unavailable"
STATUS_NOT_ATTEMPTED = "not_attempted_optional"
STATUS_NOT_ATTEMPTED_NO_NETWORK = "not_attempted_no_network"

SCIENTIFIC_INTRO = (
    "BioGRID is an interaction repository with data compiled through "
    "comprehensive curation efforts."
)

BIOGRID_HOST = "thebiogrid.org"
ALLOWED_BIOGRID_HOSTS = frozenset({"thebiogrid.org", "www.thebiogrid.org"})
INTERACTION_BODY_SELECTOR = "#interaction-body"
EXPORT_PNG_SELECTOR = '.networkExport[data-export="png"]'
VIEWER_MIN_CANVAS_WIDTH = 300
VIEWER_MIN_CANVAS_HEIGHT = 200
PNG_MIN_WIDTH = 400
PNG_MIN_HEIGHT = 300
VIEWER_STABILIZE_MS = 1_500
EXPORT_DOWNLOAD_TIMEOUT_MS = 20_000


@dataclass
class Section5bConfig:
    attempt_network_figure: bool = True
    tax_id: int = bg.TAX_ID_HUMAN
    max_results: int = bg.DEFAULT_MAX
    output_root: str | Path | None = None


def _blank(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null", "-"}:
        return ""
    return text


def _as_int(value: Any) -> int | None:
    text = _blank(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def canonical_pair_key(id_a: Any, id_b: Any) -> tuple[int, int] | None:
    a = _as_int(id_a)
    b = _as_int(id_b)
    if a is None or b is None:
        return None
    return (a, b) if a <= b else (b, a)


def _stable_sort_key(row: dict[str, Any]) -> tuple[int, int, int]:
    page = _as_int(row.get("_page_start") if "_page_start" in row else row.get("page_start")) or 0
    idx = (
        _as_int(row.get("_within_page_index") if "_within_page_index" in row else row.get("within_page_index"))
        or 0
    )
    iid = _as_int(row.get("BIOGRID_INTERACTION_ID")) or 0
    return (page, idx, iid)


def reconcile_target(
    rows: list[dict[str, Any]],
    *,
    gene_symbol: str,
    expected_entrez_id: int | str | None = None,
) -> tuple[dict[str, Any] | None, str | None, list[dict[str, Any]]]:
    """Return (target, error_code, invariant_violation_rows)."""
    symbol = gene_symbol.strip().upper()
    expected_entrez = _as_int(expected_entrez_id)
    if not rows:
        return None, None, []

    target: dict[str, Any] | None = None
    violations: list[dict[str, Any]] = []

    def _endpoint(row: dict[str, Any], side: str) -> dict[str, Any]:
        return {
            "symbol": _blank(row.get(f"OFFICIAL_SYMBOL_{side}")).upper(),
            "entrez": _as_int(row.get(f"ENTREZ_GENE_{side}")),
            "biogrid_id": _as_int(row.get(f"BIOGRID_ID_{side}")),
            "taxon": _as_int(row.get(f"ORGANISM_{side}")),
        }

    for row in rows:
        a = _endpoint(row, "A")
        b = _endpoint(row, "B")
        candidates = []
        for ep in (a, b):
            if expected_entrez is not None and ep["entrez"] == expected_entrez and ep["taxon"] == 9606:
                candidates.append(ep)
            elif expected_entrez is None and ep["symbol"] == symbol and ep["taxon"] == 9606:
                candidates.append(ep)
        if not candidates:
            # Try symbol match on human taxon only for bootstrap
            for ep in (a, b):
                if ep["symbol"] == symbol and ep["taxon"] == 9606:
                    candidates.append(ep)
        if not candidates:
            violations.append(dict(row))
            continue
        chosen = candidates[0]
        if chosen["biogrid_id"] is None or chosen["entrez"] is None:
            return None, STATUS_TARGET_MISMATCH, violations
        if target is None:
            target = {
                "target_symbol": symbol,
                "target_entrez_id": chosen["entrez"],
                "target_biogrid_id": chosen["biogrid_id"],
                "target_taxon_id": 9606,
            }
            continue
        # After establishment: require ID consistency for any endpoint claiming target
        for ep in (a, b):
            is_target_side = (
                ep["biogrid_id"] == target["target_biogrid_id"]
                or ep["entrez"] == target["target_entrez_id"]
                or (ep["symbol"] == symbol and ep["taxon"] == 9606)
            )
            if not is_target_side:
                continue
            if ep["taxon"] not in (None, 9606) and ep["taxon"] != 9606:
                return None, STATUS_TARGET_MISMATCH, violations
            if ep["biogrid_id"] not in (None, target["target_biogrid_id"]):
                return None, STATUS_TARGET_MISMATCH, violations
            if ep["entrez"] not in (None, target["target_entrez_id"]):
                return None, STATUS_TARGET_MISMATCH, violations

    if violations and target is not None:
        return target, "target_missing_on_row", violations
    if target is None:
        return None, STATUS_TARGET_MISMATCH, violations
    return target, None, violations


def _partner_from_row(row: dict[str, Any], target: dict[str, Any]) -> dict[str, Any]:
    tid = target["target_biogrid_id"]
    a_id = _as_int(row.get("BIOGRID_ID_A"))
    b_id = _as_int(row.get("BIOGRID_ID_B"))
    if a_id == tid and b_id == tid:
        side = "A"
        partner_side = "A"
        self_hit = True
    elif a_id == tid:
        side = "A"
        partner_side = "B"
        self_hit = False
    elif b_id == tid:
        side = "B"
        partner_side = "A"
        self_hit = False
    else:
        side = "A"
        partner_side = "B"
        self_hit = False
    return {
        "query_side": side,
        "partner_side": partner_side,
        "self_interaction": self_hit,
        "partner_symbol": _blank(row.get(f"OFFICIAL_SYMBOL_{partner_side}")),
        "partner_biogrid_id": _as_int(row.get(f"BIOGRID_ID_{partner_side}")),
        "partner_entrez_id": _as_int(row.get(f"ENTREZ_GENE_{partner_side}")),
        "partner_taxon_id": _as_int(row.get(f"ORGANISM_{partner_side}")),
        "organism_a": _as_int(row.get("ORGANISM_A")),
        "organism_b": _as_int(row.get("ORGANISM_B")),
    }


def build_nonredundant_pairs(
    rows: list[dict[str, Any]],
    target: dict[str, Any],
    *,
    page_api_runs: dict[int, str] | None = None,
    page_raw_ids: dict[int, str] | None = None,
) -> dict[str, Any]:
    """Aggregate unordered BioGRID-ID pairs with stable primary provenance."""
    page_api_runs = page_api_runs or {}
    page_raw_ids = page_raw_ids or {}
    ordered = sorted(rows, key=_stable_sort_key)
    buckets: dict[tuple[int, int], list[dict[str, Any]]] = {}
    for row in ordered:
        key = canonical_pair_key(row.get("BIOGRID_ID_A"), row.get("BIOGRID_ID_B"))
        if key is None:
            continue
        buckets.setdefault(key, []).append(row)

    pairs: list[dict[str, Any]] = []
    manifest: dict[str, Any] = {}
    physical = genetic = cross = self_ev = 0
    for row in ordered:
        et = _blank(row.get("EXPERIMENTAL_SYSTEM_TYPE")).lower()
        if et == "physical":
            physical += 1
        elif et == "genetic":
            genetic += 1
        oa = _as_int(row.get("ORGANISM_A"))
        ob = _as_int(row.get("ORGANISM_B"))
        if oa is not None and ob is not None and oa != ob:
            cross += 1
        if _as_int(row.get("BIOGRID_ID_A")) == _as_int(row.get("BIOGRID_ID_B")):
            self_ev += 1

    for key in sorted(buckets.keys()):
        contrib = buckets[key]
        primary = contrib[0]
        orient = _partner_from_row(primary, target)
        systems = sorted({_blank(r.get("EXPERIMENTAL_SYSTEM")) for r in contrib if _blank(r.get("EXPERIMENTAL_SYSTEM"))})
        system_types = sorted(
            {_blank(r.get("EXPERIMENTAL_SYSTEM_TYPE")).lower() for r in contrib if _blank(r.get("EXPERIMENTAL_SYSTEM_TYPE"))}
        )
        pubmeds = sorted({_blank(r.get("PUBMED_ID")) for r in contrib if _blank(r.get("PUBMED_ID"))})
        throughputs = sorted({_blank(r.get("THROUGHPUT")) for r in contrib if _blank(r.get("THROUGHPUT"))})
        mods = sorted({_blank(r.get("MODIFICATION")) for r in contrib if _blank(r.get("MODIFICATION"))})
        quals = sorted({_blank(r.get("QUALIFICATIONS")) for r in contrib if _blank(r.get("QUALIFICATIONS"))})
        page_starts = sorted({int(r.get("_page_start") or 0) for r in contrib})
        contrib_api = []
        contrib_raw = []
        for r in contrib:
            ps = int(r.get("_page_start") or 0)
            if page_api_runs.get(ps):
                contrib_api.append(page_api_runs[ps])
            if page_raw_ids.get(ps):
                contrib_raw.append(page_raw_ids[ps])
        # dedupe preserve order
        contrib_api = list(dict.fromkeys(contrib_api))
        contrib_raw = list(dict.fromkeys(contrib_raw))
        primary_page = int(primary.get("_page_start") or 0)
        pair = {
            "canonical_pair": list(key),
            "gene": target["target_symbol"],
            "target_biogrid_id": target["target_biogrid_id"],
            "partner_official_symbol": orient["partner_symbol"],
            "partner_biogrid_id": orient["partner_biogrid_id"],
            "partner_entrez_id": orient["partner_entrez_id"],
            "partner_taxon_id": orient["partner_taxon_id"],
            "self_interaction": bool(orient["self_interaction"]),
            "cross_species": bool(
                orient["organism_a"] is not None
                and orient["organism_b"] is not None
                and orient["organism_a"] != orient["organism_b"]
            ),
            "evidence_record_count": len(contrib),
            "experimental_system_types": system_types,
            "experimental_systems": systems,
            "pubmed_ids": pubmeds,
            "throughput_values": throughputs,
            "modifications": mods,
            "qualifications": quals,
            "primary_page_start": primary_page,
            "primary_biogrid_interaction_id": _as_int(primary.get("BIOGRID_INTERACTION_ID")),
            "primary_api_run_id": page_api_runs.get(primary_page),
            "primary_raw_artifact_id": page_raw_ids.get(primary_page),
            "contributing_api_run_ids": contrib_api,
            "contributing_raw_artifact_ids": contrib_raw,
            "contributing_interaction_ids": [
                _as_int(r.get("BIOGRID_INTERACTION_ID")) for r in contrib
            ],
        }
        pairs.append(pair)
        manifest[f"{key[0]}|{key[1]}"] = {
            "canonical_pair": list(key),
            "contributing_api_run_ids": contrib_api,
            "contributing_raw_artifact_ids": contrib_raw,
            "contributing_interaction_ids": pair["contributing_interaction_ids"],
            "primary_api_run_id": pair["primary_api_run_id"],
            "primary_raw_artifact_id": pair["primary_raw_artifact_id"],
        }

    return {
        "pairs": pairs,
        "derived_manifest": manifest,
        "counts": {
            "raw_evidence_record_count": len(ordered),
            "nonredundant_pair_count": len(pairs),
            "physical_evidence_record_count": physical,
            "genetic_evidence_record_count": genetic,
            "cross_species_evidence_record_count": cross,
            "self_interaction_evidence_record_count": self_ev,
        },
        "ordered_rows": ordered,
    }


def _autosize(ws: Any, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        width = 8
        for cell in col[:200]:
            width = max(width, min(max_width, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = width


def write_biogrid_workbook(
    path: Path,
    *,
    gene_symbol: str,
    target: dict[str, Any] | None,
    pairs: list[dict[str, Any]],
    evidence_rows: list[dict[str, Any]],
    provenance: list[tuple[str, Any]],
) -> None:
    """Write {GENE}_BIOGRID.xlsx. Caller computes SHA-256 after this returns."""
    wb = Workbook()
    ws_nr = wb.active
    ws_nr.title = "Nonredundant_Interactions"
    nr_headers = [
        "Gene",
        "Target BioGRID ID",
        "Partner official symbol",
        "Partner BioGRID ID",
        "Partner Entrez ID",
        "Partner organism/taxon",
        "Self interaction",
        "Cross species",
        "Evidence record count",
        "Experimental system types",
        "Experimental systems",
        "PubMed IDs",
        "Throughput values",
        "Modifications",
        "Qualifications",
    ]
    ws_nr.append(nr_headers)
    for p in pairs:
        ws_nr.append(
            [
                p.get("gene"),
                p.get("target_biogrid_id"),
                p.get("partner_official_symbol"),
                p.get("partner_biogrid_id"),
                p.get("partner_entrez_id"),
                p.get("partner_taxon_id"),
                bool(p.get("self_interaction")),
                bool(p.get("cross_species")),
                p.get("evidence_record_count"),
                "; ".join(p.get("experimental_system_types") or []),
                "; ".join(p.get("experimental_systems") or []),
                "; ".join(p.get("pubmed_ids") or []),
                "; ".join(p.get("throughput_values") or []),
                "; ".join(p.get("modifications") or []),
                "; ".join(p.get("qualifications") or []),
            ]
        )
    _autosize(ws_nr)

    ws_ev = wb.create_sheet("Evidence_Records")
    ev_headers = [
        "BIOGRID_INTERACTION_ID",
        "ENTREZ_GENE_A",
        "ENTREZ_GENE_B",
        "BIOGRID_ID_A",
        "BIOGRID_ID_B",
        "OFFICIAL_SYMBOL_A",
        "OFFICIAL_SYMBOL_B",
        "SYNONYMS_A",
        "SYNONYMS_B",
        "EXPERIMENTAL_SYSTEM",
        "EXPERIMENTAL_SYSTEM_TYPE",
        "PUBMED_AUTHOR",
        "PUBMED_ID",
        "ORGANISM_A",
        "ORGANISM_B",
        "THROUGHPUT",
        "QUANTITATION",
        "MODIFICATION",
        "QUALIFICATIONS",
        "TAGS",
        "SOURCEDB",
        "SOURCEDB_ID",
        "PUBMED_INTERACTION_COUNT",
        "page_start",
        "within_page_index",
    ]
    ws_ev.append(ev_headers)
    for row in evidence_rows:
        ws_ev.append(
            [
                row.get("BIOGRID_INTERACTION_ID"),
                row.get("ENTREZ_GENE_A"),
                row.get("ENTREZ_GENE_B"),
                row.get("BIOGRID_ID_A"),
                row.get("BIOGRID_ID_B"),
                row.get("OFFICIAL_SYMBOL_A"),
                row.get("OFFICIAL_SYMBOL_B"),
                row.get("SYNONYMS_A"),
                row.get("SYNONYMS_B"),
                row.get("EXPERIMENTAL_SYSTEM"),
                row.get("EXPERIMENTAL_SYSTEM_TYPE"),
                row.get("PUBMED_AUTHOR") or row.get("AUTHOR"),
                row.get("PUBMED_ID"),
                row.get("ORGANISM_A"),
                row.get("ORGANISM_B"),
                row.get("THROUGHPUT"),
                row.get("QUANTITATION"),
                row.get("MODIFICATION"),
                row.get("QUALIFICATIONS"),
                row.get("TAGS"),
                row.get("SOURCEDB") or row.get("SOURCE_DATABASE"),
                row.get("SOURCEDB_ID"),
                row.get("PUBMED_INTERACTION_COUNT"),
                row.get("_page_start"),
                row.get("_within_page_index"),
            ]
        )
    _autosize(ws_ev)

    ws_p = wb.create_sheet("Provenance")
    ws_p.append(["Field", "Value"])
    for k, v in provenance:
        if isinstance(v, (list, dict)):
            import json as _json

            v = _json.dumps(v, ensure_ascii=False, sort_keys=True, default=str)
        ws_p.append([k, v])
    _autosize(ws_p)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _evidence(
    *,
    dossier_run_id: str,
    gene_symbol: str,
    fact_type: str,
    key: str,
    value: dict[str, Any],
    display_text: str,
    api_run_id: str | None = None,
    raw_artifact_id: str | None = None,
) -> EvidenceRecord:
    return EvidenceRecord(
        source_id=make_source_id(bg.SOURCE_NAME, gene_symbol, AssertionType.ppi, key),
        dossier_run_id=dossier_run_id,
        gene_symbol=gene_symbol,
        official_symbol=gene_symbol,
        section=SECTION_PPI,
        subsection=SUBSECTION_5B,
        source_name=bg.SOURCE_NAME,
        source_type=SourceType.interaction_database,
        assertion_type=AssertionType.ppi,
        fact_type=fact_type,
        evidence_grade=EvidenceGrade.C,
        value=value,
        display_text=display_text,
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )


def evaluate_section_5b_complete(
    *,
    status: dict[str, Any],
    html_text: str,
    pdf_path: Path | None,
    attempt_dir: Path | None,
) -> dict[str, Any]:
    checks: dict[str, Any] = {}

    def _ok(name: str, passed: bool, detail: Any = None) -> bool:
        checks[name] = {"passed": bool(passed), "detail": detail}
        return bool(passed)

    rendering = dict(status.get("rendering_status") or {})
    summary = dict(status.get("summary") or {})
    audit = dict(status.get("audit") or {})
    scientific = str(rendering.get("scientific_status") or summary.get("scientific_status") or "")
    presentation = str(
        rendering.get("presentation_status") or summary.get("presentation_status") or ""
    )
    visual = str(rendering.get("visual_status") or summary.get("visual_status") or "")
    figure_requested = bool(
        summary.get("network_figure_requested")
        if "network_figure_requested" in summary
        else audit.get("network_figure_requested", True)
    )
    all_pass = True
    all_pass &= _ok("scientific_status_success", scientific == STATUS_SUCCESS, scientific)
    all_pass &= _ok(
        "presentation_status_success", presentation == STATUS_SUCCESS, presentation
    )

    raw_count = int(summary.get("raw_evidence_record_count") or 0)
    nr_count = int(summary.get("nonredundant_pair_count") or 0)
    xlsx_name = summary.get("supplementary_xlsx")
    xlsx_path = None
    if attempt_dir and xlsx_name:
        xlsx_path = Path(attempt_dir) / "supplementary" / str(xlsx_name)
    if xlsx_path and xlsx_path.exists():
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ev_rows = max(0, (wb["Evidence_Records"].max_row or 1) - 1)
        nr_rows = max(0, (wb["Nonredundant_Interactions"].max_row or 1) - 1)
        wb.close()
        all_pass &= _ok("raw_xlsx_evidence_rows", ev_rows == raw_count, ev_rows)
        all_pass &= _ok("nonredundant_xlsx_rows", nr_rows == nr_count, nr_rows)
        stored = summary.get("supplementary_xlsx_sha256")
        all_pass &= _ok(
            "workbook_sha_matches",
            bool(stored) and sha256_file(xlsx_path) == stored,
            stored,
        )
    else:
        all_pass &= _ok("workbook_present", False, str(xlsx_path))

    audit_json = Path(attempt_dir) / "evidence_records.json" if attempt_dir else None
    if audit_json and audit_json.exists():
        import json as _json

        inv = _json.loads(audit_json.read_text(encoding="utf-8"))
        all_pass &= _ok("raw_audit_inventory_count", len(inv) == raw_count, len(inv))

    html = html_text or ""
    major = len(re.findall(r"5\.\s*Protein-protein interaction \(PPI\) partners", html))
    sub = len(
        re.findall(
            r"<h3\b[^>]*>\s*b\.\s*BioGRID\s*</h3>",
            html,
            flags=re.IGNORECASE,
        )
    )
    all_pass &= _ok("major_5_heading_once", major == 1, major)
    all_pass &= _ok("subsection_5b_heading_once", sub == 1, sub)
    gene = str(summary.get("official_symbol") or summary.get("gene_symbol") or "")
    count_line = f"{gene} has {nr_count} unique interactions." if gene else ""
    all_pass &= _ok(
        "count_line_once",
        (not count_line) or html.count(count_line) == 1,
        html.count(count_line) if count_line else None,
    )
    all_pass &= _ok("html_pdf_exist", bool(html) and bool(pdf_path and Path(pdf_path).exists()))

    all_pass &= _ok(
        "network_figure_requested_recorded",
        "network_figure_requested" in summary or "network_figure_requested" in audit,
        figure_requested,
    )

    if scientific == STATUS_SUCCESS and figure_requested:
        all_pass &= _ok("visual_status_success", visual == STATUS_SUCCESS, visual)
        fig_rel = summary.get("network_figure_relative_path")
        fig_path = None
        if attempt_dir and fig_rel:
            fig_path = Path(attempt_dir) / str(fig_rel)
        elif summary.get("network_figure_local_path"):
            fig_path = Path(str(summary["network_figure_local_path"]))
        all_pass &= _ok(
            "network_figure_png_exists",
            bool(fig_path and fig_path.is_file()),
            str(fig_path),
        )
        stored_fig_sha = summary.get("network_figure_sha256")
        all_pass &= _ok("network_figure_sha_present", bool(stored_fig_sha), stored_fig_sha)
        if fig_path and fig_path.is_file() and stored_fig_sha:
            recomputed = sha256_file(fig_path)
            all_pass &= _ok(
                "network_figure_sha_matches",
                recomputed == stored_fig_sha,
                {"stored": stored_fig_sha, "recomputed": recomputed},
            )
            png_check = validate_network_png(fig_path.read_bytes())
            all_pass &= _ok("network_figure_png_valid", bool(png_check.get("ok")), png_check)
        else:
            all_pass &= _ok("network_figure_sha_matches", False, None)
            all_pass &= _ok("network_figure_png_valid", False, None)

        fig_hits = len(
            re.findall(
                r"<figure\b[^>]*\bsection-5b-network-figure\b",
                html,
                flags=re.IGNORECASE,
            )
        )
        all_pass &= _ok("network_figure_role_once", fig_hits == 1, fig_hits)

        capture = dict(summary.get("figure_capture") or audit.get("figure_capture") or {})
        all_pass &= _ok(
            "browser_capture_provenance",
            bool(capture.get("capture_method") and capture.get("final_url")),
            {
                "capture_method": capture.get("capture_method"),
                "final_url": capture.get("final_url"),
            },
        )
    elif scientific == STATUS_SUCCESS and not figure_requested:
        all_pass &= _ok(
            "visual_status_not_attempted_optional",
            visual == STATUS_NOT_ATTEMPTED,
            visual,
        )
    elif scientific == STATUS_NO_INTERACTIONS:
        all_pass &= _ok(
            "visual_status_not_attempted_no_network",
            visual == STATUS_NOT_ATTEMPTED_NO_NETWORK,
            visual,
        )

    return {
        "complete": all_pass,
        "section_5b_complete": all_pass,
        "checks": checks,
        "scientific_status": scientific,
        "presentation_status": presentation,
        "visual_status": visual,
        "network_figure_requested": figure_requested,
    }


def accept_section_5b_report(
    *,
    gene_symbol: str,
    attempt_dir: Path,
    acceptance: dict[str, Any],
    artifacts: dict[str, Any] | None = None,
    output_root: str | Path | None = None,
    promote_existing: bool = False,
) -> Path | None:
    paths = paths_for(output_root or get_settings().output_path)
    pointer = paths.accepted_gene_pointer(gene_symbol)
    if pointer.exists() and not promote_existing:
        import json as _json

        prior = _json.loads(pointer.read_text(encoding="utf-8"))
        if (prior.get("acceptance") or {}).get("section_5b_complete"):
            return pointer
    return accept_gene_report(
        paths,
        gene_symbol=gene_symbol,
        attempt_dir=attempt_dir,
        acceptance=acceptance,
        artifacts=artifacts,
    )


def allowed_biogrid_hostname(url: str) -> bool:
    """Exact hostname allow-list (no substring host checks)."""
    host = urlparse(url).hostname
    return host in ALLOWED_BIOGRID_HOSTS


def _png_is_nonuniform(img: Any) -> bool:
    """Reject images that are effectively a single uniform color."""
    rgb = img.convert("RGB")
    extrema = rgb.getextrema()
    if all(lo == hi for lo, hi in extrema):
        return False
    # Downsample for a cheap unique-color check on large images.
    sample = rgb.resize((min(64, rgb.width), min(64, rgb.height)))
    colors = sample.getcolors(maxcolors=64 * 64 + 1)
    if colors is None:
        return True
    if len(colors) <= 1:
        return False
    # Nearly-uniform: one color dominates >99.5% of sampled pixels.
    total = sum(count for count, _color in colors)
    top = max(count for count, _color in colors)
    if total and (top / total) >= 0.995 and len(colors) <= 3:
        return False
    return True


def validate_network_png(content: bytes) -> dict[str, Any]:
    if not content.startswith(b"\x89PNG\r\n\x1a\n"):
        return {"ok": False, "reason": "not_png", "bytes": len(content)}
    try:
        from PIL import Image

        img = Image.open(io.BytesIO(content))
        img.load()
        w, h = img.size
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "reason": f"pil:{exc}", "bytes": len(content)}
    if w < PNG_MIN_WIDTH or h < PNG_MIN_HEIGHT:
        return {
            "ok": False,
            "reason": "dimensions",
            "width": w,
            "height": h,
            "bytes": len(content),
        }
    if not _png_is_nonuniform(img):
        return {
            "ok": False,
            "reason": "blank_or_uniform",
            "width": w,
            "height": h,
            "bytes": len(content),
        }
    digest = sha256_bytes(content)
    return {
        "ok": True,
        "width": w,
        "height": h,
        "bytes": len(content),
        "sha256": digest,
    }


def _viewer_metrics_script() -> str:
    return f"""() => {{
      const root = document.querySelector('{INTERACTION_BODY_SELECTOR}');
      if (!root) {{
        return {{
          interaction_body_found: false,
          interaction_body_visible: false,
          canvas_count: 0,
          visible_canvas_count: 0,
          meaningful_canvas_count: 0,
          largest_canvas_width: 0,
          largest_canvas_height: 0,
          canvases: [],
        }};
      }}
      const style = window.getComputedStyle(root);
      const visible = style && style.display !== 'none' && style.visibility !== 'hidden'
        && root.getClientRects().length > 0;
      const canvases = Array.from(root.querySelectorAll('canvas'));
      const details = canvases.map((c) => {{
        const r = c.getBoundingClientRect();
        const w = Math.max(c.width || 0, Math.floor(r.width || 0));
        const h = Math.max(c.height || 0, Math.floor(r.height || 0));
        const cs = window.getComputedStyle(c);
        const cv = cs && cs.display !== 'none' && cs.visibility !== 'hidden'
          && r.width > 0 && r.height > 0;
        return {{ width: w, height: h, visible: cv }};
      }});
      const visibleCanvases = details.filter((d) => d.visible);
      const meaningful = visibleCanvases.filter(
        (d) => d.width >= {VIEWER_MIN_CANVAS_WIDTH} && d.height >= {VIEWER_MIN_CANVAS_HEIGHT}
      );
      let largest_w = 0;
      let largest_h = 0;
      for (const d of details) {{
        if (d.width * d.height > largest_w * largest_h) {{
          largest_w = d.width;
          largest_h = d.height;
        }}
      }}
      return {{
        interaction_body_found: true,
        interaction_body_visible: !!visible,
        canvas_count: canvases.length,
        visible_canvas_count: visibleCanvases.length,
        meaningful_canvas_count: meaningful.length,
        largest_canvas_width: largest_w,
        largest_canvas_height: largest_h,
        canvases: details,
      }};
    }}"""


def _data_url_to_png_bytes(data_url: str) -> bytes | None:
    if not data_url or not isinstance(data_url, str) or "," not in data_url:
        return None
    header, payload = data_url.split(",", 1)
    if "png" not in header.lower():
        return None
    try:
        return base64.b64decode(payload)
    except Exception:  # noqa: BLE001
        return None


def _install_export_interceptors(page: Any) -> None:
    page.evaluate(
        """() => {
          window.__biogridPngCapture = null;
          window.__biogridLastPngBlobUrl = null;
          if (window.__biogridExportHookInstalled) return;
          window.__biogridExportHookInstalled = true;
          const origClick = HTMLAnchorElement.prototype.click;
          HTMLAnchorElement.prototype.click = function(...args) {
            try {
              if (this.hasAttribute('download') || this.download) {
                window.__biogridPngCapture = {
                  href: this.getAttribute('href') || this.href || null,
                  download: this.getAttribute('download') || this.download || null,
                };
              }
            } catch (e) {}
            return origClick.apply(this, args);
          };
          const origCreate = URL.createObjectURL.bind(URL);
          URL.createObjectURL = function(obj) {
            const url = origCreate(obj);
            try {
              if (obj && typeof Blob !== 'undefined' && obj instanceof Blob) {
                const type = (obj.type || '').toLowerCase();
                if (type.includes('png') || type.includes('octet-stream') || !type) {
                  window.__biogridLastPngBlobUrl = url;
                }
              }
            } catch (e) {}
            return url;
          };
        }"""
    )


def _read_intercepted_export_png(page: Any) -> tuple[bytes | None, dict[str, Any]]:
    meta: dict[str, Any] = {}
    payload = page.evaluate(
        """async () => {
          async function blobUrlToDataUrl(url) {
            const resp = await fetch(url);
            const blob = await resp.blob();
            return await new Promise((resolve, reject) => {
              const fr = new FileReader();
              fr.onload = () => resolve(fr.result);
              fr.onerror = reject;
              fr.readAsDataURL(blob);
            });
          }
          const cap = window.__biogridPngCapture;
          const blobUrl = window.__biogridLastPngBlobUrl;
          if (cap && cap.href) {
            if (String(cap.href).startsWith('data:image/png')) {
              return { kind: 'data_url', href: cap.href, download: cap.download || null };
            }
            if (String(cap.href).startsWith('blob:')) {
              return {
                kind: 'blob_href',
                href: await blobUrlToDataUrl(cap.href),
                download: cap.download || null,
              };
            }
          }
          if (blobUrl) {
            return {
              kind: 'createObjectURL',
              href: await blobUrlToDataUrl(blobUrl),
              download: (cap && cap.download) || null,
            };
          }
          return null;
        }"""
    )
    if not payload or not isinstance(payload, dict):
        return None, meta
    meta["client_export_kind"] = payload.get("kind")
    meta["suggested_filename"] = payload.get("download")
    raw = _data_url_to_png_bytes(str(payload.get("href") or ""))
    return raw, meta


def _single_canvas_to_data_url(page: Any) -> tuple[bytes | None, dict[str, Any]]:
    meta: dict[str, Any] = {}
    try:
        data_url = page.evaluate(
            f"""() => {{
              const root = document.querySelector('{INTERACTION_BODY_SELECTOR}');
              if (!root) return {{ error: 'no_body' }};
              const canvases = Array.from(root.querySelectorAll('canvas'));
              const meaningful = canvases.filter((c) => {{
                const r = c.getBoundingClientRect();
                const w = Math.max(c.width || 0, Math.floor(r.width || 0));
                const h = Math.max(c.height || 0, Math.floor(r.height || 0));
                return w >= {VIEWER_MIN_CANVAS_WIDTH} && h >= {VIEWER_MIN_CANVAS_HEIGHT};
              }});
              if (meaningful.length !== 1) {{
                return {{ error: 'not_single_meaningful_canvas', count: meaningful.length }};
              }}
              try {{
                return {{ data_url: meaningful[0].toDataURL('image/png') }};
              }} catch (e) {{
                return {{ error: 'canvas_security:' + String(e && e.name || e) }};
              }}
            }}"""
        )
    except Exception as exc:  # noqa: BLE001
        meta["error"] = f"single_canvas_eval:{exc}"
        return None, meta
    if not isinstance(data_url, dict):
        meta["error"] = "single_canvas_unexpected_result"
        return None, meta
    if data_url.get("error"):
        meta["error"] = data_url.get("error")
        meta["count"] = data_url.get("count")
        return None, meta
    raw = _data_url_to_png_bytes(str(data_url.get("data_url") or ""))
    if raw is None:
        meta["error"] = "single_canvas_decode_failed"
    return raw, meta


def capture_biogrid_network_figure(
    *,
    biogrid_id: int,
    gene_symbol: str,
    timeout_ms: int = 120_000,
    debug_dir: Path | None = None,
) -> tuple[bytes | None, dict[str, Any]]:
    """Official Network Viewer capture via BioGRID DOM export / container screenshot."""
    symbol = gene_symbol.strip().lower()
    expected_symbol = gene_symbol.strip().upper()
    requested = f"https://{BIOGRID_HOST}/{biogrid_id}/network/homo-sapiens/{symbol}.html"
    audit: dict[str, Any] = {
        "requested_url": requested,
        "final_url": None,
        "page_title": None,
        "expected_biogrid_id": biogrid_id,
        "expected_gene_symbol": expected_symbol,
        "page_retrieval_timestamp": datetime.now(timezone.utc).isoformat(),
        "browser_engine": "chromium",
        "capture_method": None,
        "export_selector": EXPORT_PNG_SELECTOR,
        "interaction_body_selector": INTERACTION_BODY_SELECTOR,
        "methods_attempted": [],
        "failure_stage": None,
        "requires_window_cy": False,
        "requires_window_graph": False,
    }

    def _fail(stage: str, error: str) -> tuple[bytes | None, dict[str, Any]]:
        audit["failure_stage"] = stage
        audit["error"] = error
        return None, audit

    def _accept(raw: bytes, method: str, extra: dict[str, Any] | None = None) -> tuple[bytes | None, dict[str, Any]]:
        audit["methods_attempted"].append(method)
        audit["capture_method"] = method
        if extra:
            audit.update(extra)
        check = validate_network_png(raw)
        audit["png_validation"] = check
        audit["width"] = check.get("width")
        audit["height"] = check.get("height")
        audit["byte_length"] = len(raw)
        if not check.get("ok"):
            audit["failure_stage"] = "png_validation_failed"
            audit["error"] = check.get("reason")
            return None, audit
        audit["output_png_sha256"] = check.get("sha256") or sha256_bytes(raw)
        audit["failure_stage"] = None
        audit.pop("error", None)
        return raw, audit

    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:  # noqa: BLE001
        return _fail("navigation_failed", f"playwright_unavailable:{exc}")

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 1600, "height": 1200})
            try:
                page.goto(requested, wait_until="domcontentloaded", timeout=timeout_ms)
            except Exception as exc:  # noqa: BLE001
                browser.close()
                return _fail("navigation_failed", str(exc))

            final_url = page.url
            audit["final_url"] = final_url
            try:
                audit["page_title"] = page.title()
            except Exception:  # noqa: BLE001
                audit["page_title"] = None

            if not allowed_biogrid_hostname(final_url):
                browser.close()
                return _fail(
                    "identity_validation_failed",
                    f"unexpected_host:{urlparse(final_url).hostname}",
                )

            try:
                body = page.locator(INTERACTION_BODY_SELECTOR)
                body.wait_for(state="visible", timeout=timeout_ms)
            except Exception as exc:  # noqa: BLE001
                metrics = page.evaluate(_viewer_metrics_script())
                audit.update(metrics if isinstance(metrics, dict) else {})
                if debug_dir is not None:
                    try:
                        debug_dir.mkdir(parents=True, exist_ok=True)
                        page.screenshot(
                            path=str(debug_dir / "biogrid_network_debug.png"),
                            full_page=True,
                        )
                        audit["debug_screenshot"] = str(debug_dir / "biogrid_network_debug.png")
                    except Exception:  # noqa: BLE001
                        pass
                browser.close()
                return _fail("viewer_container_missing", str(exc))

            if not body.is_visible():
                browser.close()
                return _fail("viewer_container_missing", "interaction_body_not_visible")

            try:
                page.wait_for_function(
                    f"""() => {{
                      const root = document.querySelector('{INTERACTION_BODY_SELECTOR}');
                      if (!root) return false;
                      const canvases = Array.from(root.querySelectorAll('canvas'));
                      if (!canvases.length) return false;
                      return canvases.some((c) => {{
                        const r = c.getBoundingClientRect();
                        const w = Math.max(c.width || 0, Math.floor(r.width || 0));
                        const h = Math.max(c.height || 0, Math.floor(r.height || 0));
                        return w >= {VIEWER_MIN_CANVAS_WIDTH} && h >= {VIEWER_MIN_CANVAS_HEIGHT};
                      }});
                    }}""",
                    timeout=timeout_ms,
                )
            except Exception as exc:  # noqa: BLE001
                metrics = page.evaluate(_viewer_metrics_script())
                audit.update(metrics if isinstance(metrics, dict) else {})
                browser.close()
                return _fail("viewer_not_rendered", str(exc))

            page.wait_for_timeout(VIEWER_STABILIZE_MS)
            metrics = page.evaluate(_viewer_metrics_script())
            if isinstance(metrics, dict):
                audit.update(metrics)

            body_text = page.content()
            identity_ok = (
                str(biogrid_id) in body_text
                or expected_symbol in body_text.upper()
                or expected_symbol.lower() in (final_url or "").lower()
            )
            if not identity_ok:
                browser.close()
                return _fail("identity_validation_failed", "identity_not_found_on_page")

            # --- Primary: official Export to PNG button ---
            export = page.locator(EXPORT_PNG_SELECTOR)
            export_count = export.count()
            audit["export_button_found"] = export_count > 0
            audit["export_button_visible"] = False
            if export_count > 0:
                audit.update(_reveal_biogrid_export_control(page))
                try:
                    audit["export_button_visible"] = bool(export.first.is_visible())
                except Exception:  # noqa: BLE001
                    audit["export_button_visible"] = False

            if export_count > 0:
                _install_export_interceptors(page)
                # Prefer Playwright download event.
                try:
                    with page.expect_download(timeout=EXPORT_DOWNLOAD_TIMEOUT_MS) as download_info:
                        _click_biogrid_export_png(page, export)
                    download = download_info.value
                    suggested = download.suggested_filename
                    audit["suggested_filename"] = suggested
                    tmp_path = download.path()
                    if tmp_path:
                        raw = Path(tmp_path).read_bytes()
                    else:
                        dest = Path(
                            debug_dir or Path(".")
                        ) / f"_biogrid_export_{biogrid_id}.png"
                        download.save_as(str(dest))
                        raw = dest.read_bytes()
                        try:
                            dest.unlink(missing_ok=True)
                        except Exception:  # noqa: BLE001
                            pass
                    accepted, audit = _accept(
                        raw,
                        "biogrid_export_png_button",
                        {"export_path": "playwright_download"},
                    )
                    if accepted is not None:
                        browser.close()
                        return accepted, audit
                except Exception as download_exc:  # noqa: BLE001
                    audit["methods_attempted"].append("biogrid_export_png_button_download")
                    audit["export_download_error"] = str(download_exc)

                # Client-side data/blob export (same method name when bytes come from BioGRID export).
                try:
                    _reveal_biogrid_export_control(page)
                    _click_biogrid_export_png(page, export)
                    page.wait_for_timeout(1_250)
                    raw, client_meta = _read_intercepted_export_png(page)
                    if raw is not None:
                        accepted, audit = _accept(
                            raw,
                            "biogrid_export_png_button",
                            {**client_meta, "export_path": "client_side_data_or_blob"},
                        )
                        if accepted is not None:
                            browser.close()
                            return accepted, audit
                    else:
                        audit["methods_attempted"].append("biogrid_export_png_button_client")
                        audit["export_client_error"] = "no_client_side_png_captured"
                except Exception as client_exc:  # noqa: BLE001
                    audit["methods_attempted"].append("biogrid_export_png_button_client")
                    audit["export_client_error"] = str(client_exc)
            else:
                audit["methods_attempted"].append("biogrid_export_png_button")
                audit["export_button_error"] = "export_button_missing"

            # --- Secondary: screenshot the official viewer container ---
            try:
                body.scroll_into_view_if_needed()
                box = body.bounding_box()
                audit["interaction_body_bbox"] = box
                if not box or box.get("width", 0) < 100 or box.get("height", 0) < 100:
                    raise RuntimeError(f"interaction_body_bbox_too_small:{box}")
                raw = body.screenshot(type="png")
                accepted, audit = _accept(raw, "interaction_body_screenshot")
                if accepted is not None:
                    browser.close()
                    return accepted, audit
            except Exception as shot_exc:  # noqa: BLE001
                audit["methods_attempted"].append("interaction_body_screenshot")
                audit["screenshot_error"] = str(shot_exc)
                if audit.get("failure_stage") != "png_validation_failed":
                    audit["failure_stage"] = "viewer_screenshot_failed"

            # --- Tertiary: single meaningful canvas only ---
            meaningful = int(audit.get("meaningful_canvas_count") or 0)
            if meaningful == 1:
                raw, canvas_meta = _single_canvas_to_data_url(page)
                audit.update({f"single_canvas_{k}": v for k, v in canvas_meta.items()})
                if raw is not None:
                    accepted, audit = _accept(raw, "single_canvas_toDataURL")
                    if accepted is not None:
                        browser.close()
                        return accepted, audit
                else:
                    audit["methods_attempted"].append("single_canvas_toDataURL")
            else:
                audit["methods_attempted"].append("single_canvas_toDataURL_skipped")
                audit["single_canvas_skip_reason"] = (
                    f"meaningful_canvas_count={meaningful}"
                )

            if debug_dir is not None:
                try:
                    debug_dir.mkdir(parents=True, exist_ok=True)
                    page.screenshot(
                        path=str(debug_dir / "biogrid_network_debug.png"),
                        full_page=True,
                    )
                    audit["debug_screenshot"] = str(debug_dir / "biogrid_network_debug.png")
                except Exception:  # noqa: BLE001
                    pass

            browser.close()
            if not audit.get("export_button_found"):
                return _fail("export_button_missing", audit.get("export_button_error") or "missing")
            if audit.get("failure_stage") == "png_validation_failed":
                return None, audit
            if audit.get("screenshot_error"):
                return _fail(
                    "viewer_screenshot_failed",
                    str(audit.get("screenshot_error")),
                )
            return _fail(
                "export_download_failed",
                str(
                    audit.get("export_client_error")
                    or audit.get("export_download_error")
                    or "all_capture_methods_failed"
                ),
            )
    except Exception as exc:  # noqa: BLE001
        return _fail("navigation_failed", str(exc))


def _reveal_biogrid_export_control(page: Any) -> dict[str, Any]:
    """Open BioGRID export submenu so the PNG control is clickable."""
    meta: dict[str, Any] = {"menu_open_attempts": []}
    for sel in (
        'a.menu-button:has-text("Export")',
        'button:has-text("Export")',
        '.menu-label:has-text("Export")',
        "text=Export",
    ):
        try:
            loc = page.locator(sel)
            if loc.count() <= 0:
                continue
            loc.first.click(timeout=2_000)
            meta["menu_open_attempts"].append(sel)
            page.wait_for_timeout(250)
        except Exception as exc:  # noqa: BLE001
            meta["menu_open_attempts"].append(f"{sel}:err:{exc}")
    try:
        revealed = page.evaluate(
            f"""() => {{
              const btn = document.querySelector('{EXPORT_PNG_SELECTOR}');
              if (!btn) return {{ revealed: false }};
              let el = btn.parentElement;
              while (el) {{
                try {{
                  el.classList.add('open', 'active', 'show');
                  const style = el.style;
                  if (style) {{
                    style.display = 'block';
                    style.visibility = 'visible';
                    style.opacity = '1';
                    style.pointerEvents = 'auto';
                  }}
                }} catch (e) {{}}
                el = el.parentElement;
              }}
              try {{ btn.scrollIntoView({{ block: 'center', inline: 'center' }}); }} catch (e) {{}}
              const r = btn.getBoundingClientRect();
              return {{
                revealed: true,
                top: r.top,
                left: r.left,
                width: r.width,
                height: r.height,
              }};
            }}"""
        )
        if isinstance(revealed, dict):
            meta.update(revealed)
    except Exception as exc:  # noqa: BLE001
        meta["reveal_error"] = str(exc)
    return meta


def _click_biogrid_export_png(page: Any, export_locator: Any) -> None:
    """Click BioGRID Export-to-PNG via DOM (menu item is often off-viewport)."""
    clicked = page.evaluate(
        f"""() => {{
          const btn = document.querySelector('{EXPORT_PNG_SELECTOR}');
          if (!btn) return {{ ok: false, error: 'export_button_missing' }};
          try {{
            btn.dispatchEvent(new MouseEvent('mouseover', {{ bubbles: true, cancelable: true, view: window }}));
            btn.dispatchEvent(new MouseEvent('mousedown', {{ bubbles: true, cancelable: true, view: window }}));
            btn.dispatchEvent(new MouseEvent('mouseup', {{ bubbles: true, cancelable: true, view: window }}));
            btn.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true, view: window }}));
            if (typeof btn.click === 'function') btn.click();
            return {{ ok: true }};
          }} catch (e) {{
            return {{ ok: false, error: String(e && e.message || e) }};
          }}
        }}"""
    )
    if not isinstance(clicked, dict) or not clicked.get("ok"):
        # Last resort: Playwright force click.
        try:
            export_locator.first.click(force=True, timeout=5_000)
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(
                (clicked or {}).get("error") if isinstance(clicked, dict) else str(exc)
            ) from exc


def _persist_biogrid_raw(
    *,
    tool_result: Any,
    response_bytes: bytes | None,
    dossier_run_id: str,
    gene_symbol: str,
    settings: Settings,
    persist_db: bool,
    filename_hint: str,
    artifact_role: str,
    extension: str = "json",
    media_type: str = "application/json",
    transient: Any | None = None,
) -> tuple[Any, dict[str, Any] | None, str | None]:
    from gene_dossier.section_1c import (
        _persist_artifact_bytes,
        _persist_tool_result_json,
        _save_api_run_failure,
        _tool_result_to_api_run,
    )

    meta = bg.lookup_biogrid_meta(tool_result, transient)
    raw_sha = meta.get("response_body_sha256")
    api = _tool_result_to_api_run(
        tool_result, dossier_run_id=dossier_run_id, gene_symbol=gene_symbol
    )
    if not tool_result.success:
        _save_api_run_failure(api, persist_db=persist_db)
        return api, None, raw_sha

    content = response_bytes
    raw_meta = None
    if content is not None:

        def _validate_bytes(blob: bytes) -> dict[str, Any]:
            return {
                "media_type": meta.get("content_type") or media_type,
                "byte_size": len(blob),
            }

        _artifact, raw_meta = _persist_artifact_bytes(
            dossier_run_id=dossier_run_id,
            source_name=bg.SOURCE_NAME,
            content=content,
            extension=extension,
            artifact_type=extension,
            filename_hint=filename_hint,
            settings=settings,
            api_run=api,
            persist_db=persist_db,
            notes={
                "artifact_class": "external_raw",
                "artifact_origin": "biogrid",
                "artifact_role": artifact_role,
                "source_url": meta.get("requested_url") or tool_result.request_url,
                "retrieval_method": "api_bytes",
                "response_body_sha256": raw_sha,
                "exact_raw_bytes": True,
            },
            validate=_validate_bytes,
        )
        if raw_meta is not None:
            raw_meta["sha256"] = raw_meta.get("sha256") or raw_meta.get("content_hash") or raw_sha
    else:
        api, raw_meta = _persist_tool_result_json(
            tr=tool_result,
            dossier_run_id=dossier_run_id,
            gene_symbol=gene_symbol,
            settings=settings,
            persist_db=persist_db,
            filename_hint=filename_hint,
        )
    return api, raw_meta, raw_sha or (raw_meta or {}).get("sha256")


def _cached_bytes(transient: WorkflowTransientContext | None, tool_result: Any) -> bytes | None:
    if transient is None or tool_result is None:
        return None
    meta = bg.lookup_biogrid_meta(tool_result, transient)
    identity = meta.get("request_identity")
    if not identity:
        return None
    cached = transient.get_cached_request(str(identity))
    if isinstance(cached, dict):
        blob = cached.get("response_bytes")
        if isinstance(blob, (bytes, bytearray)):
            return bytes(blob)
    return None


def node_generate_section_5b_derived_artifacts(
    state: dict[str, Any],
    *,
    settings: Settings | None = None,
    persist_db: bool = True,
    transient: WorkflowTransientContext | None = None,
    config: Section5bConfig | None = None,
) -> dict[str, Any]:
    """Generate Section 5b attempt artifacts and return ``section_5b_status``."""
    from gene_dossier.workflow import bind_workflow_transient, reset_workflow_transient

    cfg = settings or get_settings()
    section_cfg = config or Section5bConfig()
    run_type = state.get("run_type")
    selected_keys = list(state.get("selected_section_keys") or [])
    if run_type != "section_bundle" or "5b" not in selected_keys:
        return state

    gene = str(state.get("gene_symbol") or "").strip()
    gene_ids = state.get("gene_ids") or {}
    if not gene:
        gene = str(gene_ids.get("symbol") or gene_ids.get("gene_symbol") or "").strip()
    if not gene:
        errors = list(state.get("errors") or [])
        errors.append("Section 5b requires a resolved gene_symbol")
        return {**state, "errors": errors}

    expected_entrez = _as_int(
        gene_ids.get("entrez_id")
        or gene_ids.get("entrez")
        or gene_ids.get("ncbi_gene_id")
        or (gene_ids.get("human") or {}).get("entrez_id")
    )

    run_id = str(state.get("dossier_run_id") or "")
    api_runs = list(state.get("api_runs") or [])
    raw_artifacts = list(state.get("raw_artifacts") or [])
    tool_results = list(state.get("tool_results") or [])
    evidence_records = list(state.get("evidence_records") or [])
    warnings: list[str] = []

    paths = paths_for(section_cfg.output_root or cfg.output_path)
    attempt_dir = paths.new_gene_attempt(gene, run_id=run_id or None)
    (attempt_dir / "figures").mkdir(exist_ok=True)
    (attempt_dir / "supplementary").mkdir(exist_ok=True)
    (attempt_dir / "raw").mkdir(exist_ok=True)

    bind_token = bind_workflow_transient(transient)
    scientific = STATUS_SOURCE_UNAVAILABLE
    visual = (
        STATUS_NOT_ATTEMPTED
        if not section_cfg.attempt_network_figure
        else STATUS_UNAVAILABLE
    )
    presentation = STATUS_FAILED
    version_str = None
    target = None
    pairs: list[dict[str, Any]] = []
    counts = {
        "raw_evidence_record_count": 0,
        "nonredundant_pair_count": 0,
        "physical_evidence_record_count": 0,
        "genetic_evidence_record_count": 0,
        "cross_species_evidence_record_count": 0,
        "self_interaction_evidence_record_count": 0,
    }
    ordered_rows: list[dict[str, Any]] = []
    derived_manifest: dict[str, Any] = {}
    page_api: dict[int, str] = {}
    page_raw: dict[int, str] = {}
    version_api_id = None
    version_raw_id = None
    figure_path = None
    figure_sha = None
    figure_audit: dict[str, Any] = {}
    workbook_sha = None
    xlsx_name = f"{gene.upper()}_BIOGRID.xlsx"
    primary_api = None
    primary_raw = None

    try:
        version_tr = bg.fetch_version(gene_symbol=gene, settings=cfg, transient=transient)
        tool_results.append(version_tr)
        v_api, v_raw, _ = _persist_biogrid_raw(
            tool_result=version_tr,
            response_bytes=_cached_bytes(transient, version_tr),
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            filename_hint=f"biogrid_version_{gene}",
            artifact_role="biogrid_version",
            extension="txt",
            media_type="text/plain",
            transient=transient,
        )
        api_runs.append(v_api)
        if v_raw:
            raw_artifacts.append(v_raw)
            version_raw_id = v_raw.get("id")
        version_api_id = getattr(v_api, "id", None)
        if version_tr.success:
            payload = bg.unwrap_biogrid_payload(version_tr.data) or {}
            version_str = _blank(payload.get("version"))

        pages, annotated, page_errors = bg.fetch_all_interactions_section_5b(
            gene,
            tax_id=section_cfg.tax_id,
            max_results=section_cfg.max_results,
            settings=cfg,
            transient=transient,
        )
        warnings.extend(page_errors)
        for page in pages:
            tool_results.append(page)
            meta = bg.lookup_biogrid_meta(page, transient)
            page_start = int(meta.get("page_start") or 0)
            p_api, p_raw, _ = _persist_biogrid_raw(
                tool_result=page,
                response_bytes=_cached_bytes(transient, page),
                dossier_run_id=run_id,
                gene_symbol=gene,
                settings=cfg,
                persist_db=persist_db,
                filename_hint=f"biogrid_interactions_{gene}_start{page_start}",
                artifact_role="biogrid_interactions_page",
                transient=transient,
            )
            api_runs.append(p_api)
            if p_raw:
                raw_artifacts.append(p_raw)
                page_raw[page_start] = str(p_raw.get("id"))
                if primary_raw is None:
                    primary_raw = str(p_raw.get("id"))
            page_api[page_start] = str(getattr(p_api, "id", "") or "")
            if primary_api is None:
                primary_api = str(getattr(p_api, "id", "") or "")

        pages_ok = bool(pages) and all(p.success for p in pages)
        if not version_tr.success or not pages_ok:
            scientific = STATUS_SOURCE_UNAVAILABLE
        elif not annotated:
            scientific = STATUS_NO_INTERACTIONS
            visual = STATUS_NOT_ATTEMPTED_NO_NETWORK
            counts = {k: 0 for k in counts}
        else:
            target, err, violations = reconcile_target(
                annotated, gene_symbol=gene, expected_entrez_id=expected_entrez
            )
            if violations:
                warnings.append(f"rows_missing_target={len(violations)}")
            if err == STATUS_TARGET_MISMATCH or target is None:
                scientific = STATUS_TARGET_MISMATCH
            elif err == "target_missing_on_row":
                scientific = STATUS_TARGET_MISMATCH
                write_json_atomic(attempt_dir / "invariant_violations.json", {"rows": violations})
            else:
                built = build_nonredundant_pairs(
                    annotated, target, page_api_runs=page_api, page_raw_ids=page_raw
                )
                pairs = built["pairs"]
                counts = built["counts"]
                ordered_rows = built["ordered_rows"]
                derived_manifest = built["derived_manifest"]
                scientific = STATUS_SUCCESS

                if section_cfg.attempt_network_figure and target.get("target_biogrid_id"):
                    png, figure_audit = capture_biogrid_network_figure(
                        biogrid_id=int(target["target_biogrid_id"]),
                        gene_symbol=gene,
                        debug_dir=attempt_dir / "figures",
                    )
                    if png:
                        figure_path = attempt_dir / "figures" / f"{gene.upper()}_BIOGRID_network.png"
                        figure_path.write_bytes(png)
                        figure_sha = sha256_bytes(png)
                        visual = STATUS_SUCCESS
                    else:
                        visual = STATUS_UNAVAILABLE
                        warnings.append(f"network_figure:{figure_audit.get('error')}")
                elif not section_cfg.attempt_network_figure:
                    visual = STATUS_NOT_ATTEMPTED
                else:
                    visual = STATUS_NOT_ATTEMPTED_NO_NETWORK

        # Workbook for success or no_interactions
        if scientific in {STATUS_SUCCESS, STATUS_NO_INTERACTIONS}:
            xlsx_path = attempt_dir / "supplementary" / xlsx_name
            provenance_rows = [
                ("BioGRID version", version_str),
                ("Gene", gene.upper()),
                ("Target Entrez ID", (target or {}).get("target_entrez_id") or expected_entrez),
                ("Target BioGRID ID", (target or {}).get("target_biogrid_id")),
                ("Target taxon ID", 9606),
                (
                    "Counting definition",
                    "non-redundant unordered BIOGRID_ID_A/B pairs; direction ignored",
                ),
                ("raw_evidence_record_count", counts["raw_evidence_record_count"]),
                ("nonredundant_pair_count", counts["nonredundant_pair_count"]),
                ("physical_evidence_record_count", counts["physical_evidence_record_count"]),
                ("genetic_evidence_record_count", counts["genetic_evidence_record_count"]),
                (
                    "cross_species_evidence_record_count",
                    counts["cross_species_evidence_record_count"],
                ),
                (
                    "self_interaction_evidence_record_count",
                    counts["self_interaction_evidence_record_count"],
                ),
                ("Parser version", PARSER_VERSION),
                ("Request filters", "selfInteractionsExcluded=false; interSpeciesExcluded=false"),
                ("Contributing interaction ApiRun IDs", list(page_api.values())),
                ("Version ApiRun ID", version_api_id),
                ("Retrieval timestamp", datetime.now(timezone.utc).isoformat()),
            ]
            write_biogrid_workbook(
                xlsx_path,
                gene_symbol=gene,
                target=target,
                pairs=pairs,
                evidence_rows=ordered_rows,
                provenance=provenance_rows,
            )
            workbook_sha = sha256_file(xlsx_path)
            write_json_atomic(attempt_dir / "evidence_records.json", ordered_rows)
            write_json_atomic(attempt_dir / "derived_pair_manifest.json", derived_manifest)
            presentation = (
                STATUS_SUCCESS if scientific == STATUS_SUCCESS else STATUS_PARTIAL
            )
        else:
            presentation = STATUS_FAILED

        # Polished evidence
        nr = counts["nonredundant_pair_count"]
        summary_value = {
            **counts,
            "biogrid_version": version_str,
            "target": target,
            "scientific_status": scientific,
            "visual_status": visual,
            "contributing_api_run_ids": list(page_api.values()),
            "contributing_raw_artifact_ids": list(page_raw.values()),
        }
        if scientific == STATUS_NO_INTERACTIONS:
            display = (
                f"No BioGRID interaction records were returned for {gene.upper()} "
                f"in release {version_str or 'unknown'}."
            )
        else:
            display = f"{gene.upper()} has {nr} unique interactions."
        evidence_records.append(
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                fact_type="section_5b_summary",
                key="section-5b-summary",
                value=summary_value,
                display_text=display,
                api_run_id=primary_api,
                raw_artifact_id=primary_raw,
            )
        )
        for pair in pairs:
            partner = pair.get("partner_official_symbol") or pair.get("partner_biogrid_id")
            evidence_records.append(
                _evidence(
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    fact_type="section_5b_nonredundant_pair",
                    key=f"pair-{pair['canonical_pair'][0]}-{pair['canonical_pair'][1]}",
                    value={
                        **pair,
                        "provenance": {
                            "contributing_api_run_ids": pair.get("contributing_api_run_ids"),
                            "contributing_raw_artifact_ids": pair.get(
                                "contributing_raw_artifact_ids"
                            ),
                        },
                    },
                    display_text=f"{gene.upper()}–{partner} BioGRID non-redundant interaction pair",
                    api_run_id=pair.get("primary_api_run_id"),
                    raw_artifact_id=pair.get("primary_raw_artifact_id"),
                )
            )
        if workbook_sha:
            evidence_records.append(
                _evidence(
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    fact_type="section_5b_supplementary_workbook",
                    key="section-5b-supplementary-xlsx",
                    value={"filename": xlsx_name, "sha256": workbook_sha},
                    display_text=f"Supplementary Material ({xlsx_name})",
                    api_run_id=primary_api,
                    raw_artifact_id=primary_raw,
                )
            )
        if visual == STATUS_SUCCESS and figure_path is not None:
            evidence_records.append(
                _evidence(
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    fact_type="section_5b_network_figure",
                    key="section-5b-network-figure",
                    value={
                        "figure_path": str(figure_path),
                        "sha256": figure_sha,
                        "capture": figure_audit,
                    },
                    display_text=f"BioGRID Network Viewer for {gene.upper()}",
                    api_run_id=None,
                    raw_artifact_id=None,
                )
            )

    finally:
        reset_workflow_transient(bind_token)

    summary = {
        "gene_symbol": gene,
        "official_symbol": gene.upper(),
        "biogrid_version": version_str,
        "target": target,
        "target_biogrid_id": (target or {}).get("target_biogrid_id"),
        "target_entrez_id": (target or {}).get("target_entrez_id") or expected_entrez,
        **counts,
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
        "network_figure_requested": bool(section_cfg.attempt_network_figure),
        "scientific_intro": SCIENTIFIC_INTRO,
        "count_line": f"{gene.upper()} has {counts['nonredundant_pair_count']} unique interactions.",
        "supplementary_xlsx": xlsx_name if workbook_sha else None,
        "supplementary_xlsx_sha256": workbook_sha,
        "network_figure_relative_path": (
            f"figures/{gene.upper()}_BIOGRID_network.png" if figure_sha else None
        ),
        "network_figure_local_path": str(figure_path) if figure_path else None,
        "network_figure_sha256": figure_sha,
        "figure_capture": figure_audit,
        "presentation_item_key": f"biogrid-{gene.lower()}",
        "warnings": warnings,
    }
    audit = {
        "gene_attempt_dir": str(attempt_dir),
        "parser_version": PARSER_VERSION,
        "biogrid_version": version_str,
        "version_api_run_id": version_api_id,
        "version_raw_artifact_id": version_raw_id,
        "page_api_run_ids": page_api,
        "page_raw_artifact_ids": page_raw,
        "network_figure_requested": bool(section_cfg.attempt_network_figure),
        "artifacts": {
            "summary.json": "summary.json",
            "evidence_records.json": "evidence_records.json",
            "derived_pair_manifest.json": "derived_pair_manifest.json",
            "manifest.json": MANIFEST_FILENAME,
            f"supplementary/{xlsx_name}": f"supplementary/{xlsx_name}" if workbook_sha else None,
            "network_figure": summary.get("network_figure_relative_path"),
        },
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
        "figure_capture": figure_audit,
    }
    status = {
        "section_key": "5b",
        "summary": summary,
        "rendering_status": {
            "scientific_status": scientific,
            "visual_status": visual,
            "presentation_status": presentation,
        },
        "audit": audit,
    }
    write_json_atomic(attempt_dir / "summary.json", summary)
    write_json_atomic(attempt_dir / "section_5b_status.json", status)
    write_json_atomic(
        attempt_dir / MANIFEST_FILENAME,
        {
            "gene_symbol": gene,
            "parser_version": PARSER_VERSION,
            "artifacts": audit["artifacts"],
            "counts": counts,
        },
    )

    return {
        **state,
        "api_runs": api_runs,
        "raw_artifacts": raw_artifacts,
        "tool_results": tool_results,
        "evidence_records": evidence_records,
        "section_5b_status": status,
        "warnings": list(state.get("warnings") or []) + warnings,
    }


__all__ = [
    "PARSER_VERSION",
    "SECTION_PPI",
    "SUBSECTION_5B",
    "SCIENTIFIC_INTRO",
    "STATUS_FAILED",
    "STATUS_NO_INTERACTIONS",
    "STATUS_NOT_ATTEMPTED",
    "STATUS_NOT_ATTEMPTED_NO_NETWORK",
    "STATUS_PARTIAL",
    "STATUS_SOURCE_UNAVAILABLE",
    "STATUS_SUCCESS",
    "STATUS_TARGET_MISMATCH",
    "STATUS_UNAVAILABLE",
    "Section5bConfig",
    "accept_section_5b_report",
    "allowed_biogrid_hostname",
    "build_nonredundant_pairs",
    "canonical_pair_key",
    "capture_biogrid_network_figure",
    "evaluate_section_5b_complete",
    "node_generate_section_5b_derived_artifacts",
    "reconcile_target",
    "validate_network_png",
    "write_biogrid_workbook",
]
