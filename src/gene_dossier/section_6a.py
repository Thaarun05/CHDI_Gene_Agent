"""Section 6a: Comparative Toxicogenomics Database (CTD) chemical–gene interactions.

Frozen / opt-in: supported in the section bundle, not in ``DEFAULT_SECTION_BUNDLE_KEYS``.
Scientific, aggregation, ranking, workbook, chart, acceptance, and slot routing are locked.
Provenance invariant: ``persist_db=False`` never creates or replaces the accepted CTD
bulk pointer, regardless of caller.

Uses the official shared bulk ``CTD_chem_gene_ixns.tsv.gz`` artifact (Section 2c-style
reuse). Source-level validation covers gzip/header/required columns only; unrelated
malformed rows are audited and skipped. Strict validation applies only to rows whose
GeneID matches the authoritative target.
"""

from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from gene_dossier.config import Settings, get_settings
from gene_dossier.models import (
    AssertionType,
    EvidenceGrade,
    EvidenceRecord,
    SourceType,
)
from gene_dossier.section_6a_sources import (
    MANIFEST_FILENAME,
    OFFICIAL_URL,
    SOURCE_KEY,
    accept_gene_report,
    accept_source,
    load_accepted_source,
    paths_for,
    sha256_bytes,
    sha256_file,
    write_json_atomic,
)
from gene_dossier.source_ids import make_source_id
from gene_dossier.tools import ctd as ctd_client
from gene_dossier.workflow import WorkflowTransientContext

logger = logging.getLogger(__name__)

SECTION_CTD = "CTD perturbations"
SUBSECTION_6A = "Comparative Toxicogenomics Database"
PARSER_VERSION = "section_6a_ctd_v1"

STATUS_SUCCESS = "success"
STATUS_SOURCE_UNAVAILABLE = "source_unavailable"
STATUS_PARTIAL = "partial"
STATUS_FAILED = "failed"
STATUS_NO_INTERACTIONS = "no_interactions"
STATUS_TARGET_MISMATCH = "target_identity_mismatch"
STATUS_UNAVAILABLE = "unavailable"
STATUS_NOT_ATTEMPTED_NO_INTERACTIONS = "not_attempted_no_interactions"

SCIENTIFIC_INTRO = (
    "The Comparative Toxicogenomics Database (CTD) is a robust, publicly available "
    "database that aims to advance understanding about how environmental exposures "
    "affect human health. It provides manually curated information about "
    "chemical–gene/protein interactions, chemical–disease and gene–disease "
    "relationships."
)

SCIENTIFIC_CAVEAT = (
    "Bar lengths reflect the number of CTD curated interaction-evidence occurrences "
    "(PubMed-linked) for each chemical, not potency, effect size, or therapeutic "
    "relevance."
)

COUNTING_DEFINITION = (
    "interaction_count = sum over filtered rows of nonempty PubMed ID tokens after "
    "split(PubMedIDs, '|'); repeated PMIDs across rows count once per row occurrence"
)
TIE_BREAK_DEFINITION = (
    "Top-10: interaction_count DESC, ChemicalName.casefold() ASC, ChemicalID ASC"
)

PNG_MIN_WIDTH = 400
PNG_MIN_HEIGHT = 200
FIGURE_FIGSIZE = (10.0, 6.0)
FIGURE_DPI = 150
BAR_COLOR = "#F66400"  # Rancho orange_link
LABEL_MAX_CHARS = 48


@dataclass
class Section6aConfig:
    force_refresh_ctd_source: bool = False
    promote_ctd_source: bool = False
    output_root: str | Path | None = None


@dataclass
class BulkSourcePayload:
    ok: bool
    content: bytes | None = None
    sha256: str | None = None
    byte_size: int | None = None
    official_url: str | None = OFFICIAL_URL
    origin: str | None = None
    api_run_id: str | None = None
    raw_artifact_id: str | None = None
    source_attempt_id: str | None = None
    attempt_dir: str | None = None
    ctd_report_created: str | None = None
    retrieval_timestamp: str | None = None
    extra_columns: list[str] = field(default_factory=list)
    fieldnames: list[str] = field(default_factory=list)
    error_type: str | None = None
    error_message: str | None = None
    pointer: dict[str, Any] | None = None


def _blank(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null", "-"}:
        return ""
    return text


def _as_int(value: Any) -> int | None:
    text = _blank(value)
    if not text:
        return None
    # Conservative GeneID parse: digits only (optional leading +), no floats/scientific.
    if not re.fullmatch(r"\+?\d+", text):
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def pubmed_occurrence_count(pubmed_field: Any) -> int:
    """Count nonempty PubMed ID tokens after ``|`` split (per-row occurrences)."""
    text = _blank(pubmed_field)
    if not text:
        return 0
    return sum(1 for tok in text.split("|") if tok.strip())


def pubmed_ids_list(pubmed_field: Any) -> list[str]:
    text = _blank(pubmed_field)
    if not text:
        return []
    return [tok.strip() for tok in text.split("|") if tok.strip()]


def filter_target_rows(
    gzip_bytes: bytes,
    *,
    target_gene_id: int,
    target_symbol: str,
) -> dict[str, Any]:
    """Stream-filter CTD bulk rows for the authoritative target gene.

    Source-level validation (gzip/header/required columns) is performed by
    ``iter_bulk_tsv_rows`` and raises on failure.

    Unparseable GeneID rows that cannot be established as belonging to the target
    are skipped and counted in ``malformed_source_row_count`` — they do **not**
    invalidate the shared bulk source.

    Rows with GeneID == target receive **strict** target-row validation. A
    conflicting or malformed target row fails closed for this gene attempt.
    """
    symbol = target_symbol.strip()
    symbol_cf = symbol.casefold()
    meta, row_iter = ctd_client.iter_bulk_tsv_rows(gzip_bytes)

    filtered: list[dict[str, Any]] = []
    malformed_source_row_count = 0
    malformed_diagnostics: list[dict[str, Any]] = []
    target_row_validation_errors: list[dict[str, Any]] = []
    source_order_index = 0  # among all data rows streamed

    for row in row_iter:
        source_order_index += 1
        gene_id_raw = row.get("GeneID")
        parsed_gid = _as_int(gene_id_raw)
        if parsed_gid is None:
            malformed_source_row_count += 1
            if len(malformed_diagnostics) < 50:
                malformed_diagnostics.append(
                    {
                        "source_row_ordinal": source_order_index,
                        "reason": "unparseable_gene_id",
                        "GeneID": gene_id_raw,
                        "ChemicalID": row.get("ChemicalID"),
                        "GeneSymbol": row.get("GeneSymbol"),
                    }
                )
            continue

        if parsed_gid != int(target_gene_id):
            continue

        # Strict target-row validation
        chemical_id = _blank(row.get("ChemicalID"))
        row_symbol = _blank(row.get("GeneSymbol"))
        if not chemical_id:
            target_row_validation_errors.append(
                {
                    "source_row_ordinal": source_order_index,
                    "reason": "empty_chemical_id",
                    "GeneID": parsed_gid,
                    "GeneSymbol": row_symbol,
                }
            )
            continue
        if not row_symbol or row_symbol.casefold() != symbol_cf:
            target_row_validation_errors.append(
                {
                    "source_row_ordinal": source_order_index,
                    "reason": "gene_symbol_mismatch",
                    "GeneID": parsed_gid,
                    "GeneSymbol": row_symbol,
                    "expected_symbol": symbol,
                }
            )
            continue

        kept = dict(row)
        kept["source_order"] = len(filtered)
        kept["_source_row_ordinal"] = source_order_index
        kept["_pubmed_occurrence_count"] = pubmed_occurrence_count(row.get("PubMedIDs"))
        filtered.append(kept)

    if target_row_validation_errors:
        return {
            "ok": False,
            "error_type": STATUS_TARGET_MISMATCH,
            "error_message": (
                f"{len(target_row_validation_errors)} target-row validation error(s)"
            ),
            "filtered_rows": [],
            "filtered_target_row_count": 0,
            "malformed_source_row_count": malformed_source_row_count,
            "malformed_diagnostics": malformed_diagnostics,
            "target_row_validation_errors": target_row_validation_errors,
            "source_meta": meta,
            "data_row_count_streamed": source_order_index,
        }

    return {
        "ok": True,
        "filtered_rows": filtered,
        "filtered_target_row_count": len(filtered),
        "malformed_source_row_count": malformed_source_row_count,
        "malformed_diagnostics": malformed_diagnostics,
        "target_row_validation_errors": [],
        "source_meta": meta,
        "data_row_count_streamed": source_order_index,
    }


def aggregate_by_chemical_id(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Aggregate filtered rows by ChemicalID with PubMed-occurrence counts."""
    buckets: dict[str, dict[str, Any]] = {}
    order: list[str] = []

    for row in rows:
        cid = _blank(row.get("ChemicalID"))
        if not cid:
            continue
        if cid not in buckets:
            order.append(cid)
            buckets[cid] = {
                "chemical_id": cid,
                "chemical_name": _blank(row.get("ChemicalName")),
                "cas_rn_values": set(),
                "raw_ctd_row_count": 0,
                "interaction_count": 0,
                "pubmed_ids": set(),
                "interaction_statements": set(),
                "interaction_actions": set(),
                "gene_forms": set(),
                "organisms": set(),
                "organism_ids": set(),
                "contributing_source_orders": [],
            }
        bucket = buckets[cid]
        if not bucket["chemical_name"] and _blank(row.get("ChemicalName")):
            bucket["chemical_name"] = _blank(row.get("ChemicalName"))
        cas = _blank(row.get("CasRN"))
        if cas:
            bucket["cas_rn_values"].add(cas)
        bucket["raw_ctd_row_count"] += 1
        bucket["interaction_count"] += int(row.get("_pubmed_occurrence_count") or 0)
        for pmid in pubmed_ids_list(row.get("PubMedIDs")):
            bucket["pubmed_ids"].add(pmid)
        interaction = _blank(row.get("Interaction"))
        if interaction:
            bucket["interaction_statements"].add(interaction)
        actions = _blank(row.get("InteractionActions"))
        if actions:
            bucket["interaction_actions"].add(actions)
        forms = _blank(row.get("GeneForms"))
        if forms:
            bucket["gene_forms"].add(forms)
        organism = _blank(row.get("Organism"))
        if organism:
            bucket["organisms"].add(organism)
        oid = _blank(row.get("OrganismID"))
        if oid:
            bucket["organism_ids"].add(oid)
        bucket["contributing_source_orders"].append(row.get("source_order"))

    out: list[dict[str, Any]] = []
    for cid in order:
        b = buckets[cid]
        out.append(
            {
                "chemical_id": cid,
                "chemical_name": b["chemical_name"] or cid,
                "cas_rn": sorted(b["cas_rn_values"]),
                "raw_ctd_row_count": b["raw_ctd_row_count"],
                "interaction_count": b["interaction_count"],
                "unique_pubmed_count": len(b["pubmed_ids"]),
                "unique_interaction_statement_count": len(b["interaction_statements"]),
                "interaction_actions": sorted(b["interaction_actions"]),
                "gene_forms": sorted(b["gene_forms"]),
                "organisms": sorted(b["organisms"]),
                "organism_ids": sorted(b["organism_ids"]),
                "contributing_source_orders": list(b["contributing_source_orders"]),
            }
        )
    return out


def rank_top_chemicals(
    chemicals: list[dict[str, Any]],
    *,
    limit: int = 10,
) -> list[dict[str, Any]]:
    """Deterministic Top-N ranking (does not mutate input order of full list)."""
    ranked = sorted(
        chemicals,
        key=lambda c: (
            -int(c.get("interaction_count") or 0),
            str(c.get("chemical_name") or "").casefold(),
            str(c.get("chemical_id") or ""),
        ),
    )
    out: list[dict[str, Any]] = []
    for i, chem in enumerate(ranked[: max(0, limit)], start=1):
        row = dict(chem)
        row["display_rank"] = i
        out.append(row)
    return out


def _truncate_label(text: str, max_chars: int = LABEL_MAX_CHARS) -> str:
    raw = text or ""
    if len(raw) <= max_chars:
        return raw
    return raw[: max_chars - 1].rstrip() + "…"


def render_top_chemicals_png(
    ranked: list[dict[str, Any]],
    *,
    dpi: int = FIGURE_DPI,
) -> tuple[bytes, dict[str, Any]]:
    """Horizontal bar chart; Rank 1 at top via plot-order reverse only.

    Matplotlib ``barh`` draws the first y-category at the bottom. Feed the
    reverse of ``ranked`` for plotting only — do not mutate ranked data /
    Top_Chemicals ranks. Chart title is omitted; the report heading carries it.
    """
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plot_rows = list(reversed(ranked))
    labels = [
        _truncate_label(str(r.get("chemical_name") or r.get("chemical_id") or ""))
        for r in plot_rows
    ]
    values = [int(r.get("interaction_count") or 0) for r in plot_rows]
    fig, ax = plt.subplots(figsize=FIGURE_FIGSIZE, dpi=dpi)
    y_pos = list(range(len(plot_rows)))
    ax.barh(y_pos, values, color=BAR_COLOR, height=0.7)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel("Interactions", fontsize=10)
    ax.set_xlim(left=0)
    if values:
        ax.set_xlim(right=max(values) * 1.12)
    fig.subplots_adjust(left=0.32, right=0.96, top=0.96, bottom=0.12)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi)
    plt.close(fig)
    content = buf.getvalue()
    meta = {
        "generator": "matplotlib_agg",
        "title": None,
        "rank_1_at_top": True,
        "plot_order": "reversed_ranked_for_barh_only",
        "n_bars": len(ranked),
        "dpi": dpi,
        "figsize": list(FIGURE_FIGSIZE),
        "bar_color": BAR_COLOR,
    }
    return content, meta


def _png_is_nonuniform(img: Any) -> bool:
    """Reject images that are effectively a single uniform color."""
    rgb = img.convert("RGB")
    extrema = rgb.getextrema()
    if all(lo == hi for lo, hi in extrema):
        return False
    sample = rgb.resize((min(64, rgb.width), min(64, rgb.height)))
    colors = sample.getcolors(maxcolors=64 * 64 + 1)
    if colors is None:
        return True
    if len(colors) <= 1:
        return False
    total = sum(count for count, _color in colors)
    top = max(count for count, _color in colors)
    if total and (top / total) >= 0.995 and len(colors) <= 3:
        return False
    return True


def validate_top_chemicals_png(content: bytes) -> dict[str, Any]:
    if not content.startswith(b"\x89PNG\r\n\x1a\n"):
        return {"ok": False, "reason": "not_png", "bytes": len(content)}
    try:
        from PIL import Image

        img = Image.open(io.BytesIO(content))
        img.load()
        w, h = img.size
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "reason": f"pil:{exc}", "bytes": len(content)}
    if w < PNG_MIN_WIDTH or h < PNG_MIN_HEIGHT:
        return {
            "ok": False,
            "reason": "dimensions",
            "width": w,
            "height": h,
            "bytes": len(content),
        }
    if not _png_is_nonuniform(img):
        return {
            "ok": False,
            "reason": "blank_or_uniform",
            "width": w,
            "height": h,
            "bytes": len(content),
        }
    return {
        "ok": True,
        "width": w,
        "height": h,
        "bytes": len(content),
        "sha256": sha256_bytes(content),
    }


def write_ctd_workbook(
    path: Path,
    *,
    gene_symbol: str,
    gene_id: int,
    filtered_rows: list[dict[str, Any]],
    chemicals: list[dict[str, Any]],
    top_chemicals: list[dict[str, Any]],
    provenance: list[tuple[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    top_ids = {str(c["chemical_id"]) for c in top_chemicals}
    rank_by_id = {str(c["chemical_id"]): c["display_rank"] for c in top_chemicals}

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Chemical_Summary"
    ws_sum.append(
        [
            "Gene",
            "GeneID",
            "ChemicalName",
            "ChemicalID",
            "CasRN",
            "Interaction Count",
            "Raw CTD Row Count",
            "Unique PubMed Count",
            "Unique Interaction Statement Count",
            "Interaction Actions",
            "Gene Forms",
            "Organisms",
            "Organism IDs",
            "Displayed In Top 10",
            "Display Rank",
        ]
    )
    for chem in chemicals:
        cid = str(chem["chemical_id"])
        ws_sum.append(
            [
                gene_symbol.upper(),
                gene_id,
                chem.get("chemical_name"),
                cid,
                "|".join(chem.get("cas_rn") or []),
                chem.get("interaction_count"),
                chem.get("raw_ctd_row_count"),
                chem.get("unique_pubmed_count"),
                chem.get("unique_interaction_statement_count"),
                "|".join(chem.get("interaction_actions") or []),
                "|".join(chem.get("gene_forms") or []),
                "|".join(chem.get("organisms") or []),
                "|".join(chem.get("organism_ids") or []),
                "yes" if cid in top_ids else "no",
                rank_by_id.get(cid),
            ]
        )

    ws_rec = wb.create_sheet("Interaction_Records")
    # Prefer required columns + extras present on any row + Source Order
    base_cols = list(ctd_client.BULK_REQUIRED_COLUMNS)
    extra_cols: list[str] = []
    for row in filtered_rows:
        for key in row:
            if key.startswith("_") or key == "source_order":
                continue
            if key not in base_cols and key not in extra_cols:
                extra_cols.append(key)
    headers = base_cols + extra_cols + ["Source Order"]
    ws_rec.append(headers)
    for row in filtered_rows:
        ws_rec.append(
            [row.get(h) if h != "Source Order" else row.get("source_order") for h in headers]
        )

    ws_top = wb.create_sheet("Top_Chemicals")
    ws_top.append(["Rank", "ChemicalName", "ChemicalID", "Interaction Count"])
    for chem in top_chemicals:
        ws_top.append(
            [
                chem.get("display_rank"),
                chem.get("chemical_name"),
                chem.get("chemical_id"),
                chem.get("interaction_count"),
            ]
        )

    ws_prov = wb.create_sheet("Provenance")
    ws_prov.append(["Field", "Value"])
    for key, value in provenance:
        if isinstance(value, (list, tuple, set)):
            value = "|".join(str(v) for v in value)
        ws_prov.append([key, value])

    wb.save(path)
    wb.close()


def _evidence(
    *,
    dossier_run_id: str,
    gene_symbol: str,
    fact_type: str,
    key: str,
    value: dict[str, Any],
    display_text: str,
    api_run_id: str | None = None,
    raw_artifact_id: str | None = None,
) -> EvidenceRecord:
    return EvidenceRecord(
        source_id=make_source_id(
            ctd_client.SOURCE_NAME, gene_symbol, AssertionType.chemical_interaction, key
        ),
        dossier_run_id=dossier_run_id,
        gene_symbol=gene_symbol,
        official_symbol=gene_symbol,
        section=SECTION_CTD,
        subsection=SUBSECTION_6A,
        source_name=ctd_client.SOURCE_NAME,
        source_type=SourceType.chemical_database,
        assertion_type=AssertionType.chemical_interaction,
        fact_type=fact_type,
        evidence_grade=EvidenceGrade.C,
        value=value,
        display_text=display_text,
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )


def download_persist_and_pin_ctd_bulk(
    *,
    paths: Any,
    dossier_run_id: str,
    settings: Settings,
    gene_symbol: str = "",
    persist_db: bool = True,
    promote: bool = True,
    origin: str = "live_download",
    api_runs: list | None = None,
    raw_artifacts: list | None = None,
    tool_results: list | None = None,
) -> BulkSourcePayload:
    """Shared CTD bulk HTTP download + ApiRun/RawArtifact persist + optional pin.

    Every successful download produces exactly one real ApiRun and one raw gzip
    RawArtifact. Accepted pointers are written only when ``promote and persist_db``
    and both ORIGINAL provenance IDs are present.

    Invariant: ``persist_db=False`` never creates or replaces
    ``accepted/sources/ctd_chem_gene_ixns.json``.
    """
    from gene_dossier.section_1c import (
        _persist_artifact_bytes,
        _save_api_run_failure,
        _tool_result_to_api_run,
    )

    api_runs = api_runs if api_runs is not None else []
    raw_artifacts = raw_artifacts if raw_artifacts is not None else []
    tool_results = tool_results if tool_results is not None else []

    def _validate_gzip_bytes(content: bytes) -> dict[str, Any]:
        return {
            "media_type": "application/gzip",
            "byte_size": len(content),
        }

    tr = ctd_client.download_chem_gene_ixns_bulk(settings=settings)
    tool_results.append(tr)
    api = _tool_result_to_api_run(
        tr, dossier_run_id=dossier_run_id, gene_symbol=gene_symbol or "CTD_BULK"
    )
    data = tr.data if isinstance(tr.data, dict) else {}
    content = data.get("content") or b""
    if not tr.success or not content:
        _save_api_run_failure(api, persist_db=persist_db)
        api_runs.append(api)
        return BulkSourcePayload(
            ok=False,
            official_url=OFFICIAL_URL,
            origin=origin,
            error_type=tr.error_type or "download_failed",
            error_message=tr.error_message,
        )

    api_runs.append(api)
    meta: dict[str, Any] | None = None
    try:
        _artifact, meta = _persist_artifact_bytes(
            dossier_run_id=dossier_run_id,
            source_name=tr.source_name,
            content=content,
            extension="gz",
            artifact_type="bin",
            filename_hint="ctd-chem-gene-ixns",
            settings=settings,
            api_run=api,
            persist_db=persist_db,
            notes={
                "artifact_class": "external_raw",
                "artifact_origin": "section_6a_ctd_bulk",
                "artifact_role": SOURCE_KEY,
                "source_url": tr.request_url,
                "retrieval_method": "http_bytes",
                "sha256": data.get("sha256"),
            },
            validate=_validate_gzip_bytes,
        )
        raw_artifacts.append(meta)
    except Exception as exc:  # noqa: BLE001
        return BulkSourcePayload(
            ok=False,
            official_url=OFFICIAL_URL,
            origin=origin,
            error_type="persist_failed",
            error_message=str(exc),
        )

    api_run_id = str(getattr(api, "id", "") or "").strip()
    raw_artifact_id = str((meta or {}).get("id") or "").strip()
    if not api_run_id or not raw_artifact_id:
        return BulkSourcePayload(
            ok=False,
            official_url=OFFICIAL_URL,
            origin=origin,
            error_type="missing_provenance_ids",
            error_message="CTD bulk download missing ApiRun or RawArtifact id",
        )

    try:
        header_meta, _rows = ctd_client.iter_bulk_tsv_rows(content)
    except ValueError as exc:
        return BulkSourcePayload(
            ok=False,
            content=content,
            sha256=str(data.get("sha256") or sha256_bytes(content)),
            byte_size=len(content),
            official_url=OFFICIAL_URL,
            origin=origin,
            api_run_id=api_run_id,
            raw_artifact_id=raw_artifact_id,
            error_type="source_validation_failed",
            error_message=str(exc),
        )

    digest = str(data.get("sha256") or sha256_bytes(content))
    attempt = paths.new_source_attempt(SOURCE_KEY)
    artifact_path = attempt / "CTD_chem_gene_ixns.tsv.gz"
    artifact_path.write_bytes(content)
    retrieval_ts = datetime.now(timezone.utc).isoformat()
    validation = {
        "required_columns_ok": True,
        "extra_columns": header_meta.get("extra_columns") or [],
        "fieldnames": header_meta.get("fieldnames") or [],
        "ctd_report_created": header_meta.get("ctd_report_created"),
        "header_probe_decompressed_bytes": header_meta.get(
            "header_probe_decompressed_bytes"
        ),
        "header_origin": header_meta.get("header_origin"),
    }
    extra = {
        "api_run_id": api_run_id,
        "raw_artifact_id": raw_artifact_id,
        "source_attempt_id": attempt.name,
        "ctd_report_created": validation["ctd_report_created"],
        "retrieval_timestamp": retrieval_ts,
        "parser_version": PARSER_VERSION,
        "final_url": data.get("final_url") or OFFICIAL_URL,
        "content_type": data.get("content_type"),
        "origin": origin,
        "dossier_run_id": dossier_run_id,
    }

    # Accepted CTD pointers require ORIGINAL resolvable ApiRun/RawArtifact rows.
    # persist_db=False => NEVER create or replace accepted/sources/ctd_chem_gene_ixns.json.
    if promote and not persist_db:
        logger.warning(
            "CTD bulk promotion requested with persist_db=False; "
            "skipping accepted source pin (api_run_id=%s raw_artifact_id=%s)",
            api_run_id,
            raw_artifact_id,
        )
    should_accept = bool(promote and persist_db)
    pointer = None
    if should_accept:
        pointer_path = accept_source(
            paths,
            source_key=SOURCE_KEY,
            attempt_dir=attempt,
            artifact_path=artifact_path,
            official_url=OFFICIAL_URL,
            sha256=digest,
            byte_size=len(content),
            validation=validation,
            extra=extra,
        )
        pointer = {
            **extra,
            "artifact_path": str(artifact_path),
            "attempt_dir": str(attempt),
            "sha256": digest,
            "byte_size": len(content),
            "official_url": OFFICIAL_URL,
        }
        write_json_atomic(
            attempt / MANIFEST_FILENAME,
            {
                "source_key": SOURCE_KEY,
                "sha256": digest,
                "accepted_pointer": str(pointer_path),
                "api_run_id": api_run_id,
                "raw_artifact_id": raw_artifact_id,
            },
        )
    else:
        note = "refresh_not_promoted"
        if promote and not persist_db:
            note = "promotion_blocked_persist_db_false"
        write_json_atomic(
            attempt / MANIFEST_FILENAME,
            {
                "source_key": SOURCE_KEY,
                "sha256": digest,
                "note": note,
                "api_run_id": api_run_id,
                "raw_artifact_id": raw_artifact_id,
            },
        )

    return BulkSourcePayload(
        ok=True,
        content=content,
        sha256=digest,
        byte_size=len(content),
        official_url=OFFICIAL_URL,
        origin=origin,
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
        source_attempt_id=attempt.name,
        attempt_dir=str(attempt),
        ctd_report_created=validation["ctd_report_created"],
        retrieval_timestamp=retrieval_ts,
        extra_columns=list(validation["extra_columns"]),
        fieldnames=list(validation["fieldnames"]),
        pointer=pointer,
    )


def _resolve_ctd_bulk_source(
    *,
    paths: Any,
    force_refresh: bool,
    promote_source: bool,
    dossier_run_id: str,
    gene_symbol: str,
    settings: Settings,
    persist_db: bool,
    api_runs: list,
    raw_artifacts: list,
    tool_results: list,
) -> BulkSourcePayload:
    """Reuse accepted CTD bulk; download only on miss or force_refresh."""
    if not force_refresh:
        record = load_accepted_source(
            paths, source_key=SOURCE_KEY, official_url=OFFICIAL_URL
        )
        if record:
            artifact = Path(str(record["artifact_path"]))
            content = artifact.read_bytes()
            validation = dict(record.get("validation") or {})
            return BulkSourcePayload(
                ok=True,
                content=content,
                sha256=str(record.get("sha256") or ""),
                byte_size=int(record.get("byte_size") or len(content)),
                official_url=str(record.get("official_url") or OFFICIAL_URL),
                origin="accepted_pointer",
                api_run_id=str(record["api_run_id"]),
                raw_artifact_id=str(record["raw_artifact_id"]),
                source_attempt_id=str(
                    record.get("source_attempt_id")
                    or Path(str(record.get("attempt_dir") or "")).name
                    or ""
                )
                or None,
                attempt_dir=str(record.get("attempt_dir") or "") or None,
                ctd_report_created=record.get("ctd_report_created")
                or validation.get("ctd_report_created"),
                retrieval_timestamp=record.get("retrieval_timestamp"),
                extra_columns=list(validation.get("extra_columns") or []),
                fieldnames=list(validation.get("fieldnames") or []),
                pointer=record,
            )

    promote = bool(
        persist_db
        and (
            promote_source
            or not load_accepted_source(
                paths, source_key=SOURCE_KEY, official_url=OFFICIAL_URL
            )
        )
    )
    return download_persist_and_pin_ctd_bulk(
        paths=paths,
        dossier_run_id=dossier_run_id,
        settings=settings,
        gene_symbol=gene_symbol,
        persist_db=persist_db,
        promote=promote,
        origin="live_download",
        api_runs=api_runs,
        raw_artifacts=raw_artifacts,
        tool_results=tool_results,
    )


def evaluate_section_6a_complete(
    *,
    status: dict[str, Any],
    html_text: str,
    pdf_path: Path | None,
    attempt_dir: Path | None,
) -> dict[str, Any]:
    checks: dict[str, Any] = {}

    def _ok(name: str, passed: bool, detail: Any = None) -> bool:
        checks[name] = {"passed": bool(passed), "detail": detail}
        return bool(passed)

    rendering = dict(status.get("rendering_status") or {})
    summary = dict(status.get("summary") or {})
    scientific = str(
        rendering.get("scientific_status") or summary.get("scientific_status") or ""
    )
    presentation = str(
        rendering.get("presentation_status") or summary.get("presentation_status") or ""
    )
    visual = str(rendering.get("visual_status") or summary.get("visual_status") or "")
    all_pass = True
    all_pass &= _ok("scientific_status_success", scientific == STATUS_SUCCESS, scientific)
    all_pass &= _ok(
        "presentation_status_success", presentation == STATUS_SUCCESS, presentation
    )

    filtered_count = int(summary.get("filtered_target_row_count") or 0)
    unique_chem = int(summary.get("unique_chemical_count") or 0)
    xlsx_name = summary.get("supplementary_xlsx")
    xlsx_path = None
    if attempt_dir and xlsx_name:
        xlsx_path = Path(attempt_dir) / "supplementary" / str(xlsx_name)

    if xlsx_path and xlsx_path.exists():
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        rec_rows = max(0, (wb["Interaction_Records"].max_row or 1) - 1)
        chem_rows = max(0, (wb["Chemical_Summary"].max_row or 1) - 1)
        top_rows = max(0, (wb["Top_Chemicals"].max_row or 1) - 1)
        # Verify interaction_count sums
        chem_sheet = wb["Chemical_Summary"]
        headers = [c.value for c in next(chem_sheet.iter_rows(min_row=1, max_row=1))]
        ic_idx = headers.index("Interaction Count") if "Interaction Count" in headers else None
        cid_idx = headers.index("ChemicalID") if "ChemicalID" in headers else None
        name_idx = headers.index("ChemicalName") if "ChemicalName" in headers else None
        chem_counts = {}
        chemicals_for_rank: list[dict[str, Any]] = []
        if ic_idx is not None and cid_idx is not None:
            for row in chem_sheet.iter_rows(min_row=2, values_only=True):
                cid = str(row[cid_idx])
                count = int(row[ic_idx] or 0)
                chem_counts[cid] = count
                chemicals_for_rank.append(
                    {
                        "chemical_id": cid,
                        "chemical_name": (
                            str(row[name_idx] or "") if name_idx is not None else cid
                        ),
                        "interaction_count": count,
                    }
                )
        rec_sheet = wb["Interaction_Records"]
        rec_headers = [c.value for c in next(rec_sheet.iter_rows(min_row=1, max_row=1))]
        pmid_idx = rec_headers.index("PubMedIDs") if "PubMedIDs" in rec_headers else None
        rcid_idx = rec_headers.index("ChemicalID") if "ChemicalID" in rec_headers else None
        summed: dict[str, int] = defaultdict(int)
        if pmid_idx is not None and rcid_idx is not None:
            for row in rec_sheet.iter_rows(min_row=2, values_only=True):
                summed[str(row[rcid_idx])] += pubmed_occurrence_count(row[pmid_idx])
        top_sheet = wb["Top_Chemicals"]
        top_ranks = []
        for row in top_sheet.iter_rows(min_row=2, values_only=True):
            top_ranks.append(
                (
                    int(row[0] or 0),
                    str(row[1] or ""),
                    str(row[2] or ""),
                    int(row[3] or 0),
                )
            )
        wb.close()

        all_pass &= _ok("interaction_records_row_count", rec_rows == filtered_count, rec_rows)
        all_pass &= _ok("chemical_summary_unique_count", chem_rows == unique_chem, chem_rows)
        counts_match = chem_counts == dict(summed) or (
            set(chem_counts) == set(summed)
            and all(chem_counts[k] == summed[k] for k in chem_counts)
        )
        all_pass &= _ok("interaction_count_equals_pubmed_sum", counts_match, {
            "summary": chem_counts,
            "from_records": dict(summed),
        })
        all_pass &= _ok("top_chemicals_le_10", top_rows <= 10, top_rows)

        expected_top = rank_top_chemicals(chemicals_for_rank, limit=10)
        expected_tuples = [
            (
                int(c["display_rank"]),
                str(c.get("chemical_name") or ""),
                str(c.get("chemical_id") or ""),
                int(c.get("interaction_count") or 0),
            )
            for c in expected_top
        ]
        all_pass &= _ok(
            "top_chemicals_workbook_matches_expected_ranking",
            top_ranks == expected_tuples,
            {"expected": expected_tuples, "workbook": top_ranks},
        )
        summary_top = summary.get("top_chemicals") or []
        summary_tuples = [
            (
                int(c.get("rank") or c.get("display_rank") or 0),
                str(c.get("chemical_name") or ""),
                str(c.get("chemical_id") or ""),
                int(c.get("interaction_count") or 0),
            )
            for c in summary_top
        ]
        all_pass &= _ok(
            "top_chemicals_summary_matches_expected_ranking",
            summary_tuples == expected_tuples,
            {"expected": expected_tuples, "summary": summary_tuples},
        )
        ranks_ok = [t[0] for t in top_ranks] == list(range(1, len(top_ranks) + 1))
        all_pass &= _ok("top_chemicals_ranks_1_to_n", ranks_ok, [t[0] for t in top_ranks])

        stored = summary.get("supplementary_xlsx_sha256")
        all_pass &= _ok(
            "workbook_sha_matches",
            bool(stored) and sha256_file(xlsx_path) == stored,
            stored,
        )
    else:
        all_pass &= _ok("workbook_present", False, str(xlsx_path))

    html = html_text or ""
    major = len(
        re.findall(
            r"6\.\s*Information on which perturbations affect the gene",
            html,
        )
    )
    sub = len(
        re.findall(
            r"<h3\b[^>]*>\s*a\.\s*Comparative Toxicogenomics Database\s*</h3>",
            html,
            flags=re.IGNORECASE,
        )
    )
    all_pass &= _ok("major_6_heading_once", major == 1, major)
    all_pass &= _ok("subsection_6a_heading_once", sub == 1, sub)
    all_pass &= _ok(
        "html_pdf_exist", bool(html) and bool(pdf_path and Path(pdf_path).exists())
    )

    if scientific == STATUS_SUCCESS:
        all_pass &= _ok("visual_status_success", visual == STATUS_SUCCESS, visual)
        fig_rel = summary.get("top_chemicals_figure_relative_path")
        fig_path = None
        if attempt_dir and fig_rel:
            fig_path = Path(attempt_dir) / str(fig_rel)
        elif summary.get("top_chemicals_figure_local_path"):
            fig_path = Path(str(summary["top_chemicals_figure_local_path"]))
        all_pass &= _ok(
            "top_chemicals_png_exists",
            bool(fig_path and fig_path.is_file()),
            str(fig_path),
        )
        stored_fig = summary.get("top_chemicals_figure_sha256")
        all_pass &= _ok("top_chemicals_sha_present", bool(stored_fig), stored_fig)
        if fig_path and fig_path.is_file() and stored_fig:
            recomputed = sha256_file(fig_path)
            all_pass &= _ok(
                "top_chemicals_sha_matches",
                recomputed == stored_fig,
                {"stored": stored_fig, "recomputed": recomputed},
            )
            png_check = validate_top_chemicals_png(fig_path.read_bytes())
            all_pass &= _ok("top_chemicals_png_valid", bool(png_check.get("ok")), png_check)
        else:
            all_pass &= _ok("top_chemicals_sha_matches", False, None)
            all_pass &= _ok("top_chemicals_png_valid", False, None)
        fig_hits = len(
            re.findall(
                r"<figure\b[^>]*\bsection-6a-top-chemicals-figure\b",
                html,
                flags=re.IGNORECASE,
            )
        )
        all_pass &= _ok("top_chemicals_figure_role_once", fig_hits == 1, fig_hits)
        intro_hits = html.count("Comparative Toxicogenomics Database (CTD)")
        all_pass &= _ok("intro_present", intro_hits >= 1, intro_hits)
    elif scientific == STATUS_NO_INTERACTIONS:
        all_pass &= _ok(
            "visual_status_not_attempted_no_interactions",
            visual == STATUS_NOT_ATTEMPTED_NO_INTERACTIONS,
            visual,
        )

    return {
        "complete": all_pass,
        "section_6a_complete": all_pass,
        "checks": checks,
        "scientific_status": scientific,
        "presentation_status": presentation,
        "visual_status": visual,
    }


def accept_section_6a_report(
    *,
    gene_symbol: str,
    attempt_dir: Path,
    acceptance: dict[str, Any],
    artifacts: dict[str, Any] | None = None,
    output_root: str | Path | None = None,
    promote_existing: bool = False,
) -> Path | None:
    paths = paths_for(output_root or get_settings().output_path)
    pointer = paths.accepted_gene_pointer(gene_symbol)
    if pointer.exists() and not promote_existing:
        import json as _json

        prior = _json.loads(pointer.read_text(encoding="utf-8"))
        if (prior.get("acceptance") or {}).get("section_6a_complete"):
            return pointer
    return accept_gene_report(
        paths,
        gene_symbol=gene_symbol,
        attempt_dir=attempt_dir,
        acceptance=acceptance,
        artifacts=artifacts,
    )


def node_generate_section_6a_derived_artifacts(
    state: dict[str, Any],
    *,
    settings: Settings | None = None,
    persist_db: bool = True,
    transient: WorkflowTransientContext | None = None,
    config: Section6aConfig | None = None,
) -> dict[str, Any]:
    """Generate Section 6a attempt artifacts and return ``section_6a_status``."""
    cfg = settings or get_settings()
    section_cfg = config or Section6aConfig()
    run_type = state.get("run_type")
    selected_keys = list(state.get("selected_section_keys") or [])
    if run_type != "section_bundle" or "6a" not in selected_keys:
        return state

    gene = str(state.get("gene_symbol") or "").strip()
    gene_ids = state.get("gene_ids") or {}
    if not gene:
        gene = str(gene_ids.get("symbol") or gene_ids.get("gene_symbol") or "").strip()
    if not gene:
        errors = list(state.get("errors") or [])
        errors.append("Section 6a requires a resolved gene_symbol")
        return {**state, "errors": errors}

    expected_entrez = _as_int(
        gene_ids.get("entrez_gene_id")
        or gene_ids.get("entrez_id")
        or gene_ids.get("entrez")
        or gene_ids.get("ncbi_gene_id")
        or (gene_ids.get("human") or {}).get("entrez_id")
        or (gene_ids.get("human") or {}).get("entrez_gene_id")
    )

    run_id = str(state.get("dossier_run_id") or "")
    api_runs = list(state.get("api_runs") or [])
    raw_artifacts = list(state.get("raw_artifacts") or [])
    tool_results = list(state.get("tool_results") or [])
    evidence_records = list(state.get("evidence_records") or [])
    warnings: list[str] = []

    paths = paths_for(section_cfg.output_root or cfg.output_path)
    attempt_dir = paths.new_gene_attempt(gene, run_id=run_id or None)
    (attempt_dir / "figures").mkdir(exist_ok=True)
    (attempt_dir / "supplementary").mkdir(exist_ok=True)
    (attempt_dir / "derived").mkdir(exist_ok=True)

    scientific = STATUS_SOURCE_UNAVAILABLE
    visual = STATUS_UNAVAILABLE
    presentation = STATUS_FAILED
    filtered_rows: list[dict[str, Any]] = []
    chemicals: list[dict[str, Any]] = []
    top_chemicals: list[dict[str, Any]] = []
    figure_path: Path | None = None
    figure_sha: str | None = None
    figure_meta: dict[str, Any] = {}
    workbook_sha: str | None = None
    xlsx_name = f"{gene.upper()}_CTD.xlsx"
    filter_audit: dict[str, Any] = {}
    bulk: BulkSourcePayload | None = None

    if expected_entrez is None:
        scientific = STATUS_TARGET_MISMATCH
        warnings.append("missing_authoritative_entrez_gene_id")
        summary_mismatch_detail = "missing_authoritative_entrez_gene_id"
    else:
        summary_mismatch_detail = None
        bulk = _resolve_ctd_bulk_source(
            paths=paths,
            force_refresh=bool(section_cfg.force_refresh_ctd_source),
            promote_source=bool(section_cfg.promote_ctd_source),
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
        )
        if not bulk.ok or not bulk.content:
            scientific = STATUS_SOURCE_UNAVAILABLE
            warnings.append(
                f"ctd_bulk:{bulk.error_type}:{bulk.error_message}"
            )
        else:
            try:
                filter_result = filter_target_rows(
                    bulk.content,
                    target_gene_id=int(expected_entrez),
                    target_symbol=gene,
                )
            except ValueError as exc:
                scientific = STATUS_SOURCE_UNAVAILABLE
                filter_audit = {"source_validation_error": str(exc)}
                warnings.append(f"ctd_source_validation:{exc}")
                filter_result = None

            if filter_result is not None:
                filter_audit = {
                    "malformed_source_row_count": filter_result.get(
                        "malformed_source_row_count"
                    ),
                    "filtered_target_row_count": filter_result.get(
                        "filtered_target_row_count"
                    ),
                    "target_row_validation_errors": filter_result.get(
                        "target_row_validation_errors"
                    ),
                    "malformed_diagnostics": filter_result.get("malformed_diagnostics"),
                    "data_row_count_streamed": filter_result.get(
                        "data_row_count_streamed"
                    ),
                    "extra_columns": (filter_result.get("source_meta") or {}).get(
                        "extra_columns"
                    )
                    or bulk.extra_columns,
                }
                write_json_atomic(attempt_dir / "filter_audit.json", filter_audit)

                if not filter_result.get("ok"):
                    scientific = STATUS_TARGET_MISMATCH
                    summary_mismatch_detail = "target_row_validation"
                else:
                    filtered_rows = list(filter_result.get("filtered_rows") or [])
                    if not filtered_rows:
                        scientific = STATUS_NO_INTERACTIONS
                        visual = STATUS_NOT_ATTEMPTED_NO_INTERACTIONS
                    else:
                        chemicals = aggregate_by_chemical_id(filtered_rows)
                        top_chemicals = rank_top_chemicals(chemicals, limit=10)
                        scientific = STATUS_SUCCESS
                        try:
                            png_bytes, figure_meta = render_top_chemicals_png(
                                top_chemicals
                            )
                            png_check = validate_top_chemicals_png(png_bytes)
                            if png_check.get("ok"):
                                figure_path = (
                                    attempt_dir
                                    / "figures"
                                    / f"{gene.upper()}_CTD_top_chemicals.png"
                                )
                                figure_path.write_bytes(png_bytes)
                                figure_sha = sha256_bytes(png_bytes)
                                visual = STATUS_SUCCESS
                            else:
                                visual = STATUS_UNAVAILABLE
                                warnings.append(f"figure_invalid:{png_check}")
                        except Exception as exc:  # noqa: BLE001
                            visual = STATUS_UNAVAILABLE
                            warnings.append(f"figure_error:{exc}")

    # Workbook for success or no_interactions
    if scientific in {STATUS_SUCCESS, STATUS_NO_INTERACTIONS} and bulk is not None:
        provenance_rows = [
            ("Official URL", bulk.official_url or OFFICIAL_URL),
            ("CTD report created", bulk.ctd_report_created),
            ("Retrieval timestamp", bulk.retrieval_timestamp),
            ("Compressed SHA-256", bulk.sha256),
            ("Compressed byte length", bulk.byte_size),
            ("Source attempt ID", bulk.source_attempt_id),
            ("ApiRun ID", bulk.api_run_id),
            ("RawArtifact ID", bulk.raw_artifact_id),
            ("Source origin", bulk.origin),
            ("Parser version", PARSER_VERSION),
            ("Target Gene Symbol", gene.upper()),
            ("Target GeneID", expected_entrez),
            ("Filtered target row count", len(filtered_rows)),
            ("Unique chemical count", len(chemicals)),
            ("Malformed source row count", filter_audit.get("malformed_source_row_count")),
            (
                "Target row validation error count",
                len(filter_audit.get("target_row_validation_errors") or []),
            ),
            ("Extra columns", bulk.extra_columns or filter_audit.get("extra_columns")),
            ("Chart counting definition", COUNTING_DEFINITION),
            ("Tie-break definition", TIE_BREAK_DEFINITION),
        ]
        xlsx_path = attempt_dir / "supplementary" / xlsx_name
        write_ctd_workbook(
            xlsx_path,
            gene_symbol=gene,
            gene_id=int(expected_entrez or 0),
            filtered_rows=filtered_rows,
            chemicals=chemicals,
            top_chemicals=top_chemicals,
            provenance=provenance_rows,
        )
        workbook_sha = sha256_file(xlsx_path)
        write_json_atomic(
            attempt_dir / "derived" / "chemicals.json",
            {"chemicals": chemicals, "top_chemicals": top_chemicals},
        )
        write_json_atomic(
            attempt_dir / "derived" / "filtered_rows.json",
            {"rows": filtered_rows},
        )
        presentation = (
            STATUS_SUCCESS if scientific == STATUS_SUCCESS else STATUS_PARTIAL
        )
    elif scientific == STATUS_TARGET_MISMATCH:
        presentation = STATUS_FAILED
    else:
        presentation = STATUS_FAILED

    # Polished evidence
    primary_api = bulk.api_run_id if bulk else None
    primary_raw = bulk.raw_artifact_id if bulk else None
    summary_value = {
        "filtered_target_row_count": len(filtered_rows),
        "unique_chemical_count": len(chemicals),
        "top_chemical_count": len(top_chemicals),
        "malformed_source_row_count": filter_audit.get("malformed_source_row_count"),
        "target_row_validation_errors": filter_audit.get("target_row_validation_errors"),
        "scientific_status": scientific,
        "visual_status": visual,
        "target_gene_id": expected_entrez,
        "source_provenance": {
            "api_run_id": primary_api,
            "raw_artifact_id": primary_raw,
            "source_attempt_id": bulk.source_attempt_id if bulk else None,
            "origin": bulk.origin if bulk else None,
            "sha256": bulk.sha256 if bulk else None,
            "official_url": (bulk.official_url if bulk else None) or OFFICIAL_URL,
        },
    }
    if scientific == STATUS_NO_INTERACTIONS:
        display = (
            f"No CTD chemical–gene interaction records were found for "
            f"{gene.upper()} (GeneID {expected_entrez})."
        )
    else:
        display = (
            f"{gene.upper()} has curated CTD chemical–gene interactions with "
            f"{len(chemicals)} unique chemicals "
            f"({len(filtered_rows)} filtered interaction records)."
        )
    evidence_records.append(
        _evidence(
            dossier_run_id=run_id,
            gene_symbol=gene,
            fact_type="section_6a_summary",
            key="section-6a-summary",
            value=summary_value,
            display_text=display,
            api_run_id=primary_api,
            raw_artifact_id=primary_raw,
        )
    )
    for chem in chemicals:
        evidence_records.append(
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                fact_type="section_6a_chemical",
                key=f"chem-{chem['chemical_id']}",
                value=chem,
                display_text=(
                    f"{chem.get('chemical_name')} ({chem['chemical_id']}): "
                    f"{chem.get('interaction_count')} interactions"
                ),
                api_run_id=primary_api,
                raw_artifact_id=primary_raw,
            )
        )
    if workbook_sha:
        evidence_records.append(
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                fact_type="section_6a_supplementary_workbook",
                key="section-6a-supplementary-xlsx",
                value={"filename": xlsx_name, "sha256": workbook_sha},
                display_text=f"Supplementary Material ({xlsx_name})",
                api_run_id=primary_api,
                raw_artifact_id=primary_raw,
            )
        )
    if visual == STATUS_SUCCESS and figure_path is not None:
        evidence_records.append(
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                fact_type="section_6a_top_chemicals_figure",
                key="section-6a-top-chemicals-figure",
                value={
                    "figure_path": str(figure_path),
                    "sha256": figure_sha,
                    "plot": figure_meta,
                    "top_chemicals": top_chemicals,
                },
                display_text=f"Top Interacting Chemicals for {gene.upper()}",
                api_run_id=primary_api,
                raw_artifact_id=primary_raw,
            )
        )

    summary = {
        "gene_symbol": gene,
        "official_symbol": gene.upper(),
        "target_gene_id": expected_entrez,
        "filtered_target_row_count": len(filtered_rows),
        "unique_chemical_count": len(chemicals),
        "top_chemical_count": len(top_chemicals),
        "top_chemicals": [
            {
                "rank": c.get("display_rank"),
                "chemical_name": c.get("chemical_name"),
                "chemical_id": c.get("chemical_id"),
                "interaction_count": c.get("interaction_count"),
            }
            for c in top_chemicals
        ],
        "malformed_source_row_count": filter_audit.get("malformed_source_row_count"),
        "target_row_validation_errors": filter_audit.get("target_row_validation_errors"),
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
        "target_mismatch_detail": summary_mismatch_detail,
        "scientific_intro": SCIENTIFIC_INTRO,
        "scientific_caveat": SCIENTIFIC_CAVEAT,
        "supplementary_xlsx": xlsx_name if workbook_sha else None,
        "supplementary_xlsx_sha256": workbook_sha,
        "top_chemicals_figure_relative_path": (
            f"figures/{gene.upper()}_CTD_top_chemicals.png" if figure_sha else None
        ),
        "top_chemicals_figure_local_path": str(figure_path) if figure_path else None,
        "top_chemicals_figure_sha256": figure_sha,
        "figure_meta": figure_meta,
        "presentation_item_key": f"ctd-{gene.lower()}",
        "ctd_source": {
            "official_url": (bulk.official_url if bulk else None) or OFFICIAL_URL,
            "sha256": bulk.sha256 if bulk else None,
            "byte_size": bulk.byte_size if bulk else None,
            "origin": bulk.origin if bulk else None,
            "api_run_id": primary_api,
            "raw_artifact_id": primary_raw,
            "source_attempt_id": bulk.source_attempt_id if bulk else None,
            "ctd_report_created": bulk.ctd_report_created if bulk else None,
            "extra_columns": (bulk.extra_columns if bulk else None)
            or filter_audit.get("extra_columns"),
        },
        "warnings": warnings,
    }
    audit = {
        "gene_attempt_dir": str(attempt_dir),
        "parser_version": PARSER_VERSION,
        "malformed_source_row_count": filter_audit.get("malformed_source_row_count"),
        "filtered_target_row_count": len(filtered_rows),
        "target_row_validation_errors": filter_audit.get("target_row_validation_errors"),
        "filter_audit": filter_audit,
        "ctd_source": summary["ctd_source"],
        "artifacts": {
            "summary.json": "summary.json",
            "filter_audit.json": "filter_audit.json",
            "manifest.json": MANIFEST_FILENAME,
            f"supplementary/{xlsx_name}": (
                f"supplementary/{xlsx_name}" if workbook_sha else None
            ),
            "top_chemicals_figure": summary.get("top_chemicals_figure_relative_path"),
        },
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
    }
    status = {
        "section_key": "6a",
        "summary": summary,
        "rendering_status": {
            "scientific_status": scientific,
            "visual_status": visual,
            "presentation_status": presentation,
        },
        "audit": audit,
    }
    write_json_atomic(attempt_dir / "summary.json", summary)
    write_json_atomic(attempt_dir / "section_6a_status.json", status)
    write_json_atomic(
        attempt_dir / MANIFEST_FILENAME,
        {
            "gene_symbol": gene,
            "parser_version": PARSER_VERSION,
            "artifacts": audit["artifacts"],
            "counts": {
                "filtered_target_row_count": len(filtered_rows),
                "unique_chemical_count": len(chemicals),
                "malformed_source_row_count": filter_audit.get(
                    "malformed_source_row_count"
                ),
            },
        },
    )

    return {
        **state,
        "api_runs": api_runs,
        "raw_artifacts": raw_artifacts,
        "tool_results": tool_results,
        "evidence_records": evidence_records,
        "section_6a_status": status,
        "warnings": list(state.get("warnings") or []) + warnings,
    }


__all__ = [
    "PARSER_VERSION",
    "SECTION_CTD",
    "SUBSECTION_6A",
    "SCIENTIFIC_INTRO",
    "SCIENTIFIC_CAVEAT",
    "STATUS_FAILED",
    "STATUS_NO_INTERACTIONS",
    "STATUS_NOT_ATTEMPTED_NO_INTERACTIONS",
    "STATUS_PARTIAL",
    "STATUS_SOURCE_UNAVAILABLE",
    "STATUS_SUCCESS",
    "STATUS_TARGET_MISMATCH",
    "STATUS_UNAVAILABLE",
    "Section6aConfig",
    "accept_section_6a_report",
    "aggregate_by_chemical_id",
    "download_persist_and_pin_ctd_bulk",
    "evaluate_section_6a_complete",
    "filter_target_rows",
    "node_generate_section_6a_derived_artifacts",
    "pubmed_occurrence_count",
    "rank_top_chemicals",
    "render_top_chemicals_png",
    "validate_top_chemicals_png",
    "write_ctd_workbook",
]
