"""Section 7a: chemical tools and small-molecule effects (opt-in).

Frozen scientific rules live in the approved plan. This module owns ChEMBL,
DrugBank, PubTator3/PubMed, PubChem, and NCATS Inxight acquisition for the
section-bundle path only.
"""

from __future__ import annotations

import json
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from gene_dossier.config import Settings, get_settings
from gene_dossier.models import (
    AssertionType,
    EvidenceGrade,
    EvidenceRecord,
    SourceType,
    ToolResult,
)
from gene_dossier.section_7a_sources import (
    MANIFEST_FILENAME,
    accept_gene_report,
    paths_for,
    sha256_bytes,
    sha256_file,
    write_json_atomic,
)
from gene_dossier.source_ids import make_source_id
from gene_dossier.tools import chembl as chembl_client
from gene_dossier.tools import drugbank as drugbank_client
from gene_dossier.tools import ncats_inxight as ncats_client
from gene_dossier.tools import pubchem as pubchem_client
from gene_dossier.tools import pubmed as pubmed_client
from gene_dossier.tools import pubtator3 as pubtator_client
from gene_dossier.workflow import WorkflowTransientContext

logger = logging.getLogger(__name__)

SECTION_CHEMICAL = "Chemical tools"
SUBSECTION_7A = "Queried in the following databases to identify small molecule inhibitors"
PARSER_VERSION = "section_7a_v1"

STATUS_SUCCESS = "success"
STATUS_NO_RESULTS = "no_results"
STATUS_NO_AUTHORITATIVE_TARGET = "no_authoritative_target"
STATUS_UNAVAILABLE = "unavailable_not_configured"
STATUS_SOURCE_ERROR = "source_error"
STATUS_LIMITATIONS = "success_with_source_limitations"
STATUS_FAILED = "failed"
STATUS_PARTIAL = "partial"
STATUS_SUCCESS_NO_TOOLS = "success_no_tools"

POLISHED_PUBMED_CAP = 7
POLISHED_PUBCHEM_CAP = 6
EXPANSION_BATCH = 25
EXPANSION_LIMIT = 100
CHEMBL_ACTIVITY_PAGE_LIMIT = 1000
CHEMBL_ASSAY_PAGE_LIMIT = 100

INHIBIT_TOKENS = frozenset(
    {
        "inhibit",
        "inhibits",
        "inhibited",
        "inhibiting",
        "inhibition",
        "inhibitor",
        "inhibitors",
        "suppress",
        "suppresses",
        "suppressed",
        "reduce",
        "reduces",
        "reduced",
        "decrease",
        "decreases",
        "decreased",
        "block",
        "blocks",
        "blocked",
        "impair",
        "impairs",
        "impaired",
        "downregulate",
        "downregulates",
        "downregulated",
        "down-regulation",
        "downregulation",
    }
)
ACTIVATE_TOKENS = frozenset(
    {
        "activate",
        "activates",
        "activated",
        "activation",
        "stimulate",
        "stimulates",
        "stimulated",
        "increase",
        "increases",
        "increased",
        "enhance",
        "enhances",
        "enhanced",
        "upregulate",
        "upregulates",
        "upregulated",
        "up-regulation",
        "upregulation",
        "agonist",
        "agonists",
    }
)
TOOL_CUE_TOKENS = frozenset(
    {
        "inhibitor",
        "inhibitors",
        "inhibition",
        "inhibits",
        "small molecule",
        "compound",
        "compounds",
        "treated with",
        "treatment with",
        "administered",
        "exposure to",
        "pharmacologic",
        "pharmacological",
        "drug",
        "drugs",
        "suppresses",
        "blocks",
        "agonist",
        "antagonist",
    }
)
PERTURBATION_QUERY_TERMS = (
    "inhibitor",
    "inhibition",
    "agonist",
    "antagonist",
    '"small molecule"',
    "compound",
    "treatment",
    "drug",
)

RELATION_PRIORITY = {
    "negative_correlate": 0,
    "positive_correlate": 1,
    "interact": 2,
}


@dataclass
class Section7aConfig:
    output_root: Path | str | None = None
    expansion_batch_size: int = EXPANSION_BATCH
    expansion_limit: int = EXPANSION_LIMIT
    polished_pubmed_cap: int = POLISHED_PUBMED_CAP
    pubmed_fallback_retmax: int = 50


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _as_int(value: Any) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _norm(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip())


def _casefold(text: Any) -> str:
    return _norm(text).casefold()


def _tokenize_sentences(text: str) -> list[str]:
    parts = re.split(r"(?<=[.!?])\s+", _norm(text))
    return [p for p in parts if p]


def _persist_tool(
    tr: ToolResult,
    *,
    dossier_run_id: str,
    gene_symbol: str,
    settings: Settings,
    persist_db: bool,
    api_runs: list,
    raw_artifacts: list,
    tool_results: list,
    filename_hint: str,
) -> tuple[str | None, str | None]:
    from gene_dossier.section_1c import (
        _persist_artifact_bytes,
        _save_api_run_failure,
        _tool_result_to_api_run,
    )

    tool_results.append(tr)
    api = _tool_result_to_api_run(
        tr, dossier_run_id=dossier_run_id, gene_symbol=gene_symbol
    )
    if not tr.success:
        _save_api_run_failure(api, persist_db=persist_db)
        api_runs.append(api)
        return str(getattr(api, "id", "") or "") or None, None
    api_runs.append(api)
    content = json.dumps(tr.data, sort_keys=True, default=str).encode("utf-8")
    try:
        _artifact, meta = _persist_artifact_bytes(
            dossier_run_id=dossier_run_id,
            source_name=tr.source_name,
            content=content,
            extension="json",
            artifact_type="json",
            filename_hint=filename_hint,
            settings=settings,
            api_run=api,
            persist_db=persist_db,
            notes={
                "artifact_class": "external_raw",
                "artifact_origin": "section_7a",
                "endpoint_name": tr.endpoint_name,
                "request_url": tr.request_url,
            },
            validate=lambda b: {"media_type": "application/json", "byte_size": len(b)},
        )
        raw_artifacts.append(meta)
        return str(getattr(api, "id", "") or "") or None, str((meta or {}).get("id") or "") or None
    except Exception as exc:  # noqa: BLE001
        logger.warning("section_7a persist failed: %s", exc)
        return str(getattr(api, "id", "") or "") or None, None


def _protein_name_from_state(state: dict[str, Any]) -> str | None:
    gene_ids = state.get("gene_ids") or {}
    for key in ("protein_name", "preferred_name"):
        val = gene_ids.get(key)
        if val:
            return _norm(val)
    for rec in state.get("evidence_records") or []:
        if getattr(rec, "assertion_type", None) != AssertionType.gene_identity:
            continue
        if str(getattr(rec, "source_name", "")).casefold() != "uniprot":
            continue
        value = getattr(rec, "value", None) or {}
        if not isinstance(value, dict):
            continue
        tax = _as_int(value.get("taxon_id") or value.get("tax_id") or getattr(rec, "taxon_id", None))
        if tax not in (None, 9606):
            continue
        name = value.get("protein_name")
        if name:
            return _norm(name)
    return None


def _uniprot_from_state(state: dict[str, Any]) -> str:
    gene_ids = state.get("gene_ids") or {}
    uniprot = str(
        gene_ids.get("uniprot_accession")
        or gene_ids.get("uniprot")
        or (gene_ids.get("human") or {}).get("uniprot_accession")
        or ""
    ).strip()
    if uniprot:
        return uniprot
    for rec in state.get("evidence_records") or []:
        if getattr(rec, "fact_type", None) != "uniprot_accession":
            continue
        value = getattr(rec, "value", None) or {}
        if not isinstance(value, dict):
            continue
        tax = _as_int(value.get("taxon_id") or value.get("tax_id") or getattr(rec, "taxon_id", None))
        if tax not in (None, 9606):
            continue
        acc = value.get("uniprot_accession") or value.get("primaryAccession")
        if acc:
            return str(acc).strip()
    return ""


def resolve_aliases_from_state(state: dict[str, Any]) -> dict[str, Any]:
    """Collect PubMed-query aliases from authoritative identity evidence only."""
    gene_ids = state.get("gene_ids") or {}
    canonical = str(
        gene_ids.get("official_symbol")
        or gene_ids.get("symbol")
        or state.get("gene_symbol")
        or ""
    ).strip()
    aliases: list[dict[str, Any]] = []
    seen: set[str] = set()

    def _add(alias: str, source: str, source_id: str | None) -> None:
        text = _norm(alias)
        if not text:
            return
        key = text.casefold()
        if key in seen or key == canonical.casefold():
            return
        seen.add(key)
        aliases.append(
            {
                "alias": text,
                "authoritative_source": source,
                "source_identifier": source_id,
                "normalized_canonical_symbol": canonical,
            }
        )

    for rec in state.get("evidence_records") or []:
        if getattr(rec, "assertion_type", None) != AssertionType.gene_identity:
            continue
        value = getattr(rec, "value", None) or {}
        if not isinstance(value, dict):
            continue
        source_name = str(getattr(rec, "source_name", "") or "")
        source_id = str(getattr(rec, "source_id", "") or getattr(rec, "id", "") or "")
        tax = _as_int(value.get("taxon_id") or value.get("taxid") or getattr(rec, "taxon_id", None))
        if tax not in (None, 9606):
            continue
        for key in ("otheraliases", "aliases", "gene_synonyms", "synonyms"):
            raw = value.get(key)
            if isinstance(raw, str):
                parts = [p.strip() for p in re.split(r"[;,|]", raw) if p.strip()]
            elif isinstance(raw, list):
                parts = [_norm(p) for p in raw if _norm(p)]
            else:
                parts = []
            for part in parts:
                _add(part, source_name or "gene_identity", source_id)

    # Optional structured gene_ids.aliases when already populated by identity node.
    for key in ("aliases", "otheraliases", "human_aliases"):
        raw = gene_ids.get(key)
        if isinstance(raw, list):
            for part in raw:
                _add(str(part), "gene_ids", key)
        elif isinstance(raw, str):
            for part in re.split(r"[;,|]", raw):
                _add(part, "gene_ids", key)

    if aliases:
        return {
            "canonical_symbol": canonical,
            "aliases": aliases,
            "alias_resolution_status": "resolved",
            "limitation": None,
        }
    return {
        "canonical_symbol": canonical,
        "aliases": [],
        "alias_resolution_status": "canonical_only",
        "limitation": "authoritative_aliases_unavailable",
    }


def build_pubmed_perturbation_term(alias_payload: dict[str, Any]) -> str:
    canonical = alias_payload.get("canonical_symbol") or ""
    names = [canonical] + [a["alias"] for a in alias_payload.get("aliases") or []]
    gene_clause = " OR ".join(
        f'{n}[Title/Abstract]' if " " in n or "-" in n else f"{n}[Title/Abstract]"
        for n in names
        if n
    )
    # Quote multi-token aliases
    parts = []
    for n in names:
        if not n:
            continue
        if re.search(r"[\s\-]", n):
            parts.append(f'"{n}"[Title/Abstract]')
        else:
            parts.append(f"{n}[Title/Abstract]")
    gene_clause = " OR ".join(parts) if parts else f"{canonical}[Title/Abstract]"
    chem_clause = " OR ".join(f"{t}[Title/Abstract]" for t in PERTURBATION_QUERY_TERMS)
    return f"({gene_clause}) AND ({chem_clause})"


def classify_tool_eligibility(text: str, chemical_name: str) -> str:
    """Classify tool eligibility using only the chemical's local sentence windows."""
    chem = _casefold(chemical_name)
    endogenous = any(
        x in chem
        for x in (
            "cholesterol",
            "sterol",
            "sterols",
            "hydroxycholesterol",
            "fatty acid",
            "fatty acids",
            "lipid",
            "lipids",
        )
    )
    sentences = _tokenize_sentences(text)
    local_sents = _chemical_local_windows(sentences, chem)
    tool_hit = False
    explicit = False
    chem_targeted_intervention = False
    for sent in local_sents:
        s = sent.casefold()
        # Endogenous chemicals need cues that treat THIS chemical as the intervention.
        if chem and (
            f"treated with {chem}" in s
            or f"treatment with {chem}" in s
            or f"{chem} treatment" in s
            or f"administered {chem}" in s
            or f"{chem} was administered" in s
            or f"exposure to {chem}" in s
            or f"{chem} as a probe" in s
            or f"{chem} as a small molecule" in s
            or re.search(
                rf"\b(treated with|treatment with|administered|exposure to)\s+{re.escape(chem)}\b",
                s,
            )
            or re.search(
                rf"\b{re.escape(chem)}\b\s*\((?:small molecule|compound|inhibitor|agonist|antagonist|probe)\)",
                s,
            )
        ):
            chem_targeted_intervention = True
        for cue in TOOL_CUE_TOKENS:
            if cue in s:
                tool_hit = True
                if any(
                    t in s
                    for t in (
                        "inhibitor",
                        "small molecule",
                        "compound",
                        "drug",
                        "agonist",
                        "antagonist",
                    )
                ):
                    explicit = True
    if endogenous and not chem_targeted_intervention:
        return "endogenous_or_contextual_chemical"
    if explicit:
        return "explicit_chemical_tool"
    if tool_hit:
        return "perturbational_chemical"
    if endogenous:
        return "endogenous_or_contextual_chemical"
    return "insufficient_tool_evidence"


def _chemical_mentioned(sentence: str, chem: str) -> bool:
    if not chem:
        return False
    s = sentence.casefold()
    if chem in s:
        return True
    return any(tok in s for tok in chem.split() if len(tok) > 3)


def _chemical_local_windows(sentences: list[str], chem: str) -> list[str]:
    """Return sentences that mention the chemical (local tool-eligibility windows)."""
    out: list[str] = []
    for sent in sentences:
        if _chemical_mentioned(sent, chem):
            out.append(sent)
    return out


def _regulator_hit(text: str) -> str | None:
    m = re.search(
        r"\b(g9a|ehmt2|scap|s1p|pcsk9|hmgcr|histone methyltransferase)\b",
        text,
        re.I,
    )
    return m.group(1) if m else None


def _passage_snippets_for_chemical(document: dict[str, Any], chemical_name: str) -> str:
    """Collect BioC passage text that mentions ``chemical_name`` (audit/backstop)."""
    chem = _casefold(chemical_name)
    if not chem:
        return ""
    snippets: list[str] = []
    for passage in document.get("passages") or []:
        if not isinstance(passage, dict):
            continue
        text = str(passage.get("text") or "")
        if chem in text.casefold():
            snippets.append(text)
    return "\n".join(snippets)


def extract_explicit_chemical_mentions(title: str, abstract: str) -> list[dict[str, Any]]:
    """Ground chemicals from explicit title/abstract tokens when annotations are empty.

    Deterministic patterns only (no gene-specific hardcodes): alphanumerics with
    digits (e.g. tool-like IDs) and ``X inhibitor`` noun phrases.
    """
    text = f"{title or ''}\n{abstract or ''}"
    found: list[dict[str, Any]] = []
    seen: set[str] = set()

    def _add(name: str, *, source: str) -> None:
        cleaned = _norm(name)
        if not cleaned or len(cleaned) < 3:
            return
        key = cleaned.casefold()
        if key in seen:
            return
        seen.add(key)
        found.append({"name": cleaned, "text": cleaned, "id": None, "source": source})

    for m in re.finditer(r"\b([A-Za-z][A-Za-z0-9-]{1,12}\d{2,}[A-Za-z0-9-]*)\b", text):
        _add(m.group(1), source="title_abstract_token")
    for m in re.finditer(
        r"\b([A-Za-z][A-Za-z0-9-]{2,40})\s+inhibitor\b",
        text,
        re.I,
    ):
        _add(m.group(1), source="title_abstract_inhibitor_phrase")
    return found


def _proximity_hit(text: str, left: str, right_tokens: set[str], window: int = 80) -> bool:
    """True when ``left`` appears within ``window`` chars of any token in ``right_tokens``."""
    if not left:
        return False
    t = text.casefold()
    for m in re.finditer(re.escape(left), t):
        start = max(0, m.start() - window)
        end = min(len(t), m.end() + window)
        chunk = t[start:end]
        if any(tok in chunk for tok in right_tokens):
            return True
    return False


def classify_evidence_span(
    *,
    title: str,
    abstract: str,
    chemical_name: str,
    gene_symbol: str,
    gene_aliases: list[str] | None = None,
) -> dict[str, Any]:
    """Local-window classifier. Never uses document-wide co-occurrence."""
    gene_names = {_casefold(gene_symbol)}
    for a in gene_aliases or []:
        if a:
            gene_names.add(_casefold(a))
    chem = _casefold(chemical_name)
    # Never treat the dossier gene symbol/aliases as a chemical tool candidate.
    if chem and chem in gene_names:
        return {
            "evidence_class": "insufficient_effect_detail",
            "tool_eligibility": "insufficient_tool_evidence",
            "spans": [],
            "sentence_count": 0,
            "supporting_span_text": "",
            "supporting_sentence_indices": [],
            "effect_target": None,
            "effect_direction": None,
            "gene_symbol": gene_symbol,
            "evidence_scope": "local_window",
            "exclusion_reason": "chemical_matches_dossier_gene",
        }
    sentences = _tokenize_sentences(f"{title}. {abstract}" if title else abstract)
    spans: list[dict[str, Any]] = []
    evidence_class = "insufficient_effect_detail"
    effect_target: str | None = None
    effect_direction: str | None = None
    supporting_indices: list[int] = []

    def _has_gene(text: str) -> bool:
        return any(g in text for g in gene_names if len(g) >= 3)

    for idx, sent in enumerate(sentences):
        s = sent.casefold()
        has_chem = _chemical_mentioned(sent, chem)
        if not has_chem:
            continue
        # Ignore extremely long figure/table blobs for effect claims.
        if len(sent) > 420:
            continue
        # Imaging counterstain / legend annotations are not chemical-tool evidence.
        if re.search(
            rf"(\(\s*{re.escape(chem)}\s*=|{re.escape(chem)}\s*=\s*(blue|green|red|white|yellow))",
            s,
        ):
            continue
        has_gene = _has_gene(s)
        has_inhibit = _proximity_hit(s, chem, INHIBIT_TOKENS, window=48)
        has_activate = _proximity_hit(s, chem, ACTIVATE_TOKENS, window=48)
        adj = sentences[idx + 1] if idx + 1 < len(sentences) else ""
        prev = sentences[idx - 1] if idx else ""
        adj_cf = adj.casefold()
        prev_cf = prev.casefold()
        nearby_gene = has_gene or (
            len(adj) <= 420 and _has_gene(adj_cf)
        ) or (
            len(prev) <= 420 and _has_gene(prev_cf)
        )
        regulator = _regulator_hit(s) or (
            _regulator_hit(adj_cf) if len(adj) <= 420 else None
        ) or (
            _regulator_hit(prev_cf) if len(prev) <= 420 else None
        )
        window_parts = [sent]
        window_indices = [idx]
        if adj and len(adj) <= 420 and (
            (has_chem and has_inhibit and not has_gene and (_has_gene(adj_cf) or _regulator_hit(adj_cf)))
            or (has_chem and regulator and _has_gene(adj_cf))
        ):
            window_parts.append(adj)
            window_indices.append(idx + 1)
        if prev and len(prev) <= 420 and has_chem and not has_gene and (
            _has_gene(prev_cf) or _regulator_hit(prev_cf)
        ):
            if idx - 1 not in window_indices and not adj:
                window_parts.insert(0, prev)
                window_indices.insert(0, idx - 1)

        if has_chem and has_inhibit and regulator and nearby_gene:
            spans.append({"sentence_index": idx, "text": " ".join(window_parts), "role": "indirect"})
            evidence_class = "indirect_pathway_effect"
            effect_target = regulator
            effect_direction = "downstream_rescue_or_modulation"
            supporting_indices = window_indices
            continue
        if has_chem and has_inhibit and not has_gene and nearby_gene:
            spans.append({"sentence_index": idx, "text": " ".join(window_parts), "role": "indirect"})
            evidence_class = "indirect_pathway_effect"
            effect_target = regulator or "upstream_regulator"
            effect_direction = "downstream_effect"
            supporting_indices = window_indices
            continue
        if has_chem and has_gene and has_inhibit:
            spans.append({"sentence_index": idx, "text": sent, "role": "negative"})
            if evidence_class not in {"indirect_pathway_effect"}:
                evidence_class = "literature_negative_effect"
                effect_target = gene_symbol
                effect_direction = "negative"
                supporting_indices = [idx]
            continue
        if has_chem and has_gene and has_activate:
            spans.append({"sentence_index": idx, "text": sent, "role": "positive"})
            if evidence_class not in {"indirect_pathway_effect", "literature_negative_effect"}:
                evidence_class = "literature_positive_effect"
                effect_target = gene_symbol
                effect_direction = "positive"
                supporting_indices = [idx]
            continue
        if has_chem and has_gene:
            spans.append({"sentence_index": idx, "text": sent, "role": "interaction"})
            if evidence_class == "insufficient_effect_detail":
                evidence_class = "literature_interaction"
                effect_target = gene_symbol
                effect_direction = "interaction"
                supporting_indices = [idx]

    supporting_span_text = ""
    if supporting_indices:
        supporting_span_text = " ".join(
            sentences[i] for i in supporting_indices if 0 <= i < len(sentences)
        )
    elif spans:
        supporting_span_text = str(spans[0].get("text") or "")
        supporting_indices = [int(spans[0]["sentence_index"])]

    tool_elig = classify_tool_eligibility(f"{title}\n{abstract}", chemical_name)
    return {
        "evidence_class": evidence_class,
        "tool_eligibility": tool_elig,
        "spans": spans,
        "sentence_count": len(sentences),
        "supporting_span_text": supporting_span_text,
        "supporting_sentence_indices": supporting_indices,
        "effect_target": effect_target,
        "effect_direction": effect_direction,
        "gene_symbol": gene_symbol,
        "evidence_scope": "local_window",
    }


def _display_eligible(entry: dict[str, Any]) -> bool:
    chem = _norm(entry.get("chemical_name") or "")
    if not chem:
        return False
    # Reject tiny tokens / punctuation fragments from noisy full-text tables.
    if len(chem) < 3 and not re.fullmatch(r"[A-Za-z]+\d+[A-Za-z0-9-]*", chem):
        return False
    if re.fullmatch(r"[A-Za-z]{1,2}", chem):
        return False
    if entry.get("tool_eligibility") not in {
        "explicit_chemical_tool",
        "perturbational_chemical",
    }:
        return False
    if entry.get("evidence_class") in {None, "insufficient_effect_detail"}:
        return False
    # Require a local supporting span; never promote document-wide-only hits.
    span = _norm(entry.get("supporting_span_text") or "")
    if not span:
        return False
    if entry.get("evidence_scope") == "document_wide":
        return False
    grounding = str(entry.get("grounding_source") or "")
    # Full-text backstop chemicals need perturbation evidence, not mere co-occurrence.
    if "fulltext" in grounding and entry.get("evidence_class") == "literature_interaction":
        return False
    return True


def rank_literature_candidates(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    tool_rank = {
        "explicit_chemical_tool": 0,
        "perturbational_chemical": 1,
        "endogenous_or_contextual_chemical": 9,
        "insufficient_tool_evidence": 9,
    }
    effect_rank = {
        "literature_negative_effect": 0,
        "literature_positive_effect": 0,
        "indirect_pathway_effect": 0,
        "literature_interaction": 1,
        "insufficient_effect_detail": 3,
    }
    direct_rank = {
        "literature_negative_effect": 0,
        "literature_positive_effect": 0,
        "indirect_pathway_effect": 1,
        "literature_interaction": 2,
        "insufficient_effect_detail": 3,
    }

    def _chem_quality(name: Any) -> int:
        text = _norm(name)
        if not text:
            return 9
        # Prefer compact pharmacological IDs / named tool compounds.
        if re.fullmatch(r"[A-Za-z]{1,6}\d{2,}[A-Za-z0-9-]*", text):
            return 0
        if "/" in text or len(text) <= 2:
            return 8
        if text.casefold() in {"oxygen", "water", "helium", "lipid", "lipids"}:
            return 8
        return 1

    def key(e: dict[str, Any]) -> tuple:
        return (
            tool_rank.get(str(e.get("tool_eligibility")), 9),
            effect_rank.get(str(e.get("evidence_class")), 9),
            direct_rank.get(str(e.get("evidence_class")), 9),
            _chem_quality(e.get("chemical_name")),
            -int(e.get("support_count") or 0),
            _casefold(e.get("chemical_name")),
            str(min(e.get("pmids") or ["99999999"])),
        )

    return sorted(entries, key=key)


def classify_chembl_activity(
    activity: dict[str, Any],
    assay: dict[str, Any] | None = None,
) -> str:
    """Classify ChEMBL activity; directness requires explicit assay relationship."""
    std = _casefold(activity.get("standard_type"))
    comment = _casefold(activity.get("activity_comment") or activity.get("data_validity_comment"))
    action = _casefold(activity.get("action_type") or activity.get("mechanism_of_action"))
    blob = f"{std} {comment} {action}"
    inhibitory = any(
        t in blob for t in ("inhibit", "antagonist", "ic50", "ki", "blocker", "ic 50")
    )
    if chembl_client.is_direct_assay_relationship(assay) and inhibitory:
        return "direct_target_evidence"
    return "target_linked_activity"


def build_literature_effect_prose(entry: dict[str, Any]) -> str:
    """Concise grounded effect prose from validated local span (never title-only)."""
    chem = _norm(entry.get("chemical_name") or "Unknown chemical")
    eclass = str(entry.get("evidence_class") or "")
    class_phrase = {
        "literature_negative_effect": "reported negative/inhibitory effect",
        "literature_positive_effect": "reported positive/activating effect",
        "literature_interaction": "reported chemical–gene interaction",
        "indirect_pathway_effect": "indirect pathway effect",
    }.get(eclass, eclass.replace("_", " "))
    span = _norm(entry.get("supporting_span_text") or "")
    gene = _norm(entry.get("gene_symbol") or "")
    effect_target = _norm(entry.get("effect_target") or "")
    if eclass == "indirect_pathway_effect":
        if span and _casefold(chem) in span.casefold():
            # Bound to one concise span; never dump an abstract.
            clipped = span if len(span) <= 320 else span[:317].rstrip() + "..."
            return f"{chem}: {class_phrase}. {clipped}"
        target = effect_target or "an upstream regulator"
        gene_bit = gene or "the dossier gene"
        return (
            f"{chem}: {class_phrase}. {chem} acts on {target}, with downstream "
            f"effect on {gene_bit}."
        )
    if span and _casefold(chem) in span.casefold():
        clipped = span if len(span) <= 320 else span[:317].rstrip() + "..."
        return f"{chem}: {class_phrase}. {clipped}"
    return f"{chem}: {class_phrase}."


def citation_label_from_esummary(row: dict[str, Any], pmid: str) -> dict[str, Any]:
    """Build deterministic author-year citation metadata from ESummary."""
    authors = row.get("authors") if isinstance(row.get("authors"), list) else []
    names: list[str] = []
    for a in authors:
        if isinstance(a, dict) and a.get("name"):
            names.append(str(a["name"]).strip())
        elif isinstance(a, str) and a.strip():
            names.append(a.strip())
    year = None
    for key in ("pubdate", "epubdate", "sortpubdate"):
        m = re.search(r"(19|20)\d{2}", str(row.get(key) or ""))
        if m:
            year = m.group(0)
            break
    label = None
    if names and year:
        last = names[0].split()[0]
        label = f"{last} et al., {year}" if len(names) > 1 else f"{last}, {year}"
    return {
        "authors": names,
        "year": year,
        "citation_label": label,
        "pmid": str(pmid),
    }


def select_polished_pubchem_rows(
    focused_rows: list[dict[str, Any]],
    *,
    cap: int = POLISHED_PUBCHEM_CAP,
) -> list[dict[str, Any]]:
    """Choose deterministic representative focused assays for PDF presentation."""

    def _group_key(row: dict[str, Any]) -> str:
        for key in ("pmid", "doi", "document_chembl_id", "source_db_id"):
            val = _norm(row.get(key) or "")
            if val:
                return f"{key}:{val.casefold()}"
        return f"aid:{row.get('aid')}"

    def _rank(row: dict[str, Any]) -> tuple:
        reason = _casefold(row.get("reason") or "")
        direct = 0 if "direct" in reason or "protein_accession" in reason else 1
        accession = 0 if row.get("protein_accession") else 1
        recon = 0 if reason else 1
        aid = int(row.get("aid") or 0)
        return (direct, accession, recon, aid)

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in focused_rows:
        groups[_group_key(row)].append(row)
    reps: list[dict[str, Any]] = []
    for key in sorted(groups.keys()):
        best = sorted(groups[key], key=_rank)[0]
        reps.append(dict(best, evidence_group=key, representative=True))
    reps.sort(key=lambda r: int(r.get("aid") or 0))
    return reps[: max(0, int(cap))]


def extract_pubchem_publication_ids(descr: dict[str, Any]) -> dict[str, str]:
    """Extract structured publication identifiers from a PubChem assay descr."""
    out: dict[str, str] = {}
    xrefs = descr.get("xref") or descr.get("db") or []
    if isinstance(xrefs, dict):
        xrefs = [xrefs]
    if not isinstance(xrefs, list):
        xrefs = []
    for xref in xrefs:
        if not isinstance(xref, dict):
            continue
        db = _casefold(xref.get("dbname") or xref.get("db") or xref.get("name") or "")
        key = str(xref.get("key") or xref.get("id") or xref.get("value") or "").strip()
        if not key:
            continue
        if "pubmed" in db or db == "pmid":
            out.setdefault("pmid", key)
        elif "doi" in db:
            out.setdefault("doi", key)
        elif "chembl" in db:
            out.setdefault("document_chembl_id", key)
        else:
            out.setdefault("source_db_id", f"{db}:{key}" if db else key)
    # Common nested aid_source / citation blobs
    for blob_key in ("aid_source", "citation", "source"):
        blob = descr.get(blob_key)
        if not isinstance(blob, dict):
            continue
        for k, dest in (("pmid", "pmid"), ("doi", "doi"), ("db", "source_db_id")):
            if blob.get(k) and dest not in out:
                out[dest] = str(blob.get(k))
    return out


def write_chembl_workbook(
    path: Path,
    *,
    gene: str,
    target: dict[str, Any] | None,
    activities: list[dict[str, Any]],
    assays: list[dict[str, Any]] | None = None,
    provenance: dict[str, Any],
) -> None:
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Activities"
    headers = [
        "molecule_chembl_id",
        "canonical_smiles",
        "standard_type",
        "standard_relation",
        "standard_value",
        "standard_units",
        "pchembl_value",
        "assay_chembl_id",
        "document_chembl_id",
        "evidence_class",
        "relationship_type",
        "confidence_score",
    ]
    ws_a.append(headers)
    for row in activities:
        ws_a.append([row.get(h) for h in headers])

    assay_rows = list(assays or [])
    if not assay_rows:
        # Fall back to IDs from activities only when assay metadata was unavailable.
        assay_rows = [
            {"assay_chembl_id": aid}
            for aid in sorted(
                {r.get("assay_chembl_id") for r in activities if r.get("assay_chembl_id")}
            )
        ]
    ws_assays = wb.create_sheet("Assays")
    assay_headers = [
        "assay_chembl_id",
        "description",
        "assay_type",
        "target_chembl_id",
        "relationship_type",
        "confidence_score",
        "confidence_description",
        "document_chembl_id",
    ]
    ws_assays.append(assay_headers)
    for row in assay_rows:
        ws_assays.append([row.get(h) for h in assay_headers])

    ws_mol = wb.create_sheet("Molecules")
    ws_mol.append(["molecule_chembl_id", "canonical_smiles"])
    seen = set()
    for r in activities:
        mid = r.get("molecule_chembl_id")
        if mid and mid not in seen:
            seen.add(mid)
            ws_mol.append([mid, r.get("canonical_smiles")])

    ws_t = wb.create_sheet("Target")
    ws_t.append(["field", "value"])
    for k, v in (target or {}).items():
        ws_t.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])
    ws_t.append(["gene", gene])

    ws_p = wb.create_sheet("Provenance")
    ws_p.append(["field", "value"])
    for k, v in provenance.items():
        ws_p.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _chembl_page_batch(payload: Any, list_key: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data = payload if isinstance(payload, dict) else {}
    batch = data.get(list_key) or []
    if not isinstance(batch, list):
        batch = []
    rows = [r for r in batch if isinstance(r, dict)]
    page_meta = data.get("page_meta") if isinstance(data.get("page_meta"), dict) else {}
    return rows, page_meta


def ncats_display_text(candidates: list[dict[str, Any]]) -> str:
    """Render NCATS polished line with candidate names and unconfirmed qualification."""
    if not candidates:
        return "NCATS Inxight: Drugs – No results"
    names = [_norm(c.get("name")) for c in candidates if _norm(c.get("name"))]
    name_bit = ", ".join(names) if names else f"{len(candidates)} candidate(s)"
    unconfirmed = all(
        c.get("evidence_class") == "facet_target_match_unconfirmed" for c in candidates
    )
    if unconfirmed:
        return (
            f"NCATS Inxight: Drugs – {name_bit}. Candidate target-facet match; "
            "direct target relationship was not independently confirmed."
        )
    return f"NCATS Inxight: Drugs – {name_bit}"


def _evidence(
    *,
    dossier_run_id: str,
    gene_symbol: str,
    source_name: str,
    fact_type: str,
    key: str,
    value: dict[str, Any],
    display_text: str,
    api_run_id: str | None = None,
    raw_artifact_id: str | None = None,
    confidence_notes: str | None = None,
) -> EvidenceRecord:
    return EvidenceRecord(
        source_id=make_source_id(
            source_name, gene_symbol, AssertionType.chemical_tool, key
        ),
        dossier_run_id=dossier_run_id,
        gene_symbol=gene_symbol,
        official_symbol=gene_symbol,
        section=SECTION_CHEMICAL,
        subsection=SUBSECTION_7A,
        source_name=source_name,
        source_type=SourceType.chemical_database,
        assertion_type=AssertionType.chemical_tool,
        fact_type=fact_type,
        organism="Homo sapiens",
        taxon_id=9606,
        evidence_grade=EvidenceGrade.C,
        manual_review_required=False,
        confidence_notes=confidence_notes,
        value=value,
        display_text=display_text,
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )


def derive_literature_status(pubtator_status: str, pubmed_status: str) -> str:
    if pubtator_status == STATUS_SUCCESS and pubmed_status in {STATUS_SUCCESS, STATUS_NO_RESULTS}:
        # usable evidence already folded into pubmed_status success
        if pubmed_status == STATUS_SUCCESS:
            return STATUS_SUCCESS
    if pubmed_status == STATUS_SUCCESS:
        if pubtator_status in {STATUS_NO_RESULTS, STATUS_SUCCESS}:
            return STATUS_SUCCESS
        if pubtator_status == STATUS_SOURCE_ERROR:
            return STATUS_LIMITATIONS
    if pubtator_status == STATUS_SUCCESS and pubmed_status == STATUS_NO_RESULTS:
        return STATUS_NO_RESULTS
    if pubtator_status == STATUS_NO_RESULTS and pubmed_status == STATUS_NO_RESULTS:
        return STATUS_NO_RESULTS
    if pubtator_status == STATUS_SOURCE_ERROR and pubmed_status == STATUS_SOURCE_ERROR:
        return STATUS_SOURCE_ERROR
    if pubmed_status == STATUS_SUCCESS:
        return STATUS_LIMITATIONS
    return pubmed_status or pubtator_status or STATUS_NO_RESULTS


def node_generate_section_7a_derived_artifacts(
    state: dict[str, Any],
    *,
    settings: Settings | None = None,
    persist_db: bool = True,
    transient: WorkflowTransientContext | None = None,
    config: Section7aConfig | None = None,
) -> dict[str, Any]:
    from gene_dossier.section_1c import _append_evidence

    cfg = settings or get_settings()
    section_cfg = config or Section7aConfig()
    run_type = state.get("run_type")
    selected_keys = list(state.get("selected_section_keys") or [])
    if run_type != "section_bundle" or "7a" not in selected_keys:
        return state

    gene = str(state.get("gene_symbol") or "").strip()
    gene_ids = dict(state.get("gene_ids") or {})
    if not gene:
        gene = str(gene_ids.get("symbol") or gene_ids.get("official_symbol") or "").strip()
    if not gene:
        errors = list(state.get("errors") or [])
        errors.append("Section 7a requires a resolved gene_symbol")
        return {**state, "errors": errors}

    entrez = _as_int(
        gene_ids.get("entrez_gene_id")
        or gene_ids.get("entrez_id")
        or gene_ids.get("ncbi_gene_id")
        or (gene_ids.get("human") or {}).get("entrez_gene_id")
    )

    run_id = str(state.get("dossier_run_id") or "")
    api_runs = list(state.get("api_runs") or [])
    raw_artifacts = list(state.get("raw_artifacts") or [])
    tool_results = list(state.get("tool_results") or [])
    evidence_records = list(state.get("evidence_records") or [])
    uniprot = _uniprot_from_state(
        {**state, "gene_ids": gene_ids, "evidence_records": evidence_records}
    )

    paths = paths_for(section_cfg.output_root or cfg.output_path)
    attempt_dir = paths.new_gene_attempt(gene, run_id=run_id or None)
    (attempt_dir / "supplementary").mkdir(exist_ok=True)
    (attempt_dir / "derived").mkdir(exist_ok=True)

    alias_payload = resolve_aliases_from_state({**state, "gene_symbol": gene, "gene_ids": gene_ids})
    alias_names = [a["alias"] for a in alias_payload.get("aliases") or []]

    source_blocks: dict[str, Any] = {}
    warnings: list[str] = []

    # ---- DrugBank ----
    db_tr = drugbank_client.fetch_status(gene, settings=cfg)
    _persist_tool(
        db_tr,
        dossier_run_id=run_id,
        gene_symbol=gene,
        settings=cfg,
        persist_db=persist_db,
        api_runs=api_runs,
        raw_artifacts=raw_artifacts,
        tool_results=tool_results,
        filename_hint="drugbank-status",
    )
    db_unavailable = db_tr.error_type in {
        "unavailable_not_configured",
        "unavailable_not_implemented",
    }
    source_blocks["drugbank"] = {
        "source_status": STATUS_UNAVAILABLE if db_unavailable else STATUS_SOURCE_ERROR,
        "display": "DrugBank – API access unavailable",
        "error_type": db_tr.error_type,
    }

    # ---- ChEMBL ----
    chembl_block: dict[str, Any] = {
        "source_status": STATUS_NO_AUTHORITATIVE_TARGET,
        "target_chembl_id": None,
        "activity_count": 0,
        "activities": [],
        "workbook": None,
        "workbook_sha256": None,
    }
    chembl_api = None
    if not uniprot:
        chembl_block["source_status"] = STATUS_NO_AUTHORITATIVE_TARGET
        chembl_block["detail"] = "missing_uniprot_accession"
    else:
        # search by uniprot then symbol
        t_uni = chembl_client.target_search(uniprot, settings=cfg)
        chembl_api, _ = _persist_tool(
            t_uni,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="chembl-target-uniprot",
        )
        targets: list[Any] = []
        if t_uni.success and isinstance(t_uni.data, dict):
            targets = list(t_uni.data.get("targets") or [])
        if not targets:
            t_sym = chembl_client.target_search(gene, settings=cfg)
            chembl_api, _ = _persist_tool(
                t_sym,
                dossier_run_id=run_id,
                gene_symbol=gene,
                settings=cfg,
                persist_db=persist_db,
                api_runs=api_runs,
                raw_artifacts=raw_artifacts,
                tool_results=tool_results,
                filename_hint="chembl-target-symbol",
            )
            if t_sym.success and isinstance(t_sym.data, dict):
                targets = list(t_sym.data.get("targets") or [])
        tid, method, detail = chembl_client.resolve_authoritative_target(
            targets, uniprot_accession=uniprot, gene_symbol=gene
        )
        chembl_block["resolution_method"] = method
        chembl_block["resolution_detail"] = detail
        if not tid:
            chembl_block["source_status"] = STATUS_NO_AUTHORITATIVE_TARGET
            chembl_block["display"] = "ChEMBL – No authoritative target"
        else:
            chembl_block["target_chembl_id"] = tid
            chembl_block["target"] = {
                "target_chembl_id": tid,
                "uniprot_accession": uniprot,
                "gene_symbol": gene,
                "resolution_method": method,
            }
            activity_page_api_run_ids: list[str] = []
            activity_page_raw_ids: list[str] = []
            acts_raw: list[dict[str, Any]] = []
            offset = 0
            page_idx = 0
            act_error = False
            while page_idx < 50:
                act_tr = chembl_client.activities_by_target(
                    tid,
                    gene_symbol=gene,
                    limit=CHEMBL_ACTIVITY_PAGE_LIMIT,
                    offset=offset,
                    settings=cfg,
                )
                api_id, raw_id = _persist_tool(
                    act_tr,
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    settings=cfg,
                    persist_db=persist_db,
                    api_runs=api_runs,
                    raw_artifacts=raw_artifacts,
                    tool_results=tool_results,
                    filename_hint=f"chembl-activities-p{page_idx}",
                )
                if api_id:
                    activity_page_api_run_ids.append(api_id)
                    chembl_api = api_id
                if raw_id:
                    activity_page_raw_ids.append(raw_id)
                if not act_tr.success:
                    act_error = True
                    break
                batch, page_meta = _chembl_page_batch(act_tr.data, "activities")
                acts_raw.extend(batch)
                if not batch:
                    break
                offset += len(batch)
                total_count = page_meta.get("total_count")
                try:
                    total_i = int(total_count) if total_count is not None else None
                except (TypeError, ValueError):
                    total_i = None
                if total_i is not None and offset >= total_i:
                    break
                if page_meta.get("next") in (None, "", False):
                    break
                page_idx += 1

            assay_page_api_run_ids: list[str] = []
            assay_page_raw_ids: list[str] = []
            assays_raw: list[dict[str, Any]] = []
            assay_map: dict[str, dict[str, Any]] = {}
            if not act_error:
                a_offset = 0
                a_page = 0
                while a_page < 50:
                    assay_tr = chembl_client.assays_by_target(
                        tid,
                        gene_symbol=gene,
                        limit=CHEMBL_ASSAY_PAGE_LIMIT,
                        offset=a_offset,
                        settings=cfg,
                    )
                    api_id, raw_id = _persist_tool(
                        assay_tr,
                        dossier_run_id=run_id,
                        gene_symbol=gene,
                        settings=cfg,
                        persist_db=persist_db,
                        api_runs=api_runs,
                        raw_artifacts=raw_artifacts,
                        tool_results=tool_results,
                        filename_hint=f"chembl-assays-p{a_page}",
                    )
                    if api_id:
                        assay_page_api_run_ids.append(api_id)
                    if raw_id:
                        assay_page_raw_ids.append(raw_id)
                    if not assay_tr.success:
                        warnings.append("chembl_assay_metadata_incomplete")
                        break
                    batch, page_meta = _chembl_page_batch(assay_tr.data, "assays")
                    for row in batch:
                        aid = str(row.get("assay_chembl_id") or "")
                        if aid:
                            assay_map[aid] = chembl_client.summarize_assay(row)
                        assays_raw.append(row)
                    if not batch:
                        break
                    a_offset += len(batch)
                    total_count = page_meta.get("total_count")
                    try:
                        total_i = int(total_count) if total_count is not None else None
                    except (TypeError, ValueError):
                        total_i = None
                    if total_i is not None and a_offset >= total_i:
                        break
                    if page_meta.get("next") in (None, "", False):
                        break
                    a_page += 1

            chembl_block["activity_page_api_run_ids"] = activity_page_api_run_ids
            chembl_block["activity_page_raw_artifact_ids"] = activity_page_raw_ids
            chembl_block["assay_page_api_run_ids"] = assay_page_api_run_ids
            chembl_block["assay_page_raw_artifact_ids"] = assay_page_raw_ids
            chembl_block["activity_pages_fetched"] = len(activity_page_api_run_ids)
            chembl_block["assay_pages_fetched"] = len(assay_page_api_run_ids)

            if act_error:
                chembl_block["source_status"] = STATUS_SOURCE_ERROR
            else:
                classified = []
                assay_summaries = list(assay_map.values())
                for a in acts_raw:
                    if a.get("target_chembl_id") and str(a.get("target_chembl_id")) != str(tid):
                        continue
                    assay_id = str(a.get("assay_chembl_id") or "")
                    assay_meta = assay_map.get(assay_id) or {}
                    row = chembl_client.summarize_activity(a)
                    row["target_chembl_id"] = tid
                    row["relationship_type"] = assay_meta.get("relationship_type")
                    row["confidence_score"] = assay_meta.get("confidence_score")
                    row["confidence_description"] = assay_meta.get(
                        "confidence_description"
                    )
                    row["evidence_class"] = classify_chembl_activity(a, assay_meta)
                    classified.append(row)
                chembl_block["activities"] = classified
                chembl_block["assays"] = assay_summaries
                chembl_block["activity_count"] = len(classified)
                chembl_block["assay_count"] = len(assay_summaries)
                chembl_block["source_status"] = (
                    STATUS_SUCCESS if classified else STATUS_NO_RESULTS
                )
                if classified:
                    xlsx_name = f"{gene.upper()}_Chembl_Inhibitor.xlsx"
                    xlsx_path = attempt_dir / "supplementary" / xlsx_name
                    write_chembl_workbook(
                        xlsx_path,
                        gene=gene,
                        target=chembl_block["target"],
                        activities=classified,
                        assays=assay_summaries,
                        provenance={
                            "parser_version": PARSER_VERSION,
                            "activity_page_api_run_ids": activity_page_api_run_ids,
                            "activity_page_raw_artifact_ids": activity_page_raw_ids,
                            "assay_page_api_run_ids": assay_page_api_run_ids,
                            "assay_page_raw_artifact_ids": assay_page_raw_ids,
                            "official_url": "https://www.ebi.ac.uk/chembl/",
                            "retrieval_timestamp": _utc_now(),
                        },
                    )
                    # hash AFTER finalize/close
                    chembl_block["workbook"] = xlsx_name
                    chembl_block["workbook_path"] = str(xlsx_path)
                    chembl_block["workbook_sha256"] = sha256_file(xlsx_path)
                    chembl_block["display"] = (
                        f"ChEMBL – see Supplementary Material ({xlsx_name})"
                    )
                else:
                    chembl_block["display"] = "ChEMBL – No results"

    source_blocks["chembl"] = chembl_block
    chembl_molecule_names = {
        _casefold(a.get("molecule_chembl_id"))
        for a in chembl_block.get("activities") or []
        if a.get("molecule_chembl_id")
    }

    # ---- PubChem ----
    pubchem_block: dict[str, Any] = {
        "source_status": STATUS_NO_RESULTS,
        "aid_count": 0,
        "focused": [],
        "polished_focused": [],
        "excluded": [],
    }
    if entrez is None:
        pubchem_block["source_status"] = STATUS_FAILED
        pubchem_block["detail"] = "missing_entrez"
        warnings.append("pubchem_missing_entrez")
    else:
        aids_tr = pubchem_client.aids_by_geneid(entrez, gene_symbol=gene, settings=cfg)
        _persist_tool(
            aids_tr,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="pubchem-aids",
        )
        if not aids_tr.success:
            pubchem_block["source_status"] = STATUS_SOURCE_ERROR
        else:
            aids = pubchem_client.extract_aid_list(aids_tr.data)
            pubchem_block["aid_count"] = len(aids)
            if not aids:
                pubchem_block["source_status"] = STATUS_NO_RESULTS
                pubchem_block["display"] = "PubChem – No results"
            else:
                # batch in chunks of 20
                focused_rows = []
                excluded_rows = []
                for i in range(0, len(aids), 20):
                    chunk = aids[i : i + 20]
                    desc_tr = pubchem_client.assay_descriptions_batch(
                        chunk, gene_symbol=gene, settings=cfg
                    )
                    _persist_tool(
                        desc_tr,
                        dossier_run_id=run_id,
                        gene_symbol=gene,
                        settings=cfg,
                        persist_db=persist_db,
                        api_runs=api_runs,
                        raw_artifacts=raw_artifacts,
                        tool_results=tool_results,
                        filename_hint=f"pubchem-desc-{i}",
                    )
                    tgt_tr = pubchem_client.assay_targets_geneid_symbol_batch(
                        chunk, gene_symbol=gene, settings=cfg
                    )
                    _persist_tool(
                        tgt_tr,
                        dossier_run_id=run_id,
                        gene_symbol=gene,
                        settings=cfg,
                        persist_db=persist_db,
                        api_runs=api_runs,
                        raw_artifacts=raw_artifacts,
                        tool_results=tool_results,
                        filename_hint=f"pubchem-tgt-{i}",
                    )
                    # map per-aid description containers
                    containers = []
                    if desc_tr.success and isinstance(desc_tr.data, dict):
                        containers = list(desc_tr.data.get("PC_AssayContainer") or [])
                    tgt_info = []
                    if tgt_tr.success and isinstance(tgt_tr.data, dict):
                        tgt_info = list(
                            (tgt_tr.data.get("InformationList") or {}).get("Information")
                            or []
                        )
                    tgt_by_aid = {
                        int(item.get("AID")): item
                        for item in tgt_info
                        if isinstance(item, dict) and item.get("AID") is not None
                    }
                    desc_by_aid: dict[int, Any] = {}
                    for c in containers:
                        if not isinstance(c, dict):
                            continue
                        descr = (c.get("assay") or {}).get("descr") or {}
                        aid_obj = descr.get("aid") or {}
                        aid_v = aid_obj.get("id") if isinstance(aid_obj, dict) else aid_obj
                        if aid_v is not None:
                            desc_by_aid[int(aid_v)] = c
                    for aid in chunk:
                        desc_payload = desc_by_aid.get(int(aid))
                        desc_for_classify = (
                            {"PC_AssayContainer": [desc_payload]}
                            if isinstance(desc_payload, dict)
                            else None
                        )
                        if int(aid) in tgt_by_aid:
                            tgt_payload = {
                                "InformationList": {
                                    "Information": [tgt_by_aid[int(aid)]]
                                }
                            }
                        else:
                            tgt_payload = {"Fault": True}
                        name = ""
                        comment_bits = []
                        encoding = gene
                        if desc_payload:
                            descr = (desc_payload.get("assay") or {}).get("descr") or {}
                            name = str(descr.get("name") or "")
                            for t in descr.get("target") or []:
                                if isinstance(t, dict):
                                    comment_bits.append(str(t.get("name") or ""))
                                    mol = t.get("mol_id") or {}
                                    if mol.get("protein_accession"):
                                        comment_bits.append(str(mol.get("protein_accession")))
                        name_cf = name.casefold()
                        if any(
                            tok in name_cf
                            for tok in (
                                "rnai",
                                "shrna",
                                "sirna",
                                "genome-wide",
                                "genome wide",
                                "library screen",
                            )
                        ):
                            focused, reason = False, "broad_screen_membership_only"
                        else:
                            focused, reason = pubchem_client.classify_focused_assay(
                                desc_for_classify,
                                tgt_payload,
                                uniprot=uniprot or None,
                                entrez=entrez,
                                symbol=gene,
                                chembl_target_id=chembl_block.get("target_chembl_id"),
                            )
                        row = {
                            "aid": int(aid),
                            "bioassay_name": name,
                            "target_gene": encoding,
                            "comment": "; ".join([b for b in comment_bits if b])[:500],
                            "focused": focused,
                            "reason": reason,
                            "exclusion_reason": None if focused else reason,
                            "protein_accession": None,
                        }
                        pubs = {}
                        if desc_payload:
                            descr = (desc_payload.get("assay") or {}).get("descr") or {}
                            if isinstance(descr, dict):
                                pubs = extract_pubchem_publication_ids(descr)
                                for t in descr.get("target") or []:
                                    if isinstance(t, dict):
                                        mol = t.get("mol_id") or {}
                                        if mol.get("protein_accession"):
                                            row["protein_accession"] = str(
                                                mol.get("protein_accession")
                                            )
                                            break
                        row.update(pubs)
                        if focused:
                            focused_rows.append(row)
                        else:
                            excluded_rows.append(row)
                polished_focused = select_polished_pubchem_rows(
                    focused_rows, cap=POLISHED_PUBCHEM_CAP
                )
                pubchem_block["focused"] = focused_rows
                pubchem_block["polished_focused"] = polished_focused
                pubchem_block["excluded"] = excluded_rows
                pubchem_block["source_status"] = (
                    STATUS_SUCCESS if focused_rows else STATUS_NO_RESULTS
                )
                pubchem_block["display"] = (
                    f"PubChem – {len(focused_rows)} focused assay(s); "
                    f"{len(polished_focused)} representative(s) shown"
                    if focused_rows
                    else "PubChem – No results"
                )

    source_blocks["pubchem"] = pubchem_block

    # ---- NCATS ----
    ncats_block: dict[str, Any] = {
        "source_status": STATUS_NO_RESULTS,
        "candidates": [],
    }
    # Prefer UniProt protein pref name search via Primary Target facet using gene symbol / common protein names from gene_ids only
    protein_label = _protein_name_from_state(
        {**state, "gene_ids": gene_ids, "evidence_records": evidence_records}
    )
    facet_labels: list[str] = []
    if protein_label:
        facet_labels.append(protein_label)
    # Official symbol as facet label — never invent gene-specific aliases
    facet_labels.append(gene)
    seen_labels = set()
    candidates = []
    ncats_ok = False
    for label in facet_labels:
        if not label or label.casefold() in seen_labels:
            continue
        seen_labels.add(label.casefold())
        s_tr = ncats_client.search_substances(
            gene_symbol=gene, primary_target=label, top=10, settings=cfg
        )
        _persist_tool(
            s_tr,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="ncats-search",
        )
        if not s_tr.success:
            continue
        ncats_ok = True
        content = s_tr.data if isinstance(s_tr.data, dict) else {}
        total = int(content.get("total") or content.get("count") or 0)
        items = content.get("content") or content.get("substances") or []
        if not isinstance(items, list):
            items = []
        for item in items:
            if not isinstance(item, dict):
                continue
            uuid = str(item.get("uuid") or item.get("id") or "")
            name = str(
                item.get("name")
                or (item.get("_name") if isinstance(item.get("_name"), str) else "")
                or ""
            )
            if not uuid:
                continue
            rel_tr = ncats_client.substance_relationships(
                uuid, gene_symbol=gene, settings=cfg
            )
            _persist_tool(
                rel_tr,
                dossier_run_id=run_id,
                gene_symbol=gene,
                settings=cfg,
                persist_db=persist_db,
                api_runs=api_runs,
                raw_artifacts=raw_artifacts,
                tool_results=tool_results,
                filename_hint=f"ncats-rel-{uuid[:8]}",
            )
            rels = rel_tr.data if rel_tr.success else []
            if isinstance(rels, dict):
                rels = rels.get("content") or rels.get("relationships") or [rels]
            if not isinstance(rels, list):
                rels = []
            confirmed = False
            for rel in rels:
                if not isinstance(rel, dict):
                    continue
                blob = json.dumps(rel).casefold()
                if gene.casefold() in blob or (uniprot and uniprot.casefold() in blob):
                    if "target" in blob or "protein" in blob:
                        confirmed = True
                        break
            evidence_class = (
                "direct_target_evidence"
                if confirmed
                else "facet_target_match_unconfirmed"
            )
            candidates.append(
                {
                    "uuid": uuid,
                    "name": name or uuid,
                    "facet_label": label,
                    "evidence_class": evidence_class,
                    "relationship_types": [
                        r.get("type") for r in rels if isinstance(r, dict)
                    ],
                }
            )
        if total == 0 and not items:
            continue
    if candidates:
        ncats_block["source_status"] = STATUS_SUCCESS
        ncats_block["candidates"] = candidates
        ncats_block["display"] = ncats_display_text(candidates)
    elif ncats_ok:
        ncats_block["source_status"] = STATUS_NO_RESULTS
        ncats_block["display"] = "NCATS Inxight: Drugs – No results"
    else:
        ncats_block["source_status"] = STATUS_SOURCE_ERROR
        ncats_block["display"] = "NCATS Inxight: Drugs – source error"
    source_blocks["ncats"] = ncats_block
    ncats_names = {_casefold(c.get("name")) for c in candidates if c.get("name")}

    # ---- PubTator + PubMed ----
    pubtator_status = STATUS_NO_RESULTS
    pubmed_status = STATUS_NO_RESULTS
    relation_inventory: list[dict[str, Any]] = []
    literature_entries: list[dict[str, Any]] = []
    fallback_pmids: list[str] = []
    expansion_meta = {
        "total_relation_candidate_count": 0,
        "expanded_relation_candidate_count": 0,
        "expansion_batch_size": section_cfg.expansion_batch_size,
        "expansion_limit": section_cfg.expansion_limit,
        "expansion_exhausted": False,
    }

    entity_id = None
    ac = pubtator_client.entity_autocomplete(
        gene, concept="gene", limit=10, gene_symbol=gene, settings=cfg
    )
    _persist_tool(
        ac,
        dossier_run_id=run_id,
        gene_symbol=gene,
        settings=cfg,
        persist_db=persist_db,
        api_runs=api_runs,
        raw_artifacts=raw_artifacts,
        tool_results=tool_results,
        filename_hint="pubtator-autocomplete",
    )
    if ac.success and isinstance(ac.data, list):
        for hit in ac.data:
            if not isinstance(hit, dict):
                continue
            db_id = str(hit.get("db_id") or "")
            eid = str(hit.get("_id") or "")
            if entrez is not None and db_id == str(entrez):
                entity_id = eid
                break
        if not entity_id and ac.data:
            # fail closed if Entrez mismatch
            pubtator_status = STATUS_SOURCE_ERROR
            warnings.append("pubtator_entrez_mismatch")
    elif not ac.success:
        pubtator_status = STATUS_SOURCE_ERROR

    if entity_id:
        for rtype in ("negative_correlate", "positive_correlate", "interact"):
            r_tr = pubtator_client.relations(
                entity_id, rtype, "Chemical", gene_symbol=gene, settings=cfg
            )
            _persist_tool(
                r_tr,
                dossier_run_id=run_id,
                gene_symbol=gene,
                settings=cfg,
                persist_db=persist_db,
                api_runs=api_runs,
                raw_artifacts=raw_artifacts,
                tool_results=tool_results,
                filename_hint=f"pubtator-rel-{rtype}",
            )
            if not r_tr.success:
                pubtator_status = STATUS_SOURCE_ERROR
                continue
            rows = r_tr.data if isinstance(r_tr.data, list) else []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                # source is chemical, target gene (validated probe shape)
                chem_id = str(row.get("source") or "")
                chem_name = chem_id.replace("@CHEMICAL_", "").replace("_", " ")
                if row.get("target") and str(row.get("target")).startswith("@CHEMICAL_"):
                    chem_id = str(row.get("target"))
                    chem_name = chem_id.replace("@CHEMICAL_", "").replace("_", " ")
                relation_inventory.append(
                    {
                        "chemical_entity_id": chem_id,
                        "chemical_name": chem_name,
                        "relation_type": row.get("type") or rtype,
                        "support_count": int(row.get("publications") or 0),
                        "candidate_expanded": False,
                        "expansion_rank": None,
                        "exclusion_reason": None,
                    }
                )
        if relation_inventory and pubtator_status != STATUS_SOURCE_ERROR:
            pubtator_status = STATUS_SUCCESS
        elif pubtator_status != STATUS_SOURCE_ERROR:
            pubtator_status = STATUS_NO_RESULTS

    expansion_meta["total_relation_candidate_count"] = len(relation_inventory)

    def _force_expand(name: str, entity_id: str) -> bool:
        cf = _casefold(name)
        if cf in ncats_names:
            return True
        if any(cf and cf in m for m in chembl_molecule_names):
            return True
        return False

    # rank relations
    ranked_relations = sorted(
        relation_inventory,
        key=lambda r: (
            RELATION_PRIORITY.get(str(r.get("relation_type")), 9),
            -int(r.get("support_count") or 0),
            _casefold(r.get("chemical_name")),
            str(r.get("chemical_entity_id") or ""),
        ),
    )
    for i, r in enumerate(ranked_relations):
        r["relation_rank"] = i + 1

    def _expand_relation(rel: dict[str, Any], expansion_rank: int) -> list[dict[str, Any]]:
        chem_id = str(rel.get("chemical_entity_id") or "")
        if not chem_id or not entity_id:
            return []
        rel["candidate_expanded"] = True
        rel["expansion_rank"] = expansion_rank
        search = pubtator_client.search_relations(
            str(rel.get("relation_type") or "negative_correlate"),
            chem_id,
            entity_id,
            page=1,
            gene_symbol=gene,
            settings=cfg,
        )
        _persist_tool(
            search,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="pubtator-search",
        )
        pmids = pubtator_client.extract_search_pmids(search.data) if search.success else []
        out = []
        if not pmids:
            return out
        # fetch titles/abstracts for up to 5 pmids per chemical
        use = pmids[:5]
        es = pubmed_client.esummary(use, gene_symbol=gene, settings=cfg)
        _persist_tool(
            es,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="pubmed-esummary",
        )
        ef = pubmed_client.efetch_abstracts(use, gene_symbol=gene, settings=cfg)
        _persist_tool(
            ef,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="pubmed-efetch",
        )
        titles: dict[str, str] = {}
        citations: dict[str, dict[str, Any]] = {}
        if es.success and isinstance(es.data, dict):
            result = (es.data.get("result") or {})
            for pid in use:
                row = result.get(str(pid)) or {}
                if isinstance(row, dict):
                    titles[str(pid)] = str(row.get("title") or "")
                    citations[str(pid)] = citation_label_from_esummary(row, str(pid))
        abstracts: dict[str, str] = {}
        raw_xml = ""
        if ef.success and isinstance(ef.data, dict):
            raw_xml = str(ef.data.get("raw_text") or ef.data.get("xml") or "")
        # crude abstract split by PMID
        for pid in use:
            m = re.search(
                rf"<PMID[^>]*>{pid}</PMID>.*?<AbstractText[^>]*>(.*?)</AbstractText>",
                raw_xml,
                re.I | re.S,
            )
            if m:
                abstracts[pid] = re.sub(r"<[^>]+>", "", m.group(1))
            else:
                abstracts[pid] = ""
        for pid in use:
            title = titles.get(pid, "")
            abstract = abstracts.get(pid, "")
            classified = classify_evidence_span(
                title=title,
                abstract=abstract,
                chemical_name=str(rel.get("chemical_name") or ""),
                gene_symbol=gene,
                gene_aliases=alias_names,
            )
            cite = citations.get(pid) or {}
            out.append(
                {
                    "pmid": pid,
                    "chemical_entity_id": chem_id,
                    "chemical_name": rel.get("chemical_name"),
                    "relation_type": rel.get("relation_type"),
                    "support_count": rel.get("support_count") or 0,
                    "pmids": [pid],
                    "title": title,
                    "discovery_source": "pubtator_relation",
                    "citation_label": cite.get("citation_label"),
                    "authors": cite.get("authors"),
                    "year": cite.get("year"),
                    **classified,
                }
            )
        return out

    # expand in batches until 7 display-eligible
    expanded_count = 0
    force_ids = {
        str(r.get("chemical_entity_id"))
        for r in ranked_relations
        if _force_expand(str(r.get("chemical_name") or ""), str(r.get("chemical_entity_id") or ""))
    }
    queue = list(ranked_relations)
    # move force-expand to front while preserving relative order
    queue.sort(key=lambda r: (0 if str(r.get("chemical_entity_id")) in force_ids else 1, r.get("relation_rank") or 0))

    idx = 0
    while True:
        displayable = [e for e in literature_entries if _display_eligible(e)]
        if len(rank_literature_candidates(displayable)) >= section_cfg.polished_pubmed_cap:
            break
        if expanded_count >= section_cfg.expansion_limit:
            expansion_meta["expansion_exhausted"] = False
            break
        if idx >= len(queue):
            expansion_meta["expansion_exhausted"] = True
            break
        batch = queue[idx : idx + section_cfg.expansion_batch_size]
        idx += section_cfg.expansion_batch_size
        for rel in batch:
            if expanded_count >= section_cfg.expansion_limit:
                break
            if rel.get("candidate_expanded"):
                continue
            expanded_count += 1
            literature_entries.extend(_expand_relation(rel, expanded_count))
        # continue loop for next batch if needed
        if len(batch) == 0:
            break

    expansion_meta["expanded_relation_candidate_count"] = expanded_count

    # PubMed fallback if needed
    displayable = [e for e in literature_entries if _display_eligible(e)]
    need_fallback = len(rank_literature_candidates(displayable)) < section_cfg.polished_pubmed_cap
    if need_fallback:
        term = build_pubmed_perturbation_term(alias_payload)
        es = pubmed_client.esearch_custom(
            term,
            retmax=section_cfg.pubmed_fallback_retmax,
            gene_symbol=gene,
            settings=cfg,
        )
        _persist_tool(
            es,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            api_runs=api_runs,
            raw_artifacts=raw_artifacts,
            tool_results=tool_results,
            filename_hint="pubmed-fallback-search",
        )
        if es.success:
            pubmed_status = STATUS_SUCCESS  # provisional; may become no_results
            fallback_pmids = pubmed_client.extract_id_list(es)
            if fallback_pmids:
                # Primary grounding: title/abstract PubTator annotations only.
                ann = pubtator_client.fetch_publication_annotations(
                    fallback_pmids[:20], gene_symbol=gene, full=False, settings=cfg
                )
                _persist_tool(
                    ann,
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    settings=cfg,
                    persist_db=persist_db,
                    api_runs=api_runs,
                    raw_artifacts=raw_artifacts,
                    tool_results=tool_results,
                    filename_hint="pubtator-biocjson-ta",
                )
                # Full-text annotations remain audit/backstop only.
                ann_full = pubtator_client.fetch_publication_annotations(
                    fallback_pmids[:20], gene_symbol=gene, full=True, settings=cfg
                )
                _persist_tool(
                    ann_full,
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    settings=cfg,
                    persist_db=persist_db,
                    api_runs=api_runs,
                    raw_artifacts=raw_artifacts,
                    tool_results=tool_results,
                    filename_hint="pubtator-biocjson-full",
                )
                ef = pubmed_client.efetch_abstracts(
                    fallback_pmids[:20], gene_symbol=gene, settings=cfg
                )
                _persist_tool(
                    ef,
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    settings=cfg,
                    persist_db=persist_db,
                    api_runs=api_runs,
                    raw_artifacts=raw_artifacts,
                    tool_results=tool_results,
                    filename_hint="pubmed-fallback-efetch",
                )
                esum = pubmed_client.esummary(
                    fallback_pmids[:20], gene_symbol=gene, settings=cfg
                )
                _persist_tool(
                    esum,
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    settings=cfg,
                    persist_db=persist_db,
                    api_runs=api_runs,
                    raw_artifacts=raw_artifacts,
                    tool_results=tool_results,
                    filename_hint="pubmed-fallback-esummary",
                )
                titles = {}
                citations: dict[str, dict[str, Any]] = {}
                if esum.success and isinstance(esum.data, dict):
                    result = esum.data.get("result") or {}
                    for pid in fallback_pmids[:20]:
                        row = result.get(str(pid)) or {}
                        if isinstance(row, dict):
                            titles[str(pid)] = str(row.get("title") or "")
                            citations[str(pid)] = citation_label_from_esummary(
                                row, str(pid)
                            )
                raw_xml = ""
                if ef.success and isinstance(ef.data, dict):
                    raw_xml = str(ef.data.get("raw_text") or "")
                chem_by_pmid: dict[str, list[dict[str, Any]]] = defaultdict(list)
                if ann.success:
                    for doc in pubtator_client.biocjson_documents(ann.data):
                        did = str(doc.get("_id") or doc.get("id") or "")
                        pmid = did.split("|")[0] if did else ""
                        for chem in pubtator_client.extract_chemical_entities(
                            {"PubTator3": [doc]}
                        ):
                            row = dict(chem)
                            row["grounding_source"] = "pubtator_title_abstract"
                            if not row.get("name"):
                                row["name"] = row.get("text")
                            chem_by_pmid[pmid].append(row)
                full_docs_by_pmid: dict[str, dict[str, Any]] = {}
                fulltext_chem_by_pmid: dict[str, list[dict[str, Any]]] = defaultdict(list)
                if ann_full.success:
                    for doc in pubtator_client.biocjson_documents(ann_full.data):
                        did = str(doc.get("_id") or doc.get("id") or "")
                        pmid = did.split("|")[0] if did else ""
                        if pmid:
                            full_docs_by_pmid[pmid] = doc
                        for chem in pubtator_client.extract_chemical_entities(
                            {"PubTator3": [doc]}
                        ):
                            row = dict(chem)
                            row["grounding_source"] = "pubtator_fulltext_audit"
                            if not row.get("name"):
                                row["name"] = row.get("text")
                            fulltext_chem_by_pmid[pmid].append(row)
                # Full PubMed XML for ChemicalList + reliable title/abstract
                full = pubmed_client.efetch_pubmed_xml(
                    fallback_pmids[:20], gene_symbol=gene, settings=cfg
                )
                _persist_tool(
                    full,
                    dossier_run_id=run_id,
                    gene_symbol=gene,
                    settings=cfg,
                    persist_db=persist_db,
                    api_runs=api_runs,
                    raw_artifacts=raw_artifacts,
                    tool_results=tool_results,
                    filename_hint="pubmed-fallback-fullxml",
                )
                full_xml = ""
                if full.success and isinstance(full.data, dict):
                    full_xml = str(full.data.get("raw_text") or "")
                mesh_by_pmid = pubmed_client.extract_medline_chemicals(full_xml)
                ta_by_pmid = pubmed_client.extract_pmid_title_abstract(full_xml)
                for pid in fallback_pmids[:20]:
                    meta = ta_by_pmid.get(str(pid)) or {}
                    title = str(meta.get("title") or titles.get(str(pid), ""))
                    abstract = str(meta.get("abstract") or "")
                    if not abstract:
                        m = re.search(
                            rf"<PMID[^>]*>{pid}</PMID>.*?<AbstractText[^>]*>(.*?)</AbstractText>",
                            raw_xml,
                            re.I | re.S,
                        )
                        abstract = re.sub(r"<[^>]+>", "", m.group(1)) if m else ""
                    cite = citations.get(str(pid)) or {}
                    chems = list(chem_by_pmid.get(str(pid)) or [])
                    if not chems:
                        for mesh in mesh_by_pmid.get(str(pid)) or []:
                            mesh_name = str(mesh.get("name") or "")
                            # ChemicalList often includes gene/protein substances; skip those.
                            if re.search(r"\bprotein\b", mesh_name, re.I):
                                continue
                            if _casefold(mesh_name) in {_casefold(gene)} | {
                                _casefold(a) for a in alias_names
                            }:
                                continue
                            chems.append(
                                {
                                    "name": mesh.get("name"),
                                    "text": mesh.get("name"),
                                    "id": None,
                                    "grounding_source": "pubmed_chemical_list",
                                }
                            )
                    if not chems:
                        for hit in extract_explicit_chemical_mentions(title, abstract):
                            hit["grounding_source"] = hit.get("source")
                            chems.append(hit)
                    # Full-text chemicals may enter polished only via local windows.
                    ta_names = {_casefold(c.get("name") or c.get("text")) for c in chems}
                    for ft_chem in fulltext_chem_by_pmid.get(str(pid)) or []:
                        cname = str(
                            ft_chem.get("name") or ft_chem.get("text") or ft_chem.get("id") or ""
                        )
                        if not cname or _casefold(cname) in ta_names:
                            continue
                        # Keep audit entry even when not polished.
                        doc = full_docs_by_pmid.get(str(pid)) or {}
                        snippets = _passage_snippets_for_chemical(doc, cname)
                        classified_ft = classify_evidence_span(
                            title=title,
                            abstract=snippets or "",
                            chemical_name=cname,
                            gene_symbol=gene,
                            gene_aliases=alias_names,
                        )
                        entry_ft = {
                            "pmid": str(pid),
                            "chemical_entity_id": ft_chem.get("id")
                            or ft_chem.get("accession"),
                            "chemical_name": cname,
                            "mention_text": ft_chem.get("text"),
                            "grounding_source": "pubtator_fulltext_local_window",
                            "discovery_source": "pubmed_perturbation_search",
                            "pmids": [str(pid)],
                            "support_count": 0,
                            "title": title,
                            "citation_label": cite.get("citation_label"),
                            "authors": cite.get("authors"),
                            "year": cite.get("year"),
                            **classified_ft,
                        }
                        if not _display_eligible(entry_ft):
                            entry_ft["exclusion_reason"] = (
                                "fulltext_without_local_perturbation_gene_window"
                            )
                            entry_ft["displayed"] = False
                        literature_entries.append(entry_ft)
                    if not chems and not any(
                        e.get("pmid") == str(pid) for e in literature_entries
                    ):
                        literature_entries.append(
                            {
                                "pmid": str(pid),
                                "chemical_name": None,
                                "discovery_source": "pubmed_perturbation_search",
                                "exclusion_reason": "no_grounded_chemical_candidate",
                                "evidence_class": "insufficient_effect_detail",
                                "tool_eligibility": "insufficient_tool_evidence",
                                "pmids": [str(pid)],
                                "support_count": 0,
                                "title": title,
                            }
                        )
                        continue
                    for chem in chems:
                        cname = str(
                            chem.get("name") or chem.get("text") or chem.get("id") or ""
                        )
                        # Title/abstract chemicals: classify on TA text only.
                        classified = classify_evidence_span(
                            title=title,
                            abstract=abstract,
                            chemical_name=cname,
                            gene_symbol=gene,
                            gene_aliases=alias_names,
                        )
                        entry = {
                            "pmid": str(pid),
                            "chemical_entity_id": chem.get("id")
                            or chem.get("accession"),
                            "chemical_name": cname,
                            "mention_text": chem.get("text"),
                            "grounding_source": chem.get("grounding_source"),
                            "discovery_source": "pubmed_perturbation_search",
                            "pmids": [str(pid)],
                            "support_count": 0,
                            "title": title,
                            "citation_label": cite.get("citation_label"),
                            "authors": cite.get("authors"),
                            "year": cite.get("year"),
                            **classified,
                        }
                        if not _display_eligible(entry):
                            entry["exclusion_reason"] = (
                                entry.get("exclusion_reason")
                                or "not_display_eligible_local_window"
                            )
                        literature_entries.append(entry)
            else:
                pubmed_status = STATUS_NO_RESULTS
        else:
            pubmed_status = STATUS_SOURCE_ERROR

    # Deduplicate by pmid+chemical
    dedup: dict[tuple[str, str], dict[str, Any]] = {}
    for e in literature_entries:
        key = (str(e.get("pmid") or ""), _casefold(e.get("chemical_name")))
        prior = dedup.get(key)
        if prior is None:
            dedup[key] = e
            continue
        # Prefer title/abstract grounding over full-text audit for the same chem.
        prior_ft = "fulltext" in str(prior.get("grounding_source") or "")
        cur_ft = "fulltext" in str(e.get("grounding_source") or "")
        if prior_ft and not cur_ft:
            dedup[key] = e
    literature_entries = list(dedup.values())

    ranked_lit = rank_literature_candidates(
        [e for e in literature_entries if _display_eligible(e) and not e.get("exclusion_reason")]
    )
    polished_lit = []
    seen_chem: set[str] = set()
    for entry in ranked_lit:
        chem_key = _casefold(entry.get("chemical_name"))
        if chem_key and chem_key in seen_chem:
            continue
        if chem_key:
            seen_chem.add(chem_key)
        polished_lit.append(entry)
        if len(polished_lit) >= section_cfg.polished_pubmed_cap:
            break
    for e in literature_entries:
        e["displayed"] = e in polished_lit
        if not e.get("displayed") and not e.get("exclusion_reason"):
            if not _display_eligible(e):
                e["exclusion_reason"] = e.get("exclusion_reason") or "not_display_eligible"
            elif e not in polished_lit:
                e["exclusion_reason"] = "beyond_polished_cap"

    if polished_lit:
        pubmed_status = STATUS_SUCCESS
    elif pubmed_status != STATUS_SOURCE_ERROR and pubtator_status != STATUS_SOURCE_ERROR:
        if not literature_entries:
            pubmed_status = STATUS_NO_RESULTS
        else:
            pubmed_status = STATUS_NO_RESULTS

    literature_status = derive_literature_status(pubtator_status, pubmed_status)
    source_blocks["pubmed"] = {
        "source_status": literature_status,
        "pubtator_status": pubtator_status,
        "pubmed_status": pubmed_status,
        "entity_id": entity_id,
        "polished": polished_lit,
        "inventory_count": len(literature_entries),
        "fallback_pmids": fallback_pmids,
        "expansion": expansion_meta,
        "alias_resolution": alias_payload,
        "display": (
            f"PubMed – {len(polished_lit)} literature tool(s)"
            if polished_lit
            else "PubMed – No results"
        ),
    }
    source_blocks["pubtator"] = {"source_status": pubtator_status, "entity_id": entity_id}

    # Overall scientific status
    public_error = any(
        source_blocks[k]["source_status"] == STATUS_SOURCE_ERROR
        for k in ("chembl", "pubchem", "ncats", "pubmed")
    )
    has_tools = bool(
        (chembl_block.get("activity_count") or 0) > 0
        or pubchem_block.get("focused")
        or polished_lit
        or any(
            c.get("evidence_class") != "facet_target_match_unconfirmed"
            for c in ncats_block.get("candidates")
            or []
        )
        or ncats_block.get("candidates")
    )
    # NCATS unconfirmed still counts as a usable candidate record for section non-empty
    if ncats_block.get("candidates"):
        has_tools = True

    if entrez is None:
        overall = STATUS_FAILED
    elif public_error and has_tools:
        overall = STATUS_PARTIAL
    elif public_error and not has_tools:
        overall = STATUS_PARTIAL
    elif has_tools and source_blocks["drugbank"]["source_status"] == STATUS_UNAVAILABLE:
        overall = STATUS_LIMITATIONS
    elif has_tools:
        overall = STATUS_SUCCESS
    else:
        overall = STATUS_SUCCESS_NO_TOOLS

    # Build audit WITHOUT sha field, write, then hash externally
    audit = {
        "parser_version": PARSER_VERSION,
        "gene_symbol": gene,
        "entrez_gene_id": entrez,
        "uniprot_accession": uniprot,
        "alias_resolution": alias_payload,
        "sources": {
            "chembl": {
                k: v
                for k, v in chembl_block.items()
                if k not in {"activities", "assays"}
            },
            "chembl_activities": chembl_block.get("activities") or [],
            "chembl_assays": chembl_block.get("assays") or [],
            "drugbank": source_blocks["drugbank"],
            "pubchem": pubchem_block,
            "ncats": ncats_block,
            "pubmed": {
                "pubtator_status": pubtator_status,
                "pubmed_status": pubmed_status,
                "literature_status": literature_status,
                "relation_inventory": ranked_relations,
                "literature_entries": literature_entries,
                "polished": polished_lit,
                "fallback_pmids": fallback_pmids,
                "expansion": expansion_meta,
            },
        },
        "warnings": warnings,
        "overall_scientific_status": overall,
    }
    audit_path = attempt_dir / "section_7a_audit.json"
    write_json_atomic(audit_path, audit)
    audit_sha = sha256_file(audit_path)

    summary = {
        "gene_symbol": gene,
        "scientific_status": overall,
        "presentation_status": STATUS_SUCCESS if overall != STATUS_FAILED else STATUS_FAILED,
        "visual_status": STATUS_SUCCESS,
        "source_blocks": {
            "chembl": {
                "source_status": chembl_block.get("source_status"),
                "display": chembl_block.get("display"),
                "target_chembl_id": chembl_block.get("target_chembl_id"),
                "activity_count": chembl_block.get("activity_count"),
                "assay_count": chembl_block.get("assay_count"),
                "activity_pages_fetched": chembl_block.get("activity_pages_fetched"),
                "assay_pages_fetched": chembl_block.get("assay_pages_fetched"),
                "activity_page_api_run_ids": chembl_block.get(
                    "activity_page_api_run_ids"
                )
                or [],
                "activity_page_raw_artifact_ids": chembl_block.get(
                    "activity_page_raw_artifact_ids"
                )
                or [],
                "assay_page_api_run_ids": chembl_block.get("assay_page_api_run_ids")
                or [],
                "assay_page_raw_artifact_ids": chembl_block.get(
                    "assay_page_raw_artifact_ids"
                )
                or [],
                "workbook": chembl_block.get("workbook"),
                "workbook_sha256": chembl_block.get("workbook_sha256"),
            },
            "drugbank": source_blocks["drugbank"],
            "pubmed": {
                "source_status": literature_status,
                "display": source_blocks["pubmed"]["display"],
                "polished_count": len(polished_lit),
                "pubtator_status": pubtator_status,
                "pubmed_status": pubmed_status,
            },
            "pubchem": {
                "source_status": pubchem_block.get("source_status"),
                "display": pubchem_block.get("display"),
                "aid_count": pubchem_block.get("aid_count"),
                "focused_count": len(pubchem_block.get("focused") or []),
                "polished_focused_count": len(pubchem_block.get("polished_focused") or []),
                "excluded_count": len(pubchem_block.get("excluded") or []),
                "focused": pubchem_block.get("focused") or [],
                "polished_focused": pubchem_block.get("polished_focused") or [],
                "excluded": pubchem_block.get("excluded") or [],
            },
            "ncats": {
                "source_status": ncats_block.get("source_status"),
                "display": ncats_block.get("display"),
                "candidates": ncats_block.get("candidates") or [],
            },
        },
        "polished_literature": polished_lit,
        "section_7a_audit_path": str(audit_path),
        "section_7a_audit_sha256": audit_sha,
        "chembl_workbook_sha256": chembl_block.get("workbook_sha256"),
        "attempt_dir": str(attempt_dir),
        "parser_version": PARSER_VERSION,
        "scientific_intro": (
            f"{gene} gene was queried in the following databases to identify any known "
            "inhibitors or agonists and their effect(s) on the gene."
        ),
    }

    # manifest stores external hashes only
    write_json_atomic(
        attempt_dir / MANIFEST_FILENAME,
        {
            "section_key": "7a",
            "gene_symbol": gene,
            "dossier_run_id": run_id,
            "section_7a_audit_sha256": audit_sha,
            "chembl_workbook_sha256": chembl_block.get("workbook_sha256"),
            "chembl_workbook": chembl_block.get("workbook"),
            "parser_version": PARSER_VERSION,
            "created_at": _utc_now(),
        },
    )
    write_json_atomic(attempt_dir / "summary.json", summary)
    write_json_atomic(attempt_dir / "section_7a_status.json", {
        "section_key": "7a",
        "summary": summary,
        "rendering_status": {
            "scientific_status": overall,
            "presentation_status": summary["presentation_status"],
            "visual_status": STATUS_SUCCESS,
        },
        "audit": {
            "gene_attempt_dir": str(attempt_dir),
            "attempt_dir": str(attempt_dir),
            "section_7a_audit_sha256": audit_sha,
        },
    })

    # evidence records (sparse)
    _append_evidence(
        evidence_records,
        _evidence(
            dossier_run_id=run_id,
            gene_symbol=gene,
            source_name="Section7a",
            fact_type="section_7a_summary",
            key="summary",
            value={"summary": summary},
            display_text=summary["scientific_intro"],
        ),
        persist_db=persist_db,
    )
    if chembl_block.get("workbook"):
        _append_evidence(
            evidence_records,
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                source_name="ChEMBL",
                fact_type="section_7a_chembl_workbook",
                key="workbook",
                value={
                    "workbook": chembl_block.get("workbook"),
                    "sha256": chembl_block.get("workbook_sha256"),
                    "local_path": chembl_block.get("workbook_path"),
                },
                display_text=chembl_block.get("display") or "",
                api_run_id=chembl_api,
            ),
            persist_db=persist_db,
        )
    for i, lit in enumerate(polished_lit):
        _append_evidence(
            evidence_records,
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                source_name="PubMed",
                fact_type="section_7a_pubmed_tool",
                key=f"lit-{i}-{lit.get('pmid')}",
                value=lit,
                display_text=build_literature_effect_prose(lit),
                confidence_notes=str(lit.get("evidence_class")),
            ),
            persist_db=persist_db,
        )
    for i, row in enumerate(pubchem_block.get("polished_focused") or []):
        _append_evidence(
            evidence_records,
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                source_name="PubChem",
                fact_type="section_7a_pubchem_assay",
                key=f"aid-{row.get('aid')}",
                value=row,
                display_text=f"AID {row.get('aid')}: {row.get('bioassay_name')}",
            ),
            persist_db=persist_db,
        )
    for i, cand in enumerate(ncats_block.get("candidates") or []):
        _append_evidence(
            evidence_records,
            _evidence(
                dossier_run_id=run_id,
                gene_symbol=gene,
                source_name="NCATS Inxight",
                fact_type="section_7a_ncats_candidate",
                key=f"ncats-{i}",
                value=cand,
                display_text=f"{cand.get('name')} ({cand.get('evidence_class')})",
            ),
            persist_db=persist_db,
        )
    _append_evidence(
        evidence_records,
        _evidence(
            dossier_run_id=run_id,
            gene_symbol=gene,
            source_name="DrugBank",
            fact_type="section_7a_drugbank_status",
            key="drugbank",
            value=source_blocks["drugbank"],
            display_text=source_blocks["drugbank"]["display"],
        ),
        persist_db=persist_db,
    )

    status = {
        "section_key": "7a",
        "summary": summary,
        "rendering_status": {
            "scientific_status": overall,
            "presentation_status": summary["presentation_status"],
            "visual_status": STATUS_SUCCESS,
        },
        "audit": {
            "gene_attempt_dir": str(attempt_dir),
            "attempt_dir": str(attempt_dir),
            "section_7a_audit_sha256": audit_sha,
            "chembl_workbook_sha256": chembl_block.get("workbook_sha256"),
            "artifacts": {
                "section_7a_audit_sha256": audit_sha,
                "chembl_workbook_sha256": chembl_block.get("workbook_sha256"),
                "chembl_workbook": chembl_block.get("workbook"),
            },
        },
    }
    return {
        **state,
        "api_runs": api_runs,
        "raw_artifacts": raw_artifacts,
        "tool_results": tool_results,
        "evidence_records": evidence_records,
        "section_7a_status": status,
        "warnings": list(state.get("warnings") or []) + warnings,
    }


def evaluate_section_7a_complete(
    *,
    status: dict[str, Any],
    html_text: str,
    pdf_path: Path | None,
    attempt_dir: Path | None,
) -> dict[str, Any]:
    checks: dict[str, Any] = {}

    def _ok(name: str, passed: bool, detail: Any = None) -> bool:
        checks[name] = {"passed": bool(passed), "detail": detail}
        return bool(passed)

    summary = dict(status.get("summary") or {})
    blocks = dict(summary.get("source_blocks") or {})
    all_pass = True

    all_pass &= _ok("identity_present", bool(summary.get("gene_symbol")))
    for name in ("chembl", "drugbank", "pubmed", "pubchem", "ncats"):
        st = (blocks.get(name) or {}).get("source_status")
        all_pass &= _ok(f"source_status_{name}", bool(st), st)

    chembl = blocks.get("chembl") or {}
    if chembl.get("source_status") == STATUS_NO_AUTHORITATIVE_TARGET:
        all_pass &= _ok(
            "chembl_no_activity_without_target",
            int(chembl.get("activity_count") or 0) == 0
            and not (chembl.get("activity_page_api_run_ids") or []),
            {
                "activity_count": chembl.get("activity_count"),
                "pages": chembl.get("activity_page_api_run_ids"),
            },
        )
    if chembl.get("target_chembl_id"):
        tid = str(chembl.get("target_chembl_id"))
        acts = []
        # Prefer attempt audit for detailed activity/assay rows.
        audit_acts: list[dict[str, Any]] = []
        audit_assays: list[dict[str, Any]] = []
        if attempt_dir and (attempt_dir / "section_7a_audit.json").exists():
            try:
                audit_payload = json.loads(
                    (attempt_dir / "section_7a_audit.json").read_text(encoding="utf-8")
                )
                audit_acts = list(
                    (audit_payload.get("sources") or {}).get("chembl_activities") or []
                )
                audit_assays = list(
                    (audit_payload.get("sources") or {}).get("chembl_assays") or []
                )
            except Exception as exc:  # noqa: BLE001
                all_pass &= _ok("chembl_audit_readable", False, str(exc))
        acts = audit_acts
        all_pass &= _ok(
            "chembl_activities_exact_target",
            all(str(a.get("target_chembl_id") or tid) == tid for a in acts),
            len(acts),
        )
        assay_by_id = {
            str(a.get("assay_chembl_id")): a
            for a in audit_assays
            if a.get("assay_chembl_id")
        }
        direct_ok = True
        for a in acts:
            if a.get("evidence_class") != "direct_target_evidence":
                continue
            meta = assay_by_id.get(str(a.get("assay_chembl_id") or "")) or a
            if not chembl_client.is_direct_assay_relationship(meta):
                direct_ok = False
                break
        all_pass &= _ok("chembl_direct_requires_relationship_metadata", direct_ok)
        page_ids = list(chembl.get("activity_page_api_run_ids") or [])
        raw_ids = list(chembl.get("activity_page_raw_artifact_ids") or [])
        pages = int(chembl.get("activity_pages_fetched") or len(page_ids) or 0)
        all_pass &= _ok(
            "chembl_activity_pages_have_api_runs",
            pages == 0 or (len(page_ids) == pages and len(set(page_ids)) == pages),
            {"pages": pages, "api_runs": page_ids},
        )
        all_pass &= _ok(
            "chembl_activity_pages_have_raw_artifacts",
            pages == 0 or (len(raw_ids) == pages and len(set(raw_ids)) == pages),
            {"pages": pages, "raw": raw_ids},
        )

    if chembl.get("workbook_sha256") and attempt_dir:
        wb = attempt_dir / "supplementary" / str(chembl.get("workbook") or "")
        if wb.exists():
            all_pass &= _ok(
                "chembl_workbook_sha_matches",
                sha256_file(wb) == chembl.get("workbook_sha256"),
            )
            all_pass &= _ok("chembl_workbook_exists", True)
        else:
            all_pass &= _ok("chembl_workbook_exists", False, str(wb))

    audit_path = Path(str(summary.get("section_7a_audit_path") or ""))
    if attempt_dir and (attempt_dir / "section_7a_audit.json").exists():
        audit_path = attempt_dir / "section_7a_audit.json"
    if audit_path.exists():
        recomputed = sha256_file(audit_path)
        all_pass &= _ok(
            "audit_sha_matches",
            recomputed == summary.get("section_7a_audit_sha256"),
            {"expected": summary.get("section_7a_audit_sha256"), "actual": recomputed},
        )
        try:
            payload = json.loads(audit_path.read_text(encoding="utf-8"))
            all_pass &= _ok(
                "audit_has_no_self_sha",
                "section_7a_audit_sha256" not in payload and "sha256" not in payload,
            )
        except Exception as exc:  # noqa: BLE001
            all_pass &= _ok("audit_readable", False, str(exc))

    all_pass &= _ok("html_exists", bool(html_text))
    all_pass &= _ok("pdf_exists", bool(pdf_path and Path(pdf_path).exists()))
    all_pass &= _ok(
        "major7_once",
        html_text.count('id="section-7"') == 1,
        html_text.count('id="section-7"'),
    )
    all_pass &= _ok(
        "subsection_7a_once",
        html_text.count("subsection-7a") == 1,
        html_text.count("subsection-7a"),
    )
    major7_html = html_text
    m7 = re.search(
        r'id="section-7".*?(?:id="section-\d+"|$)',
        html_text,
        re.I | re.S,
    )
    if m7:
        major7_html = m7.group(0)
    positions = []
    for label in ("ChEMBL", "DrugBank", "PubMed", "PubChem", "NCATS"):
        positions.append(major7_html.find(label))
    all_pass &= _ok(
        "source_block_order",
        all(p >= 0 for p in positions) and positions == sorted(positions),
        positions,
    )
    all_pass &= _ok("no_srebf3_leak", "SREBF3" not in html_text)

    polished = list(summary.get("polished_literature") or [])
    all_pass &= _ok("pubmed_cap", len(polished) <= POLISHED_PUBMED_CAP, len(polished))
    for lit in polished:
        pmid = str(lit.get("pmid") or "")
        all_pass &= _ok(
            f"lit_{pmid}_grounded_chemical",
            bool(_norm(lit.get("chemical_name"))),
        )
        all_pass &= _ok(f"lit_{pmid}_has_pmid", bool(pmid))
        all_pass &= _ok(
            f"lit_{pmid}_local_span",
            bool(_norm(lit.get("supporting_span_text")))
            and lit.get("evidence_scope") != "document_wide",
            lit.get("supporting_sentence_indices"),
        )
        all_pass &= _ok(
            f"lit_{pmid}_eligible",
            lit.get("tool_eligibility")
            in {"explicit_chemical_tool", "perturbational_chemical"},
        )
        all_pass &= _ok(
            f"lit_{pmid}_evidence_class",
            lit.get("evidence_class")
            in {
                "literature_negative_effect",
                "literature_positive_effect",
                "literature_interaction",
                "indirect_pathway_effect",
            },
            lit.get("evidence_class"),
        )
        if lit.get("evidence_class") == "indirect_pathway_effect":
            prose = build_literature_effect_prose(lit).casefold()
            all_pass &= _ok(
                f"lit_{pmid}_indirect_not_direct_claim",
                "directly inhibits" not in prose
                and "direct inhibitor of" not in prose,
            )
        # Citation once per entry: no inline PMID duplication with author-year.
        if pmid and major7_html:
            cite = str(lit.get("citation_label") or "")
            prose_text = build_literature_effect_prose(lit)
            inline_pmid = bool(re.search(rf"\bPMID\s*{re.escape(pmid)}\b", prose_text, re.I))
            if cite:
                all_pass &= _ok(
                    f"lit_{pmid}_single_citation",
                    (not inline_pmid) and (cite in major7_html),
                    {"inline_pmid": inline_pmid, "cite_in_html": cite in major7_html},
                )
            else:
                pmid_hits = len(
                    re.findall(rf"\bPMID\s*{re.escape(pmid)}\b", major7_html, re.I)
                )
                all_pass &= _ok(
                    f"lit_{pmid}_single_citation",
                    pmid_hits == 1 and not inline_pmid,
                    pmid_hits,
                )

    pubchem = blocks.get("pubchem") or {}
    focused = list(pubchem.get("focused") or [])
    polished_pc = list(pubchem.get("polished_focused") or [])
    focused_aids = {int(r.get("aid")) for r in focused if r.get("aid") is not None}
    all_pass &= _ok(
        "pubchem_polished_subset",
        all(int(r.get("aid")) in focused_aids for r in polished_pc if r.get("aid") is not None),
    )
    all_pass &= _ok(
        "pubchem_polished_focused_rule",
        all(bool(r.get("focused")) for r in polished_pc),
    )
    all_pass &= _ok(
        "pubchem_polished_cap",
        len(polished_pc) <= POLISHED_PUBCHEM_CAP,
        len(polished_pc),
    )
    for row in list(pubchem.get("excluded") or []):
        all_pass &= _ok(
            f"pubchem_excluded_{row.get('aid')}_has_reason",
            bool(row.get("exclusion_reason") or row.get("reason")),
        )

    ncats = blocks.get("ncats") or {}
    for cand in ncats.get("candidates") or []:
        if cand.get("evidence_class") == "facet_target_match_unconfirmed":
            all_pass &= _ok(
                f"ncats_{cand.get('name')}_not_direct",
                cand.get("evidence_class") != "direct_target_evidence",
            )
    if ncats.get("candidates"):
        display = str(ncats.get("display") or "")
        named = any(
            _norm(c.get("name")) and _norm(c.get("name")) in display
            for c in ncats.get("candidates") or []
        )
        all_pass &= _ok("ncats_names_in_display", named, display)

    drugbank = blocks.get("drugbank") or {}
    all_pass &= _ok(
        "drugbank_unavailable_not_no_results",
        drugbank.get("source_status") == STATUS_UNAVAILABLE,
        drugbank.get("source_status"),
    )

    # Bundle invariants (static).
    from gene_dossier.section_bundle import (
        DEFAULT_SECTION_BUNDLE_KEYS,
        SUPPORTED_SECTION_BUNDLE_KEYS,
    )

    all_pass &= _ok("bundle_7a_supported", "7a" in SUPPORTED_SECTION_BUNDLE_KEYS)
    all_pass &= _ok("bundle_7a_not_default", "7a" not in DEFAULT_SECTION_BUNDLE_KEYS)
    all_pass &= _ok(
        "bundle_defaults_1a_4a",
        DEFAULT_SECTION_BUNDLE_KEYS
        == (
            "1a",
            "1b",
            "1c",
            "1d",
            "1e",
            "2a",
            "2b",
            "2c",
            "3a",
            "4a",
        ),
    )

    return {"complete": all_pass, "checks": checks}


def accept_section_7a_report(
    *,
    gene_symbol: str,
    attempt_dir: Path,
    acceptance: dict[str, Any],
    artifacts: dict[str, Any] | None = None,
    output_root: Path | str | None = None,
    promote_existing: bool = False,
) -> Path | None:
    """Pin accepted Section 7a attempt; do not replace a prior complete pin without promote."""
    from gene_dossier.config import get_settings as _get_settings

    paths = paths_for(output_root or _get_settings().output_path)
    pointer = paths.accepted_gene_pointer(gene_symbol)
    if pointer.exists() and not promote_existing:
        try:
            prior = json.loads(pointer.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            prior = {}
        if (prior.get("acceptance") or {}).get("section_7a_complete"):
            return pointer
    return accept_gene_report(
        paths,
        gene_symbol=gene_symbol,
        attempt_dir=attempt_dir,
        acceptance=acceptance,
        artifacts=artifacts,
    )


__all__ = [
    "SECTION_CHEMICAL",
    "SUBSECTION_7A",
    "PARSER_VERSION",
    "POLISHED_PUBCHEM_CAP",
    "Section7aConfig",
    "resolve_aliases_from_state",
    "build_pubmed_perturbation_term",
    "classify_tool_eligibility",
    "extract_explicit_chemical_mentions",
    "classify_evidence_span",
    "classify_chembl_activity",
    "build_literature_effect_prose",
    "select_polished_pubchem_rows",
    "ncats_display_text",
    "rank_literature_candidates",
    "derive_literature_status",
    "node_generate_section_7a_derived_artifacts",
    "evaluate_section_7a_complete",
    "accept_section_7a_report",
    "write_chembl_workbook",
]
