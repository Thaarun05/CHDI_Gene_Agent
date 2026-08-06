"""Bundle-only Section 4a Harmonizome TF associations.

Owns the single Harmonizome gene associations request, allowlisted parsing,
display selection, supplementary XLSX, and section statuses. Accepted gene
pointers are written only after post-render evaluation in ``section_bundle``.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from gene_dossier.config import Settings, get_settings
from gene_dossier.models import (
    AssertionType,
    EvidenceGrade,
    EvidenceRecord,
    SourceCoverageResult,
    SourceStatus,
    SourceType,
)
from gene_dossier.section_1c import _append_evidence
from gene_dossier.section_4a_sources import (
    MANIFEST_FILENAME,
    Section4aPaths,
    accept_gene_report,
    paths_for,
    sha256_bytes,
    sha256_file,
    write_json_atomic,
)
from gene_dossier.source_ids import make_source_id, slugify
from gene_dossier.tools import harmonizome as hz
from gene_dossier.tools.harmonizome_section4a import (
    CURATED_TF_DATASET_ORDER,
    PARSER_VERSION,
    PREDICTED_TF_DATASET_ORDER,
    SUPPLEMENTARY_SCOPE,
    collect_section_4a_harmonizome,
    gene_page_url,
)
from gene_dossier.workflow import DossierState, WorkflowTransientContext

logger = logging.getLogger(__name__)

SECTION_TF = "Transcription factors that drive the gene’s expression"
SUBSECTION_4A = "Harmonizome Integrated Knowledge About Genes & Proteins"

STATUS_SUCCESS = "success"
STATUS_SOURCE_UNAVAILABLE = "source_unavailable"
STATUS_PARTIAL = "partial"
STATUS_FAILED = "failed"
STATUS_NO_ASSOCIATIONS = "no_associations"
STATUS_GENE_MISMATCH = "gene_mismatch"

SCIENTIFIC_CAVEAT = (
    "Harmonizome associations link this gene to transcription-factor datasets; "
    "they do not by themselves prove that a listed factor drives expression of "
    "the gene in a specific tissue or disease context."
)

ENCODE_BLURB = (
    "ENCODE Transcription Factor Binding Site Profiles and Targets summarize "
    "ChIP-seq-derived factor–target relationships from the ENCODE encyclopedia."
)
CHEA_BLURB = (
    "ChEA Transcription Factor Binding Site Profiles and Targets compile "
    "ChIP enrichment analysis results from published experiments."
)
MOTIFMAP_BLURB = (
    "MotifMap Predicted Transcription Factor Targets combine motif scanning "
    "with comparative genomics to nominate candidate regulators."
)
JASPAR_BLURB = (
    "JASPAR Predicted Transcription Factor Targets are based on position "
    "frequency matrices from the JASPAR database."
)


@dataclass(frozen=True)
class Section4aConfig:
    output_root: str | Path | None = None
    max_displayed_curated_associations: int = 14
    max_displayed_predicted_associations: int = 25

    def __post_init__(self) -> None:
        if int(self.max_displayed_curated_associations) < 1:
            raise ValueError("max_displayed_curated_associations must be >= 1")
        if int(self.max_displayed_predicted_associations) < 1:
            raise ValueError("max_displayed_predicted_associations must be >= 1")


def _evidence(
    *,
    dossier_run_id: str,
    gene_symbol: str,
    fact_type: str,
    key: str,
    value: dict[str, Any],
    display_text: str,
    api_run_id: str | None = None,
    raw_artifact_id: str | None = None,
) -> EvidenceRecord:
    return EvidenceRecord(
        source_id=make_source_id(
            hz.SOURCE_NAME,
            gene_symbol,
            AssertionType.transcription_factor_association,
            key,
        ),
        dossier_run_id=dossier_run_id,
        gene_symbol=gene_symbol,
        official_symbol=gene_symbol,
        section=SECTION_TF,
        subsection=SUBSECTION_4A,
        source_name=hz.SOURCE_NAME,
        source_type=SourceType.expression_database,
        assertion_type=AssertionType.transcription_factor_association,
        fact_type=fact_type,
        evidence_grade=EvidenceGrade.B,
        value=value,
        display_text=display_text,
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )


def _blank(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null"}:
        return ""
    return text


def _autosize(ws: Any, max_width: int = 48) -> None:
    for idx, column in enumerate(ws.columns, start=1):
        length = 0
        for cell in column:
            if cell.value is None:
                continue
            length = max(length, min(max_width, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(idx)].width = max(12, length + 2)


def write_harmonizome_workbook(
    path: Path,
    *,
    gene_symbol: str,
    collection: dict[str, Any],
    audit_meta: dict[str, Any],
) -> None:
    """Write supplementary XLSX. Caller must compute SHA-256 after this returns.

    Never writes the workbook's own final SHA-256 into any cell.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Gene symbol", gene_symbol),
        ("Official symbol", collection.get("official_symbol") or gene_symbol),
        ("Gene page URL", collection.get("gene_page_url") or gene_page_url(gene_symbol)),
        ("Supplementary scope", SUPPLEMENTARY_SCOPE),
        ("Parser version", PARSER_VERSION),
        ("Total Harmonizome association count", collection.get("total_association_count")),
        ("In-scope association count", collection.get("in_scope_association_count")),
        ("Curated association count", collection.get("curated_total")),
        ("Predicted association count", collection.get("predicted_total")),
        ("Displayed curated count", collection.get("displayed_curated_count")),
        ("Displayed predicted count", collection.get("displayed_predicted_count")),
        ("Out-of-scope association count", collection.get("out_of_scope_total")),
        ("Generation timestamp", audit_meta.get("generation_timestamp")),
        ("Raw Harmonizome response SHA-256", audit_meta.get("raw_response_sha256")),
        ("Source artifact ID", audit_meta.get("raw_artifact_id")),
        ("ApiRun ID", audit_meta.get("api_run_id")),
    ]
    for name in CURATED_TF_DATASET_ORDER:
        summary_rows.append(
            (f"Curated count: {name}", (collection.get("curated_counts") or {}).get(name, 0))
        )
    for name in PREDICTED_TF_DATASET_ORDER:
        summary_rows.append(
            (
                f"Predicted count: {name}",
                (collection.get("predicted_counts") or {}).get(name, 0),
            )
        )
    ws.append(["Field", "Value"])
    for row in summary_rows:
        ws.append(list(row))
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.freeze_panes = "A2"
    _autosize(ws)

    # Curated Associations — all allowlisted curated rows
    ws_c = wb.create_sheet("Curated Associations")
    ws_c.append(
        [
            "Gene",
            "Association",
            "Dataset",
            "Tissue/Cells",
            "Organism",
            "Genome Build",
            "PubMed ID",
            "Gene set name",
            "Gene set URL",
            "Parse status",
            "Source order",
            "Displayed",
            "Displayed Rank",
            "Selection Reason",
            "Raw Artifact ID",
            "ApiRun ID",
            "Organism (audit)",
            "Organism derivation",
            "Threshold value",
            "Standardized value",
        ]
    )
    for name in CURATED_TF_DATASET_ORDER:
        rows = [
            r
            for r in (collection.get("curated_records") or [])
            if r.get("dataset_name") == name
        ]
        rows.sort(key=lambda r: int(r.get("source_order") or 0))
        for r in rows:
            ws_c.append(
                [
                    gene_symbol,
                    _blank(r.get("association")),
                    _blank(r.get("dataset_name")),
                    _blank(r.get("tissue_cells")),
                    _blank(r.get("organism")),
                    _blank(r.get("genome_build")),
                    _blank(r.get("pubmed_id")),
                    _blank(r.get("gene_set_name")),
                    _blank(r.get("href")),
                    _blank(r.get("parse_status")),
                    r.get("source_order"),
                    bool(r.get("displayed")),
                    r.get("displayed_rank"),
                    _blank(r.get("selection_reason")),
                    _blank(audit_meta.get("raw_artifact_id")),
                    _blank(audit_meta.get("api_run_id")),
                    _blank(r.get("organism_audit")),
                    _blank(r.get("organism_derivation")),
                    r.get("threshold_value"),
                    r.get("standardized_value"),
                ]
            )
    for cell in ws_c[1]:
        cell.font = Font(bold=True)
    ws_c.freeze_panes = "A2"
    _autosize(ws_c)

    # Predicted Associations
    ws_p = wb.create_sheet("Predicted Associations")
    ws_p.append(
        [
            "Gene",
            "Predicted Association",
            "Dataset",
            "Gene set name",
            "Gene set URL",
            "Parse status",
            "Source order",
            "Displayed",
            "Displayed Rank",
            "Selection Reason",
            "Raw Artifact ID",
            "ApiRun ID",
            "Threshold value",
            "Standardized value",
        ]
    )
    for name in PREDICTED_TF_DATASET_ORDER:
        rows = [
            r
            for r in (collection.get("predicted_records") or [])
            if r.get("dataset_name") == name
        ]
        rows.sort(key=lambda r: int(r.get("source_order") or 0))
        for r in rows:
            ws_p.append(
                [
                    gene_symbol,
                    _blank(r.get("association")),
                    _blank(r.get("dataset_name")),
                    _blank(r.get("gene_set_name")),
                    _blank(r.get("href")),
                    _blank(r.get("parse_status")),
                    r.get("source_order"),
                    bool(r.get("displayed")),
                    r.get("displayed_rank"),
                    _blank(r.get("selection_reason")),
                    _blank(audit_meta.get("raw_artifact_id")),
                    _blank(audit_meta.get("api_run_id")),
                    r.get("threshold_value"),
                    r.get("standardized_value"),
                ]
            )
    for cell in ws_p[1]:
        cell.font = Font(bold=True)
    ws_p.freeze_panes = "A2"
    _autosize(ws_p)

    # Out-of-Scope Dataset Summary
    ws_o = wb.create_sheet("Out-of-Scope Dataset Summary")
    ws_o.append(["Dataset name", "Association count"])
    for row in collection.get("out_of_scope_summary") or []:
        ws_o.append([row.get("dataset_name"), row.get("association_count")])
    for cell in ws_o[1]:
        cell.font = Font(bold=True)
    ws_o.freeze_panes = "A2"
    _autosize(ws_o)

    # Audit — never include workbook final SHA-256
    ws_a = wb.create_sheet("Audit")
    audit_rows = [
        ("Supplementary scope", SUPPLEMENTARY_SCOPE),
        ("Parser version", PARSER_VERSION),
        ("Generation timestamp", audit_meta.get("generation_timestamp")),
        ("Raw Harmonizome response SHA-256", audit_meta.get("raw_response_sha256")),
        ("Source artifact ID", audit_meta.get("raw_artifact_id")),
        ("ApiRun ID", audit_meta.get("api_run_id")),
        ("Requested URL", audit_meta.get("requested_url")),
        ("Final URL", audit_meta.get("final_url")),
        ("Gene page URL", collection.get("gene_page_url") or gene_page_url(gene_symbol)),
        ("Scientific status", audit_meta.get("scientific_status")),
        ("Presentation status", audit_meta.get("presentation_status")),
    ]
    ws_a.append(["Field", "Value"])
    for row in audit_rows:
        ws_a.append(list(row))
    for cell in ws_a[1]:
        cell.font = Font(bold=True)
    ws_a.freeze_panes = "A2"
    _autosize(ws_a)

    wb.save(path)
    wb.close()


def evaluate_section_4a_complete(
    *,
    status: dict[str, Any],
    attempt_dir: Path | None = None,
    html_path: Path | None = None,
    pdf_path: Path | None = None,
    presentation_blocks: list[Any] | None = None,
    major_html: str | None = None,
) -> dict[str, Any]:
    """Evaluate Section 4a complete acceptance with per-check results."""
    import json as _json
    import re as _re

    rendering = dict(status.get("rendering_status") or status)
    summary = dict(status.get("summary") or {})
    audit = dict(status.get("audit") or {})
    scientific = str(rendering.get("scientific_status") or "")
    presentation = str(rendering.get("presentation_status") or "")
    curated = int(summary.get("curated_total") or 0)
    predicted = int(summary.get("predicted_total") or 0)
    attempt = Path(attempt_dir) if attempt_dir else Path(str(audit.get("gene_attempt_dir") or ""))
    checks: dict[str, Any] = {}

    def _ok(name: str, passed: bool, detail: Any = None) -> bool:
        checks[name] = {"passed": bool(passed), "detail": detail}
        return bool(passed)

    all_pass = True
    all_pass &= _ok("scientific_status_success", scientific == STATUS_SUCCESS, scientific)
    all_pass &= _ok(
        "presentation_status_success", presentation == STATUS_SUCCESS, presentation
    )
    all_pass &= _ok(
        "association_total_nonzero",
        (curated + predicted) >= 1,
        {"curated_total": curated, "predicted_total": predicted},
    )

    curated_path = attempt / "curated_associations.json" if attempt else None
    predicted_path = attempt / "predicted_associations.json" if attempt else None
    curated_records: list[Any] = []
    predicted_records: list[Any] = []
    if curated_path and curated_path.is_file():
        curated_records = list(
            (_json.loads(curated_path.read_text(encoding="utf-8")).get("records") or [])
        )
    if predicted_path and predicted_path.is_file():
        predicted_records = list(
            (
                _json.loads(predicted_path.read_text(encoding="utf-8")).get("records")
                or []
            )
        )
    all_pass &= _ok(
        "curated_total_matches_records",
        curated == len(curated_records),
        {"curated_total": curated, "record_count": len(curated_records)},
    )
    all_pass &= _ok(
        "predicted_total_matches_records",
        predicted == len(predicted_records),
        {"predicted_total": predicted, "record_count": len(predicted_records)},
    )

    xlsx_name = summary.get("supplementary_xlsx")
    xlsx_path = (
        attempt / "supplementary" / str(xlsx_name)
        if attempt and xlsx_name
        else None
    )
    stored_sha = summary.get("supplementary_xlsx_sha256") or audit.get(
        "supplementary_xlsx_sha256"
    )
    xlsx_exists = bool(xlsx_path and xlsx_path.is_file())
    all_pass &= _ok("supplementary_xlsx_exists", xlsx_exists, str(xlsx_path))
    all_pass &= _ok(
        "supplementary_xlsx_sha_present", bool(stored_sha), stored_sha
    )
    recalculated = sha256_file(xlsx_path) if xlsx_exists else None
    all_pass &= _ok(
        "supplementary_xlsx_sha_matches",
        bool(stored_sha) and recalculated == stored_sha,
        {"stored": stored_sha, "recalculated": recalculated},
    )

    html_ok = bool(html_path and Path(html_path).is_file())
    pdf_ok = bool(pdf_path and Path(pdf_path).is_file())
    all_pass &= _ok("html_render_exists", html_ok, str(html_path) if html_path else None)
    all_pass &= _ok("pdf_render_exists", pdf_ok, str(pdf_path) if pdf_path else None)

    html_text = ""
    if html_ok:
        html_text = Path(html_path).read_text(encoding="utf-8", errors="replace")
    elif major_html:
        html_text = major_html

    major_count = len(
        _re.findall(
            r'<h2 class="major-heading"[^>]*>\s*4\.\s*Transcription factors',
            html_text,
        )
    )
    sub_count = len(
        _re.findall(
            r'<h3 class="sub-heading"[^>]*>\s*a\.\s*Harmonizome Integrated Knowledge',
            html_text,
        )
    )
    all_pass &= _ok("major_4_heading_once", major_count == 1, major_count)
    all_pass &= _ok("subsection_4a_heading_once", sub_count == 1, sub_count)

    roles = [str(getattr(b, "presentation_role", None) or "") for b in (presentation_blocks or [])]
    if not roles and html_text:
        # Fallback role presence via known polished lines when blocks unavailable.
        roles = []
    curated_count_roles = roles.count("section_4a_curated_count")
    predicted_count_roles = roles.count("section_4a_predicted_count")
    curated_table_roles = roles.count("section_4a_curated_table")
    predicted_table_roles = roles.count("section_4a_predicted_table")
    supp_roles = roles.count("section_4a_supplementary_note")
    if presentation_blocks is not None:
        all_pass &= _ok("curated_count_role_once", curated_count_roles == 1, curated_count_roles)
        all_pass &= _ok(
            "predicted_count_role_once", predicted_count_roles == 1, predicted_count_roles
        )
        if curated > 0:
            all_pass &= _ok(
                "curated_table_role_at_least_once",
                curated_table_roles >= 1,
                curated_table_roles,
            )
        else:
            checks["curated_table_role_at_least_once"] = {
                "passed": True,
                "detail": "skipped_zero_curated",
            }
        if predicted > 0:
            all_pass &= _ok(
                "predicted_table_role_at_least_once",
                predicted_table_roles >= 1,
                predicted_table_roles,
            )
        else:
            checks["predicted_table_role_at_least_once"] = {
                "passed": True,
                "detail": "skipped_zero_predicted",
            }
        all_pass &= _ok("supplementary_note_role_once", supp_roles == 1, supp_roles)
    else:
        # Soft presence checks from HTML prose when blocks not supplied.
        all_pass &= _ok(
            "curated_count_role_once",
            "transcription factor associations" in html_text,
            "html_fallback",
        )
        all_pass &= _ok(
            "predicted_count_role_once",
            "predicted transcription factor associations" in html_text,
            "html_fallback",
        )
        if curated > 0:
            all_pass &= _ok(
                "curated_table_role_at_least_once",
                "<table" in html_text,
                "html_fallback",
            )
        if predicted > 0:
            all_pass &= _ok(
                "predicted_table_role_at_least_once",
                "Predicted Association" in html_text or "<table" in html_text,
                "html_fallback",
            )
        all_pass &= _ok(
            "supplementary_note_role_once",
            "Supplementary Material" in html_text,
            "html_fallback",
        )

    return {
        "complete": bool(all_pass),
        "section_4a_complete": bool(all_pass),
        "checks": checks,
        "scientific_status": scientific,
        "presentation_status": presentation,
        "curated_total": curated,
        "predicted_total": predicted,
        "html_render_status": "success" if html_ok else "failed",
        "pdf_render_status": "success" if pdf_ok else "failed",
        "supplementary_xlsx_sha256": stored_sha,
    }


def accept_section_4a_report(
    paths: Section4aPaths,
    *,
    gene_symbol: str,
    attempt_dir: Path,
    acceptance: dict[str, Any],
    artifacts: dict[str, Any] | None = None,
    promote_existing: bool = False,
) -> Path | None:
    """Accept a complete Section 4a attempt; preserve prior success unless promoted."""
    pointer = paths.accepted_gene_pointer(gene_symbol)
    if pointer.is_file():
        try:
            import json

            existing = json.loads(pointer.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            existing = {}
        prior = dict(existing.get("acceptance") or {})
        prior_ok = prior.get("section_4a_complete") is True
        if prior_ok and not promote_existing:
            return None
    return accept_gene_report(
        paths,
        gene_symbol=gene_symbol,
        attempt_dir=attempt_dir,
        acceptance=acceptance,
        artifacts=artifacts or {},
    )


def _harmonizome_meta(tool_result: Any) -> dict[str, Any]:
    data = tool_result.data if tool_result is not None and isinstance(tool_result.data, dict) else {}
    meta = data.get("_harmonizome_meta")
    return dict(meta) if isinstance(meta, dict) else {}


def _find_existing_gene_associations_tool_result(state: DossierState, gene: str) -> Any | None:
    target = gene.strip().upper()
    for tr in state.get("tool_results") or []:
        if getattr(tr, "source_name", None) != hz.SOURCE_NAME:
            continue
        if not getattr(tr, "success", False):
            continue
        params = dict(getattr(tr, "request_params", None) or {})
        show = str(params.get("showAssociations") or "").lower()
        if show not in {"true", "1"}:
            # fetch_tf_associations nests params; accept gene_symbol match + associations
            if not isinstance(getattr(tr, "data", None), dict):
                continue
        if str(getattr(tr, "gene_symbol", "") or "").strip().upper() != target:
            continue
        data = tr.data if isinstance(tr.data, dict) else {}
        if "associations" in data or (isinstance(data.get("raw"), dict) and "associations" in (data.get("raw") or {})):
            return tr
        endpoint = str(getattr(tr, "endpoint_name", "") or "")
        if endpoint in {"gene_associations", "fetch_gene_associations"}:
            return tr
    return None


def _persist_harmonizome_raw_response(
    *,
    tool_result: Any,
    response_bytes: bytes | None,
    dossier_run_id: str,
    gene_symbol: str,
    settings: Settings,
    persist_db: bool,
    existing_api_runs: list[Any],
    existing_raw: list[dict[str, Any]],
) -> tuple[Any, dict[str, Any] | None, str | None]:
    """Persist exact response bytes when available; reuse existing ApiRun/raw if present."""
    from gene_dossier.section_1c import (
        _persist_artifact_bytes,
        _save_api_run_failure,
        _tool_result_to_api_run,
    )

    meta = _harmonizome_meta(tool_result)
    raw_sha = meta.get("response_body_sha256")
    # Reuse matching ApiRun / raw artifact by sha or request identity.
    for api in existing_api_runs:
        if getattr(api, "source_name", None) != hz.SOURCE_NAME:
            continue
        if str(getattr(api, "gene_symbol", "") or "").strip().upper() != gene_symbol.strip().upper():
            continue
        for raw in existing_raw:
            if raw.get("api_run_id") == getattr(api, "id", None) or raw.get("id") == getattr(api, "raw_artifact_id", None):
                sha = raw.get("sha256") or raw.get("content_hash") or raw.get("expected_sha256")
                if raw_sha and sha == raw_sha:
                    return api, raw, raw_sha
                if raw_sha is None and sha:
                    return api, raw, sha

    api = _tool_result_to_api_run(
        tool_result, dossier_run_id=dossier_run_id, gene_symbol=gene_symbol
    )
    if not tool_result.success:
        _save_api_run_failure(api, persist_db=persist_db)
        return api, None, raw_sha

    content = response_bytes
    if content is None and raw_sha and isinstance(tool_result.data, dict):
        # Fallback only when bytes unavailable: still persist parsed payload separately.
        content = None
    raw_meta = None
    if content is not None:
        def _validate_bytes(blob: bytes) -> dict[str, Any]:
            return {
                "media_type": meta.get("content_type") or "application/json",
                "byte_size": len(blob),
            }

        _artifact, raw_meta = _persist_artifact_bytes(
            dossier_run_id=dossier_run_id,
            source_name=hz.SOURCE_NAME,
            content=content,
            extension="json",
            artifact_type="json",
            filename_hint=f"harmonizome-gene-{slugify(gene_symbol.lower())}-raw",
            settings=settings,
            api_run=api,
            persist_db=persist_db,
            notes={
                "artifact_class": "external_raw",
                "artifact_origin": "harmonizome_gene_associations",
                "artifact_role": "gene_associations_raw_bytes",
                "source_url": meta.get("requested_url") or tool_result.request_url,
                "retrieval_method": "api_bytes",
                "response_body_sha256": raw_sha,
                "decoding_method": meta.get("decoding_method"),
                "utf8_replacement_char_count": meta.get("utf8_replacement_char_count"),
                "exact_raw_bytes": True,
            },
            validate=_validate_bytes,
        )
        if raw_meta is not None:
            raw_meta["sha256"] = raw_meta.get("sha256") or raw_meta.get("content_hash") or raw_sha
    else:
        # Preserve prior behavior as secondary path when bytes missing.
        from gene_dossier.section_1c import _persist_tool_result_json

        # Strip meta-only noise? Keep payload; SHA recorded from meta when present.
        api, raw_meta = _persist_tool_result_json(
            tr=tool_result,
            dossier_run_id=dossier_run_id,
            gene_symbol=gene_symbol,
            settings=settings,
            persist_db=persist_db,
            filename_hint=f"harmonizome-gene-{slugify(gene_symbol.lower())}",
        )
        if raw_sha and raw_meta is not None:
            raw_meta = {**raw_meta, "parsed_json_only": True, "response_body_sha256": raw_sha}
    return api, raw_meta, raw_sha or (raw_meta or {}).get("sha256")


def node_generate_section_4a_derived_artifacts(
    state: DossierState,
    *,
    settings: Settings | None = None,
    persist_db: bool = True,
    transient: WorkflowTransientContext | None = None,
    config: Section4aConfig | None = None,
) -> DossierState:
    """Generate Section 4a attempt artifacts and return ``section_4a_status``.

    Does not update accepted gene pointers. Never calls gene_set.
    """
    from gene_dossier.tools.harmonizome import normalize_harmonizome_request_identity
    from gene_dossier.workflow import (
        bind_workflow_transient,
        reset_workflow_transient,
    )

    cfg = settings or get_settings()
    section_cfg = config or Section4aConfig()
    run_type = state.get("run_type")
    selected_keys = list(state.get("selected_section_keys") or [])
    if run_type != "section_bundle" or "4a" not in selected_keys:
        return state

    gene = str(state.get("gene_symbol") or "").strip()
    if not gene:
        gene_ids = state.get("gene_ids") or {}
        gene = str(gene_ids.get("symbol") or gene_ids.get("gene_symbol") or "").strip()
    if not gene:
        errors = list(state.get("errors") or [])
        errors.append("Section 4a requires a resolved gene_symbol")
        return {**state, "errors": errors}

    existing_tr = _find_existing_gene_associations_tool_result(state, gene)
    bind_token = bind_workflow_transient(transient)
    try:
        # Prefer payload.raw for fetch_tf_associations wrappers.
        reuse_tr = existing_tr
        if (
            reuse_tr is not None
            and isinstance(reuse_tr.data, dict)
            and "associations" not in reuse_tr.data
            and isinstance(reuse_tr.data.get("raw"), dict)
        ):
            from gene_dossier.models import ToolResult as _TR

            raw_payload = dict(reuse_tr.data.get("raw") or {})
            reuse_tr = _TR(
                source_name=reuse_tr.source_name,
                endpoint_name="gene_associations",
                success=True,
                gene_symbol=gene,
                request_url=reuse_tr.request_url,
                request_params={
                    "gene_symbol": gene,
                    "showAssociations": "true",
                },
                status_code=reuse_tr.status_code,
                data=raw_payload,
            )

        collected = collect_section_4a_harmonizome(
            gene,
            max_displayed_curated=section_cfg.max_displayed_curated_associations,
            max_displayed_predicted=section_cfg.max_displayed_predicted_associations,
            settings=cfg,
            transient=transient,
            tool_result=reuse_tr,
        )
    finally:
        reset_workflow_transient(bind_token)

    tool_result = collected.get("tool_result")
    run_id = str(state.get("dossier_run_id") or "")
    api_runs = list(state.get("api_runs") or [])
    raw_artifacts = list(state.get("raw_artifacts") or [])
    api_run = None
    raw_meta = None
    raw_sha = None
    response_bytes = None
    if tool_result is not None and transient is not None:
        meta = _harmonizome_meta(tool_result)
        identity = meta.get("request_identity") or normalize_harmonizome_request_identity(
            method="GET",
            url_path=f"{hz.HARMONIZOME_BASE}/gene/{gene}",
            gene_symbol=gene,
            query_params={"showAssociations": "true"},
        )
        cached = transient.get_cached_request(str(identity))
        if isinstance(cached, dict):
            response_bytes = cached.get("response_bytes")
    if tool_result is not None:
        api_run, raw_meta, raw_sha = _persist_harmonizome_raw_response(
            tool_result=tool_result,
            response_bytes=response_bytes if isinstance(response_bytes, (bytes, bytearray)) else None,
            dossier_run_id=run_id,
            gene_symbol=gene,
            settings=cfg,
            persist_db=persist_db,
            existing_api_runs=api_runs,
            existing_raw=raw_artifacts,
        )
        if api_run is not None and all(getattr(a, "id", None) != getattr(api_run, "id", None) for a in api_runs):
            api_runs.append(api_run)
        if raw_meta and all(r.get("id") != raw_meta.get("id") for r in raw_artifacts):
            raw_artifacts.append(raw_meta)

    paths = paths_for(section_cfg.output_root or cfg.output_path)
    attempt_dir = paths.new_gene_attempt(gene, run_id=run_id)
    supp_dir = attempt_dir / "supplementary"
    supp_dir.mkdir(parents=True, exist_ok=True)

    official = str(collected.get("official_symbol") or gene)
    collection = collected.get("collection") or {
        "query_gene": gene,
        "official_symbol": official,
        "parser_version": PARSER_VERSION,
        "supplementary_scope": SUPPLEMENTARY_SCOPE,
        "total_association_count": 0,
        "in_scope_association_count": 0,
        "curated_total": 0,
        "predicted_total": 0,
        "displayed_curated_count": 0,
        "displayed_predicted_count": 0,
        "curated_counts": {n: 0 for n in CURATED_TF_DATASET_ORDER},
        "predicted_counts": {n: 0 for n in PREDICTED_TF_DATASET_ORDER},
        "curated_records": [],
        "predicted_records": [],
        "curated_display": [],
        "predicted_display": [],
        "out_of_scope_summary": [],
        "out_of_scope_total": 0,
        "gene_page_url": gene_page_url(official),
        "max_displayed_curated": section_cfg.max_displayed_curated_associations,
        "max_displayed_predicted": section_cfg.max_displayed_predicted_associations,
    }

    scientific = str(collected.get("scientific_status") or STATUS_SOURCE_UNAVAILABLE)
    presentation = str(collected.get("presentation_status") or STATUS_FAILED)
    generation_ts = datetime.now(timezone.utc).isoformat()
    hz_meta = _harmonizome_meta(tool_result)

    api_run_id = getattr(api_run, "id", None) if api_run else None
    raw_artifact_id = (raw_meta or {}).get("id") if raw_meta else None
    audit_meta = {
        "generation_timestamp": generation_ts,
        "raw_response_sha256": raw_sha or hz_meta.get("response_body_sha256"),
        "raw_artifact_id": raw_artifact_id,
        "api_run_id": api_run_id,
        "requested_url": hz_meta.get("requested_url"),
        "final_url": hz_meta.get("final_url"),
        "redirect_history": hz_meta.get("redirect_history"),
        "content_type": hz_meta.get("content_type"),
        "retrieved_at": hz_meta.get("retrieved_at"),
        "response_byte_length": hz_meta.get("response_byte_length"),
        "decoding_method": hz_meta.get("decoding_method"),
        "utf8_replacement_char_count": hz_meta.get("utf8_replacement_char_count"),
        "decode_warning": hz_meta.get("decode_warning"),
        "scientific_status": scientific,
        "presentation_status": presentation,
        "parser_version": PARSER_VERSION,
        "supplementary_scope": SUPPLEMENTARY_SCOPE,
    }

    xlsx_name = f"{official}_Harmonizome.xlsx"
    xlsx_path = supp_dir / xlsx_name
    workbook_sha = None
    if scientific == STATUS_SUCCESS and (
        int(collection.get("curated_total") or 0)
        + int(collection.get("predicted_total") or 0)
        > 0
    ):
        write_harmonizome_workbook(
            xlsx_path,
            gene_symbol=official,
            collection=collection,
            audit_meta=audit_meta,
        )
        workbook_sha = sha256_file(xlsx_path)

    summary_payload = {
        "gene_symbol": gene,
        "official_symbol": official,
        "scientific_status": scientific,
        "presentation_status": presentation,
        "total_association_count": collection.get("total_association_count"),
        "in_scope_association_count": collection.get("in_scope_association_count"),
        "curated_total": collection.get("curated_total"),
        "predicted_total": collection.get("predicted_total"),
        "displayed_curated_count": collection.get("displayed_curated_count"),
        "displayed_predicted_count": collection.get("displayed_predicted_count"),
        "curated_counts": collection.get("curated_counts"),
        "predicted_counts": collection.get("predicted_counts"),
        "curated_display": collection.get("curated_display"),
        "predicted_display": collection.get("predicted_display"),
        "out_of_scope_total": collection.get("out_of_scope_total"),
        "gene_page_url": collection.get("gene_page_url"),
        "parser_version": PARSER_VERSION,
        "supplementary_scope": SUPPLEMENTARY_SCOPE,
        "scientific_caveat": SCIENTIFIC_CAVEAT,
        "source_blurbs": {
            "encode": ENCODE_BLURB,
            "chea": CHEA_BLURB,
            "motifmap": MOTIFMAP_BLURB,
            "jaspar": JASPAR_BLURB,
        },
        "supplementary_xlsx": xlsx_name if workbook_sha else None,
        "supplementary_xlsx_sha256": workbook_sha,
        "raw_response_sha256": audit_meta.get("raw_response_sha256"),
        "max_displayed_curated": collection.get("max_displayed_curated"),
        "max_displayed_predicted": collection.get("max_displayed_predicted"),
        "presentation_item_key": f"harmonizome-{slugify(official.lower())}",
    }
    write_json_atomic(attempt_dir / "summary.json", summary_payload)
    write_json_atomic(
        attempt_dir / "curated_associations.json",
        {"records": collection.get("curated_records") or []},
    )
    write_json_atomic(
        attempt_dir / "predicted_associations.json",
        {"records": collection.get("predicted_records") or []},
    )
    write_json_atomic(
        attempt_dir / "out_of_scope_summary.json",
        {"datasets": collection.get("out_of_scope_summary") or []},
    )

    evidence_records = list(state.get("evidence_records") or [])
    summary_rec = _evidence(
        dossier_run_id=run_id,
        gene_symbol=official,
        fact_type="section_4a_summary",
        key="section-4a-summary",
        value={
            **summary_payload,
            "attempt_dir": str(attempt_dir),
            "api_run_id": api_run_id,
            "raw_artifact_id": raw_artifact_id,
        },
        display_text=(
            f"{official} Harmonizome TF associations "
            f"(curated={summary_payload['curated_total']}, "
            f"predicted={summary_payload['predicted_total']})."
        ),
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )
    _append_evidence(evidence_records, summary_rec, persist_db=persist_db)

    for i, row in enumerate(collection.get("curated_display") or []):
        rec = _evidence(
            dossier_run_id=run_id,
            gene_symbol=official,
            fact_type="section_4a_curated_row",
            key=f"section-4a-curated-{i}",
            value={**row, "api_run_id": api_run_id, "raw_artifact_id": raw_artifact_id},
            display_text=_blank(row.get("association")) or f"curated-{i}",
            api_run_id=api_run_id,
            raw_artifact_id=raw_artifact_id,
        )
        _append_evidence(evidence_records, rec, persist_db=persist_db)

    for i, row in enumerate(collection.get("predicted_display") or []):
        rec = _evidence(
            dossier_run_id=run_id,
            gene_symbol=official,
            fact_type="section_4a_predicted_row",
            key=f"section-4a-predicted-{i}",
            value={**row, "api_run_id": api_run_id, "raw_artifact_id": raw_artifact_id},
            display_text=_blank(row.get("predicted_association")) or f"predicted-{i}",
            api_run_id=api_run_id,
            raw_artifact_id=raw_artifact_id,
        )
        _append_evidence(evidence_records, rec, persist_db=persist_db)

    if workbook_sha:
        supp_rec = _evidence(
            dossier_run_id=run_id,
            gene_symbol=official,
            fact_type="section_4a_supplementary_workbook",
            key="section-4a-supplementary-xlsx",
            value={
                "filename": xlsx_name,
                "relative_path": f"supplementary/{xlsx_name}",
                "local_artifact_path": str(xlsx_path),
                "sha256": workbook_sha,
                "supplementary_scope": SUPPLEMENTARY_SCOPE,
                "api_run_id": api_run_id,
                "raw_artifact_id": raw_artifact_id,
                "media_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            },
            display_text=f"Supplementary Material ({xlsx_name})",
            api_run_id=api_run_id,
            raw_artifact_id=raw_artifact_id,
        )
        _append_evidence(evidence_records, supp_rec, persist_db=persist_db)

    write_json_atomic(
        attempt_dir / "evidence.json",
        {
            "summary_evidence_id": summary_rec.id,
            "supplementary_xlsx_sha256": workbook_sha,
            "raw_response_sha256": audit_meta.get("raw_response_sha256"),
            "api_run_id": api_run_id,
            "raw_artifact_id": raw_artifact_id,
        },
    )

    manifest = {
        "gene_symbol": gene,
        "official_symbol": official,
        "attempt_dir": str(attempt_dir),
        "scientific_status": scientific,
        "presentation_status": presentation,
        "supplementary_xlsx": xlsx_name if workbook_sha else None,
        "supplementary_xlsx_sha256": workbook_sha,
        "raw_response_sha256": audit_meta.get("raw_response_sha256"),
        "api_run_id": api_run_id,
        "raw_artifact_id": raw_artifact_id,
        "parser_version": PARSER_VERSION,
        "supplementary_scope": SUPPLEMENTARY_SCOPE,
        "created_at": generation_ts,
        "requested_url": audit_meta.get("requested_url"),
        "final_url": audit_meta.get("final_url"),
        "decoding_method": audit_meta.get("decoding_method"),
    }
    write_json_atomic(attempt_dir / MANIFEST_FILENAME, manifest)

    coverage = list(state.get("coverage") or [])
    if scientific == STATUS_SUCCESS:
        cov_status = SourceStatus.success
    elif scientific == STATUS_NO_ASSOCIATIONS:
        cov_status = SourceStatus.partial
    else:
        cov_status = SourceStatus.failed
    coverage.append(
        SourceCoverageResult(
            dossier_run_id=run_id,
            source_name=hz.SOURCE_NAME,
            status=cov_status,
            evidence_record_count=1
            + len(collection.get("curated_display") or [])
            + len(collection.get("predicted_display") or [])
            + (1 if workbook_sha else 0),
            error_message=None
            if scientific == STATUS_SUCCESS
            else f"section_4a scientific_status={scientific}",
            notes=f"presentation_status={presentation}",
            report_sections_supported=[SECTION_TF],
        )
    )

    section_status = {
        "section_key": "4a",
        "summary": summary_payload,
        "rendering_status": {
            "scientific_status": scientific,
            "presentation_status": presentation,
            "overall_status": (
                STATUS_SUCCESS
                if scientific == STATUS_SUCCESS and presentation == STATUS_SUCCESS
                else presentation
                if scientific == STATUS_SUCCESS
                else scientific
            ),
        },
        "audit": {
            "gene_attempt_dir": str(attempt_dir),
            "artifacts": {
                "summary.json": "summary.json",
                "curated_associations.json": "curated_associations.json",
                "predicted_associations.json": "predicted_associations.json",
                "out_of_scope_summary.json": "out_of_scope_summary.json",
                "evidence.json": "evidence.json",
                MANIFEST_FILENAME: MANIFEST_FILENAME,
                **(
                    {f"supplementary/{xlsx_name}": f"supplementary/{xlsx_name}"}
                    if workbook_sha
                    else {}
                ),
            },
            "supplementary_xlsx_sha256": workbook_sha,
            "raw_response_sha256": audit_meta.get("raw_response_sha256"),
            "api_run_id": api_run_id,
            "raw_artifact_id": raw_artifact_id,
            "requested_url": audit_meta.get("requested_url"),
            "final_url": audit_meta.get("final_url"),
            "redirect_history": audit_meta.get("redirect_history"),
            "decoding_method": audit_meta.get("decoding_method"),
            "utf8_replacement_char_count": audit_meta.get("utf8_replacement_char_count"),
            "decode_warning": audit_meta.get("decode_warning"),
            "parser_version": PARSER_VERSION,
            "supplementary_scope": SUPPLEMENTARY_SCOPE,
        },
    }
    write_json_atomic(attempt_dir / "section_4a_status.json", section_status)

    tool_results = list(state.get("tool_results") or [])
    reused_existing = existing_tr is not None
    if tool_result is not None and not reused_existing:
        tool_results.append(tool_result)

    return {
        **state,
        "evidence_records": evidence_records,
        "coverage": coverage,
        "api_runs": api_runs,
        "raw_artifacts": raw_artifacts,
        "section_4a_status": section_status,
        "tool_results": tool_results,
    }



__all__ = [
    "CHEA_BLURB",
    "ENCODE_BLURB",
    "JASPAR_BLURB",
    "MOTIFMAP_BLURB",
    "SCIENTIFIC_CAVEAT",
    "SECTION_TF",
    "STATUS_FAILED",
    "STATUS_GENE_MISMATCH",
    "STATUS_NO_ASSOCIATIONS",
    "STATUS_PARTIAL",
    "STATUS_SOURCE_UNAVAILABLE",
    "STATUS_SUCCESS",
    "SUBSECTION_4A",
    "Section4aConfig",
    "accept_section_4a_report",
    "evaluate_section_4a_complete",
    "node_generate_section_4a_derived_artifacts",
    "write_harmonizome_workbook",
]
