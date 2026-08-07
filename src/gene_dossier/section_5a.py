"""Bundle-only Section 5a STRING functional network.

Owns STRING identifier resolution, bounded ``/network`` JSON, optional official
PNG, ``get_link``, supplementary XLSX, and section statuses. Accepted gene
pointers are written only after post-render evaluation in ``section_bundle``.
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from gene_dossier.config import Settings, get_settings
from gene_dossier.models import (
    AssertionType,
    EvidenceGrade,
    EvidenceRecord,
    SourceCoverageResult,
    SourceStatus,
    SourceType,
)
from gene_dossier.section_1c import _append_evidence
from gene_dossier.section_5a_sources import (
    MANIFEST_FILENAME,
    Section5aPaths,
    accept_gene_report,
    paths_for,
    sha256_file,
    write_json_atomic,
)
from gene_dossier.source_ids import make_source_id, slugify
from gene_dossier.tools import string_db as sd
from gene_dossier.workflow import DossierState, WorkflowTransientContext

logger = logging.getLogger(__name__)

SECTION_PPI = "Protein-protein interaction (PPI) partners"
SUBSECTION_5A = "STRING"
PARSER_VERSION = "section_5a_string_v1"
SUPPLEMENTARY_SCOPE = (
    "bounded STRING functional network returned by /api/json/network with "
    "configured add_nodes and required_score"
)
DIRECT_QUERY_ASSOCIATION_DEF = (
    "Direct Query Association: an edge in the bounded STRING functional network "
    "where one endpoint is the resolved query protein. It does not imply "
    "experimentally demonstrated physical binding."
)

STATUS_SUCCESS = "success"
STATUS_SOURCE_UNAVAILABLE = "source_unavailable"
STATUS_PARTIAL = "partial"
STATUS_FAILED = "failed"
STATUS_NO_INTERACTIONS = "no_interactions"
STATUS_IDENTIFIER_NOT_FOUND = "identifier_not_found"
STATUS_IDENTIFIER_AMBIGUOUS = "identifier_ambiguous"
STATUS_UNAVAILABLE = "unavailable"
STATUS_NOT_ATTEMPTED = "not_attempted_optional"

SCIENTIFIC_INTRO = (
    "STRING is a database of known and predicted protein-protein interactions "
    "(PPIs). Functional networks integrate physical and functional associations "
    "supported by experimental evidence, curated databases, genomic context, "
    "co-expression, text mining and other evidence channels. A functional "
    "association does not necessarily indicate direct physical binding."
)

STRING_HOMEPAGE = "https://string-db.org/"


@dataclass(frozen=True)
class Section5aConfig:
    output_root: str | Path | None = None
    species_taxon_id: int = sd.SPECIES_HUMAN
    add_nodes: int = sd.DEFAULT_ADD_NODES
    required_score: int = sd.DEFAULT_REQUIRED_SCORE
    network_type: str = sd.DEFAULT_NETWORK_TYPE
    attempt_network_figure: bool = True
    include_interaction_partners_diagnostic: bool = False

    def __post_init__(self) -> None:
        if int(self.add_nodes) != 30:
            raise ValueError("add_nodes must be exactly 30 for Section 5a v1")
        if int(self.required_score) < 0 or int(self.required_score) > 1000:
            raise ValueError("required_score must be between 0 and 1000")
        if int(self.species_taxon_id) <= 0:
            raise ValueError("species_taxon_id must be a positive NCBI taxon id")
        if str(self.network_type or "").strip().lower() != "functional":
            raise ValueError("network_type must be 'functional' for Section 5a v1")


@dataclass(frozen=True)
class NetworkEdgeRecord:
    string_id_a: str
    string_id_b: str
    preferred_name_a: str
    preferred_name_b: str
    ncbi_taxon_id: int
    combined_score: float
    neighborhood_score: float | None
    fusion_score: float | None
    phylogenetic_profile_score: float | None
    coexpression_score: float | None
    experimental_score: float | None
    database_score: float | None
    textmining_score: float | None
    source_order: int
    direct_query_association: bool
    raw_artifact_id: str | None = None
    api_run_id: str | None = None

    def undirected_key(self) -> tuple[str, str]:
        a, b = self.string_id_a, self.string_id_b
        return (a, b) if a <= b else (b, a)

    def as_dict(self) -> dict[str, Any]:
        return {
            "string_id_a": self.string_id_a,
            "string_id_b": self.string_id_b,
            "preferred_name_a": self.preferred_name_a,
            "preferred_name_b": self.preferred_name_b,
            "ncbi_taxon_id": self.ncbi_taxon_id,
            "combined_score": self.combined_score,
            "neighborhood_score": self.neighborhood_score,
            "fusion_score": self.fusion_score,
            "phylogenetic_profile_score": self.phylogenetic_profile_score,
            "coexpression_score": self.coexpression_score,
            "experimental_score": self.experimental_score,
            "database_score": self.database_score,
            "textmining_score": self.textmining_score,
            "source_order": self.source_order,
            "direct_query_association": self.direct_query_association,
            "raw_artifact_id": self.raw_artifact_id,
            "api_run_id": self.api_run_id,
        }


def _evidence(
    *,
    dossier_run_id: str,
    gene_symbol: str,
    fact_type: str,
    key: str,
    value: dict[str, Any],
    display_text: str,
    api_run_id: str | None = None,
    raw_artifact_id: str | None = None,
) -> EvidenceRecord:
    return EvidenceRecord(
        source_id=make_source_id(
            sd.SOURCE_NAME,
            gene_symbol,
            AssertionType.ppi,
            key,
        ),
        dossier_run_id=dossier_run_id,
        gene_symbol=gene_symbol,
        official_symbol=gene_symbol,
        section=SECTION_PPI,
        subsection=SUBSECTION_5A,
        source_name=sd.SOURCE_NAME,
        source_type=SourceType.interaction_database,
        assertion_type=AssertionType.ppi,
        fact_type=fact_type,
        evidence_grade=EvidenceGrade.C,
        value=value,
        display_text=display_text,
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )


def _blank(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null"}:
        return ""
    return text


def _autosize(ws: Any, max_width: int = 48) -> None:
    for idx, column in enumerate(ws.columns, start=1):
        length = 0
        for cell in column:
            if cell.value is None:
                continue
            length = max(length, min(max_width, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(idx)].width = max(12, length + 2)


def _score_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def map_network_edge(
    row: dict[str, Any],
    *,
    source_order: int,
    query_string_id: str,
    species_taxon_id: int,
    required_score: int,
    raw_artifact_id: str | None = None,
    api_run_id: str | None = None,
) -> tuple[NetworkEdgeRecord | None, str | None]:
    """Map one STRING /network row; return (record, warning)."""
    sid_a = str(row.get("stringId_A") or row.get("string_id_a") or "").strip()
    sid_b = str(row.get("stringId_B") or row.get("string_id_b") or "").strip()
    if not sid_a or not sid_b:
        return None, "missing string ids"
    if sid_a == sid_b:
        return None, "self-edge rejected"
    taxon = sd.normalize_taxon_id(row.get("ncbiTaxonId") or row.get("ncbi_taxon_id"))
    if taxon is None:
        return None, "invalid ncbiTaxonId"
    if taxon != int(species_taxon_id):
        return None, f"taxon mismatch: {taxon}"
    prefix = f"{species_taxon_id}."
    if not sid_a.startswith(prefix) or not sid_b.startswith(prefix):
        return None, "string id species prefix mismatch"
    combined = _score_or_none(row.get("score"))
    if combined is None:
        return None, "missing combined score"
    if not (0.0 <= combined <= 1.0):
        return None, "combined score out of range"
    threshold = float(required_score) / 1000.0
    if combined + 1e-12 < threshold:
        return None, f"below threshold {threshold:.3f}"

    partials = {
        "neighborhood_score": _score_or_none(row.get("nscore")),
        "fusion_score": _score_or_none(row.get("fscore")),
        "phylogenetic_profile_score": _score_or_none(row.get("pscore")),
        "coexpression_score": _score_or_none(row.get("ascore")),
        "experimental_score": _score_or_none(row.get("escore")),
        "database_score": _score_or_none(row.get("dscore")),
        "textmining_score": _score_or_none(row.get("tscore")),
    }
    for name, val in partials.items():
        if val is not None and not (0.0 <= val <= 1.0):
            return None, f"{name} out of range"

    direct = query_string_id in {sid_a, sid_b}
    record = NetworkEdgeRecord(
        string_id_a=sid_a,
        string_id_b=sid_b,
        preferred_name_a=str(row.get("preferredName_A") or row.get("preferred_name_a") or ""),
        preferred_name_b=str(row.get("preferredName_B") or row.get("preferred_name_b") or ""),
        ncbi_taxon_id=int(taxon),
        combined_score=float(combined),
        neighborhood_score=partials["neighborhood_score"],
        fusion_score=partials["fusion_score"],
        phylogenetic_profile_score=partials["phylogenetic_profile_score"],
        coexpression_score=partials["coexpression_score"],
        experimental_score=partials["experimental_score"],
        database_score=partials["database_score"],
        textmining_score=partials["textmining_score"],
        source_order=source_order,
        direct_query_association=direct,
        raw_artifact_id=raw_artifact_id,
        api_run_id=api_run_id,
    )
    return record, None


def parse_network_edge(
    row: dict[str, Any],
    *,
    source_order: int,
    query_string_id: str,
    species_taxon_id: int,
    required_score: int,
    raw_artifact_id: str | None = None,
    api_run_id: str | None = None,
) -> tuple[NetworkEdgeRecord | None, str | None]:
    """Alias for :func:`map_network_edge` (normalize_taxon_id on ncbiTaxonId)."""
    return map_network_edge(
        row,
        source_order=source_order,
        query_string_id=query_string_id,
        species_taxon_id=species_taxon_id,
        required_score=required_score,
        raw_artifact_id=raw_artifact_id,
        api_run_id=api_run_id,
    )


def canonicalize_network(
    rows: list[dict[str, Any]],
    *,
    query_string_id: str,
    species_taxon_id: int,
    required_score: int,
    raw_artifact_id: str | None = None,
    api_run_id: str | None = None,
) -> dict[str, Any]:
    """Validate, undirected-canonicalize, and compute network statistics."""
    warnings: list[str] = []
    raw_records: list[NetworkEdgeRecord] = []
    for idx, row in enumerate(rows):
        rec, warn = parse_network_edge(
            row,
            source_order=idx + 1,
            query_string_id=query_string_id,
            species_taxon_id=species_taxon_id,
            required_score=required_score,
            raw_artifact_id=raw_artifact_id,
            api_run_id=api_run_id,
        )
        if warn:
            warnings.append(f"row {idx + 1}: {warn}")
        if rec is not None:
            raw_records.append(rec)

    by_key: dict[tuple[str, str], NetworkEdgeRecord] = {}
    duplicate_count = 0
    for rec in raw_records:
        key = rec.undirected_key()
        if key in by_key:
            duplicate_count += 1
            continue
        by_key[key] = rec
    canonical = list(by_key.values())
    if duplicate_count:
        warnings.append(f"dropped {duplicate_count} duplicate undirected edge(s)")

    nodes: dict[str, str] = {}
    for rec in canonical:
        nodes[rec.string_id_a] = rec.preferred_name_a
        nodes[rec.string_id_b] = rec.preferred_name_b
    if query_string_id and query_string_id not in nodes and not canonical:
        # empty network still records query if known from resolution
        pass

    degree: dict[str, int] = {sid: 0 for sid in nodes}
    for rec in canonical:
        degree[rec.string_id_a] = degree.get(rec.string_id_a, 0) + 1
        degree[rec.string_id_b] = degree.get(rec.string_id_b, 0) + 1

    direct = [r for r in canonical if r.direct_query_association]
    neighbor = [r for r in canonical if not r.direct_query_association]
    scores = [r.combined_score for r in canonical]

    node_rows = [
        {
            "string_id": sid,
            "preferred_name": name,
            "is_query_protein": sid == query_string_id,
            "degree_within_bounded_network": degree.get(sid, 0),
        }
        for sid, name in sorted(nodes.items(), key=lambda kv: (kv[1].upper(), kv[0]))
    ]

    return {
        "canonical_edges": [r.as_dict() for r in canonical],
        "raw_accepted_edges": [r.as_dict() for r in raw_records],
        "direct_query_partners": [
            r.as_dict()
            for r in sorted(
                direct,
                key=lambda r: (-r.combined_score, r.string_id_b if r.string_id_a == query_string_id else r.string_id_a),
            )
        ],
        "nodes": node_rows,
        "warnings": warnings,
        "stats": {
            "unique_node_count": len(nodes),
            "unique_edge_count": len(canonical),
            "direct_query_edge_count": len(direct),
            "neighbor_neighbor_edge_count": len(neighbor),
            "min_combined_score": min(scores) if scores else None,
            "max_combined_score": max(scores) if scores else None,
            "duplicate_undirected_edge_count": duplicate_count,
        },
    }


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True, default=str)
    return value


def write_string_workbook(
    path: Path,
    *,
    gene_symbol: str,
    summary: dict[str, Any],
    canonical_edges: list[dict[str, Any]],
    direct_partners: list[dict[str, Any]],
    nodes: list[dict[str, Any]],
    audit_meta: dict[str, Any],
    query_string_id: str,
) -> None:
    """Write supplementary XLSX. Caller computes SHA-256 after this returns."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Gene", gene_symbol),
        ("STRING ID", query_string_id),
        ("STRING Version", summary.get("string_version") or sd.STRING_VERSION),
        ("Species", summary.get("taxon_name") or "Homo sapiens"),
        ("Taxon ID", summary.get("species_taxon_id")),
        ("Network Type", summary.get("network_type")),
        ("Required Score", summary.get("required_score")),
        ("Added Nodes", summary.get("add_nodes")),
        ("Unique Node Count", summary.get("unique_node_count")),
        ("Unique Association Count", summary.get("unique_edge_count")),
        ("Direct Query Association Count", summary.get("direct_query_edge_count")),
        (
            "Neighbor-Neighbor Association Count",
            summary.get("neighbor_neighbor_edge_count"),
        ),
        ("Minimum Combined Score", summary.get("min_combined_score")),
        ("Maximum Combined Score", summary.get("max_combined_score")),
        ("STRING Network URL", summary.get("string_network_url")),
        ("Structured Raw Artifact ID", audit_meta.get("network_raw_artifact_id")),
        ("Structured ApiRun ID", audit_meta.get("network_api_run_id")),
        ("Structured Response SHA-256", audit_meta.get("network_response_sha256")),
        ("Figure Artifact ID", audit_meta.get("figure_raw_artifact_id")),
        ("Figure ApiRun ID", audit_meta.get("figure_api_run_id")),
        ("Figure SHA-256", audit_meta.get("figure_sha256")),
        ("Supplementary Scope", SUPPLEMENTARY_SCOPE),
        ("Direct Query Association Definition", DIRECT_QUERY_ASSOCIATION_DEF),
        ("Generation Timestamp", audit_meta.get("generation_timestamp")),
    ]
    ws.append(["Field", "Value"])
    for row in summary_rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    _autosize(ws)

    ws_a = wb.create_sheet("Network Associations")
    ws_a.append(
        [
            "Gene",
            "STRING ID A",
            "Protein A",
            "STRING ID B",
            "Protein B",
            "Taxon ID",
            "Combined Score",
            "Neighborhood Score",
            "Fusion Score",
            "Phylogenetic Profile Score",
            "Coexpression Score",
            "Experimental Score",
            "Database Score",
            "Text-mining Score",
            "Direct Query Association",
            "Source Order",
            "Raw Artifact ID",
            "ApiRun ID",
        ]
    )
    for r in canonical_edges:
        ws_a.append(
            [
                gene_symbol,
                r.get("string_id_a"),
                r.get("preferred_name_a"),
                r.get("string_id_b"),
                r.get("preferred_name_b"),
                r.get("ncbi_taxon_id"),
                r.get("combined_score"),
                r.get("neighborhood_score"),
                r.get("fusion_score"),
                r.get("phylogenetic_profile_score"),
                r.get("coexpression_score"),
                r.get("experimental_score"),
                r.get("database_score"),
                r.get("textmining_score"),
                bool(r.get("direct_query_association")),
                r.get("source_order"),
                r.get("raw_artifact_id"),
                r.get("api_run_id"),
            ]
        )
    for cell in ws_a[1]:
        cell.font = Font(bold=True)
    ws_a.freeze_panes = "A2"
    _autosize(ws_a)

    ws_d = wb.create_sheet("Direct Query Partners")
    ws_d.append(
        [
            "Gene",
            "Query STRING ID",
            "Partner STRING ID",
            "Partner Protein",
            "Combined Score",
            "Neighborhood Score",
            "Fusion Score",
            "Phylogenetic Profile Score",
            "Coexpression Score",
            "Experimental Score",
            "Database Score",
            "Text-mining Score",
        ]
    )
    for r in direct_partners:
        if r.get("string_id_a") == query_string_id:
            partner_id = r.get("string_id_b")
            partner_name = r.get("preferred_name_b")
        else:
            partner_id = r.get("string_id_a")
            partner_name = r.get("preferred_name_a")
        ws_d.append(
            [
                gene_symbol,
                query_string_id,
                partner_id,
                partner_name,
                r.get("combined_score"),
                r.get("neighborhood_score"),
                r.get("fusion_score"),
                r.get("phylogenetic_profile_score"),
                r.get("coexpression_score"),
                r.get("experimental_score"),
                r.get("database_score"),
                r.get("textmining_score"),
            ]
        )
    for cell in ws_d[1]:
        cell.font = Font(bold=True)
    ws_d.freeze_panes = "A2"
    _autosize(ws_d)

    ws_n = wb.create_sheet("Network Nodes")
    ws_n.append(
        ["STRING ID", "Preferred Name", "Is Query Protein", "Degree Within Bounded Network"]
    )
    for n in nodes:
        ws_n.append(
            [
                n.get("string_id"),
                n.get("preferred_name"),
                bool(n.get("is_query_protein")),
                n.get("degree_within_bounded_network"),
            ]
        )
    for cell in ws_n[1]:
        cell.font = Font(bold=True)
    ws_n.freeze_panes = "A2"
    _autosize(ws_n)

    ws_audit = wb.create_sheet("Audit")
    audit_rows = [
        ("Supplementary scope", SUPPLEMENTARY_SCOPE),
        ("Parser version", PARSER_VERSION),
        ("STRING version", sd.STRING_VERSION),
        ("STRING base URL", sd.STRING_BASE_URL),
        ("Generation timestamp", audit_meta.get("generation_timestamp")),
        ("Resolve requested URL", audit_meta.get("resolve_requested_url")),
        ("Resolve final URL", audit_meta.get("resolve_final_url")),
        ("Network requested URL", audit_meta.get("network_requested_url")),
        ("Network final URL", audit_meta.get("network_final_url")),
        ("Network response SHA-256", audit_meta.get("network_response_sha256")),
        ("Figure requested URL", audit_meta.get("figure_requested_url")),
        ("Figure final URL", audit_meta.get("figure_final_url")),
        ("Figure SHA-256", audit_meta.get("figure_sha256")),
        ("Link requested URL", audit_meta.get("link_requested_url")),
        ("Link final URL", audit_meta.get("link_final_url")),
        ("Scientific status", audit_meta.get("scientific_status")),
        ("Visual status", audit_meta.get("visual_status")),
        ("Presentation status", audit_meta.get("presentation_status")),
        ("Network configuration", _excel_cell_value(audit_meta.get("network_configuration"))),
        ("Figure configuration", _excel_cell_value(audit_meta.get("figure_configuration"))),
        ("Validation warnings", "; ".join(audit_meta.get("warnings") or [])),
        ("Direct Query Association Definition", DIRECT_QUERY_ASSOCIATION_DEF),
    ]
    ws_audit.append(["Field", "Value"])
    for row in audit_rows:
        ws_audit.append(list(row))
    for cell in ws_audit[1]:
        cell.font = Font(bold=True)
    ws_audit.freeze_panes = "A2"
    _autosize(ws_audit)

    wb.save(path)
    wb.close()


def evaluate_section_5a_complete(
    *,
    status: dict[str, Any],
    attempt_dir: Path | None = None,
    html_path: Path | None = None,
    pdf_path: Path | None = None,
    presentation_blocks: list[Any] | None = None,
    major_html: str | None = None,
) -> dict[str, Any]:
    """Evaluate Section 5a complete acceptance with per-check results."""
    import json as _json
    import re as _re

    rendering = dict(status.get("rendering_status") or status)
    summary = dict(status.get("summary") or {})
    audit = dict(status.get("audit") or {})
    scientific = str(rendering.get("scientific_status") or "")
    presentation = str(rendering.get("presentation_status") or "")
    visual = str(rendering.get("visual_status") or "")
    attempt = Path(attempt_dir) if attempt_dir else Path(str(audit.get("gene_attempt_dir") or ""))
    checks: dict[str, Any] = {}

    def _ok(name: str, passed: bool, detail: Any = None) -> bool:
        checks[name] = {"passed": bool(passed), "detail": detail}
        return bool(passed)

    all_pass = True
    all_pass &= _ok("scientific_status_success", scientific == STATUS_SUCCESS, scientific)
    all_pass &= _ok(
        "presentation_status_success", presentation == STATUS_SUCCESS, presentation
    )
    all_pass &= _ok(
        "identifier_resolved",
        bool(summary.get("resolved_string_id")),
        summary.get("resolved_string_id"),
    )

    net_path = attempt / "network_associations.json" if attempt else None
    nodes_path = attempt / "network_nodes.json" if attempt else None
    edges: list[Any] = []
    nodes: list[Any] = []
    if net_path and net_path.is_file():
        try:
            edges = list((_json.loads(net_path.read_text(encoding="utf-8")).get("edges")) or [])
        except (OSError, ValueError):
            edges = []
    if nodes_path and nodes_path.is_file():
        try:
            nodes = list((_json.loads(nodes_path.read_text(encoding="utf-8")).get("nodes")) or [])
        except (OSError, ValueError):
            nodes = []

    unique_edges = int(summary.get("unique_edge_count") or 0)
    unique_nodes = int(summary.get("unique_node_count") or 0)
    direct = int(summary.get("direct_query_edge_count") or 0)
    neighbor = int(summary.get("neighbor_neighbor_edge_count") or 0)
    all_pass &= _ok("network_associations_exist", bool(net_path and net_path.is_file()), str(net_path))
    all_pass &= _ok(
        "network_response_sha_present",
        bool(summary.get("network_response_sha256") or audit.get("network_response_sha256")),
        summary.get("network_response_sha256"),
    )
    all_pass &= _ok(
        "unique_edge_count_matches_file",
        unique_edges == len(edges),
        {"summary": unique_edges, "file": len(edges)},
    )
    all_pass &= _ok(
        "unique_node_count_matches_file",
        unique_nodes == len(nodes),
        {"summary": unique_nodes, "file": len(nodes)},
    )
    all_pass &= _ok(
        "direct_plus_neighbor_equals_unique",
        direct + neighbor == unique_edges,
        {"direct": direct, "neighbor": neighbor, "unique": unique_edges},
    )

    xlsx_name = summary.get("supplementary_xlsx")
    xlsx_path = (
        attempt / "supplementary" / str(xlsx_name)
        if attempt and xlsx_name
        else None
    )
    all_pass &= _ok("xlsx_exists", bool(xlsx_path and xlsx_path.is_file()), str(xlsx_path))
    stored_sha = summary.get("supplementary_xlsx_sha256")
    all_pass &= _ok("xlsx_sha_present", bool(stored_sha), stored_sha)
    if xlsx_path and xlsx_path.is_file() and stored_sha:
        all_pass &= _ok(
            "xlsx_sha_matches",
            sha256_file(xlsx_path) == stored_sha,
            stored_sha,
        )
    else:
        all_pass &= _ok("xlsx_sha_matches", False, None)

    figure_attempted = visual != STATUS_NOT_ATTEMPTED
    if figure_attempted:
        all_pass &= _ok("visual_status_success", visual == STATUS_SUCCESS, visual)
        fig_rel = (audit.get("artifacts") or {}).get("network_figure") or summary.get(
            "network_figure_relative_path"
        )
        fig_path = attempt / str(fig_rel) if attempt and fig_rel else None
        all_pass &= _ok("png_exists", bool(fig_path and fig_path.is_file()), str(fig_path))
        all_pass &= _ok("png_sha_present", bool(summary.get("network_figure_sha256")), None)
        if fig_path and fig_path.is_file():
            try:
                from PIL import Image
                import io

                Image.open(io.BytesIO(fig_path.read_bytes())).verify()
                decoded = True
            except Exception as exc:  # noqa: BLE001
                decoded = False
                checks["png_decodes"] = {"passed": False, "detail": str(exc)}
            else:
                checks["png_decodes"] = {"passed": True, "detail": None}
            all_pass &= bool(decoded)
        else:
            all_pass &= _ok("png_decodes", False, None)

    roles = [
        str(getattr(b, "presentation_role", None) or "")
        for b in (presentation_blocks or [])
    ]
    all_pass &= _ok("intro_role_once", roles.count("section_5a_intro") == 1, roles.count("section_5a_intro"))
    all_pass &= _ok(
        "supplementary_note_once",
        roles.count("section_5a_supplementary_note") == 1,
        roles.count("section_5a_supplementary_note"),
    )
    if figure_attempted and visual == STATUS_SUCCESS:
        all_pass &= _ok(
            "network_figure_role_once",
            roles.count("section_5a_network_figure") == 1,
            roles.count("section_5a_network_figure"),
        )
        all_pass &= _ok(
            "network_legend_role_once",
            roles.count("section_5a_network_legend") == 1,
            roles.count("section_5a_network_legend"),
        )

    html_text = ""
    if html_path and Path(html_path).is_file():
        html_text = Path(html_path).read_text(encoding="utf-8", errors="replace")
    elif major_html:
        html_text = major_html
    all_pass &= _ok("html_exists", bool(html_text), None)
    if html_text:
        major_count = len(_re.findall(r">\s*5\.\s*Protein-protein interaction", html_text))
        sub_count = len(_re.findall(r">\s*a\.\s*STRING\s*<", html_text))
        all_pass &= _ok("major_5_heading_once", major_count == 1, major_count)
        all_pass &= _ok("subsection_5a_heading_once", sub_count == 1, sub_count)
    all_pass &= _ok(
        "pdf_exists",
        bool(pdf_path and Path(pdf_path).is_file()),
        str(pdf_path) if pdf_path else None,
    )

    return {
        "complete": bool(all_pass),
        "section_5a_complete": bool(all_pass),
        "checks": checks,
    }


def accept_section_5a_report(
    paths: Section5aPaths,
    *,
    gene_symbol: str,
    attempt_dir: Path,
    acceptance: dict[str, Any],
    artifacts: dict[str, Any] | None = None,
    promote_existing: bool = False,
) -> Path | None:
    """Accept a complete Section 5a attempt; preserve prior success unless promoted."""
    pointer = paths.accepted_gene_pointer(gene_symbol)
    if pointer.is_file():
        try:
            import json

            existing = json.loads(pointer.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            existing = {}
        prior = dict(existing.get("acceptance") or {})
        prior_ok = prior.get("section_5a_complete") is True
        if prior_ok and not promote_existing:
            return None
    return accept_gene_report(
        paths,
        gene_symbol=gene_symbol,
        attempt_dir=attempt_dir,
        acceptance=acceptance,
        artifacts=artifacts or {},
    )


def _string_meta(tool_result: Any, transient: Any | None = None) -> dict[str, Any]:
    return sd.lookup_string_meta(tool_result, transient)


def _persist_string_raw(
    *,
    tool_result: Any,
    response_bytes: bytes | None,
    dossier_run_id: str,
    gene_symbol: str,
    settings: Settings,
    persist_db: bool,
    filename_hint: str,
    artifact_role: str,
    existing_api_runs: list[Any],
    existing_raw: list[dict[str, Any]],
    extension: str = "json",
    media_type: str = "application/json",
    transient: Any | None = None,
) -> tuple[Any, dict[str, Any] | None, str | None]:
    from gene_dossier.section_1c import (
        _persist_artifact_bytes,
        _save_api_run_failure,
        _tool_result_to_api_run,
    )

    meta = _string_meta(tool_result, transient)
    if isinstance(tool_result.data, dict) and tool_result.data.get("sha256"):
        raw_sha = tool_result.data.get("sha256")
    else:
        raw_sha = meta.get("response_body_sha256")
    for api in existing_api_runs:
        if getattr(api, "source_name", None) != sd.SOURCE_NAME:
            continue
        if str(getattr(api, "gene_symbol", "") or "").strip().upper() != gene_symbol.strip().upper():
            continue
        for raw in existing_raw:
            if raw.get("api_run_id") == getattr(api, "id", None) or raw.get("id") == getattr(
                api, "raw_artifact_id", None
            ):
                sha = raw.get("sha256") or raw.get("content_hash") or raw.get("expected_sha256")
                if raw_sha and sha == raw_sha:
                    return api, raw, raw_sha

    api = _tool_result_to_api_run(
        tool_result, dossier_run_id=dossier_run_id, gene_symbol=gene_symbol
    )
    if not tool_result.success:
        _save_api_run_failure(api, persist_db=persist_db)
        return api, None, raw_sha

    content = response_bytes
    if content is None and isinstance(tool_result.data, dict):
        maybe = tool_result.data.get("content")
        if isinstance(maybe, (bytes, bytearray)):
            content = bytes(maybe)
    raw_meta = None
    if content is not None:

        def _validate_bytes(blob: bytes) -> dict[str, Any]:
            data = tool_result.data if isinstance(tool_result.data, dict) else {}
            return {
                "media_type": meta.get("content_type")
                or data.get("content_type")
                or media_type,
                "byte_size": len(blob),
            }

        _artifact, raw_meta = _persist_artifact_bytes(
            dossier_run_id=dossier_run_id,
            source_name=sd.SOURCE_NAME,
            content=content,
            extension=extension,
            artifact_type=extension,
            filename_hint=filename_hint,
            settings=settings,
            api_run=api,
            persist_db=persist_db,
            notes={
                "artifact_class": "external_raw",
                "artifact_origin": "string_db",
                "artifact_role": artifact_role,
                "source_url": meta.get("requested_url") or tool_result.request_url,
                "retrieval_method": "api_bytes",
                "response_body_sha256": raw_sha,
                "exact_raw_bytes": True,
                "string_version": sd.STRING_VERSION,
            },
            validate=_validate_bytes,
        )
        if raw_meta is not None:
            raw_meta["sha256"] = raw_meta.get("sha256") or raw_meta.get("content_hash") or raw_sha
    else:
        from gene_dossier.section_1c import _persist_tool_result_json

        api, raw_meta = _persist_tool_result_json(
            tr=tool_result,
            dossier_run_id=dossier_run_id,
            gene_symbol=gene_symbol,
            settings=settings,
            persist_db=persist_db,
            filename_hint=filename_hint,
        )
        if raw_sha and raw_meta is not None:
            raw_meta = {**raw_meta, "parsed_json_only": True, "response_body_sha256": raw_sha}
    return api, raw_meta, raw_sha or (raw_meta or {}).get("sha256")


def _cached_bytes(transient: WorkflowTransientContext | None, tool_result: Any) -> bytes | None:
    if transient is None or tool_result is None:
        return None
    meta = _string_meta(tool_result, transient)
    identity = meta.get("request_identity")
    if not identity:
        return None
    cached = transient.get_cached_request(str(identity))
    if isinstance(cached, dict):
        blob = cached.get("response_bytes")
        if isinstance(blob, (bytes, bytearray)):
            return bytes(blob)
    return None


def node_generate_section_5a_derived_artifacts(
    state: DossierState,
    *,
    settings: Settings | None = None,
    persist_db: bool = True,
    transient: WorkflowTransientContext | None = None,
    config: Section5aConfig | None = None,
) -> DossierState:
    """Generate Section 5a attempt artifacts and return ``section_5a_status``.

    Does not update accepted gene pointers. Does not call ``interaction_partners``
    unless ``include_interaction_partners_diagnostic`` is set.
    """
    from gene_dossier.workflow import bind_workflow_transient, reset_workflow_transient

    cfg = settings or get_settings()
    section_cfg = config or Section5aConfig()
    run_type = state.get("run_type")
    selected_keys = list(state.get("selected_section_keys") or [])
    if run_type != "section_bundle" or "5a" not in selected_keys:
        return state

    gene = str(state.get("gene_symbol") or "").strip()
    if not gene:
        gene_ids = state.get("gene_ids") or {}
        gene = str(gene_ids.get("symbol") or gene_ids.get("gene_symbol") or "").strip()
    if not gene:
        errors = list(state.get("errors") or [])
        errors.append("Section 5a requires a resolved gene_symbol")
        return {**state, "errors": errors}

    run_id = str(state.get("dossier_run_id") or "")
    api_runs = list(state.get("api_runs") or [])
    raw_artifacts = list(state.get("raw_artifacts") or [])
    tool_results = list(state.get("tool_results") or [])
    evidence_records = list(state.get("evidence_records") or [])
    warnings: list[str] = []

    bind_token = bind_workflow_transient(transient)
    try:
        resolve_tr = sd.resolve_string_identifier(
            gene,
            species=section_cfg.species_taxon_id,
            settings=cfg,
            transient=transient,
            caller_identity=sd.SECTION_5A_CALLER_IDENTITY,
        )
        tool_results.append(resolve_tr)

        scientific = STATUS_SOURCE_UNAVAILABLE
        visual = (
            STATUS_NOT_ATTEMPTED
            if not section_cfg.attempt_network_figure
            else STATUS_UNAVAILABLE
        )
        presentation = STATUS_FAILED
        string_id = None
        preferred_name = gene
        taxon_name = "Homo sapiens"
        network_tr = None
        image_tr = None
        link_tr = None
        canonical: dict[str, Any] = {
            "canonical_edges": [],
            "direct_query_partners": [],
            "nodes": [],
            "warnings": [],
            "stats": {
                "unique_node_count": 0,
                "unique_edge_count": 0,
                "direct_query_edge_count": 0,
                "neighbor_neighbor_edge_count": 0,
                "min_combined_score": None,
                "max_combined_score": None,
            },
        }
        string_network_url = None
        network_api = None
        network_raw = None
        network_sha = None
        figure_api = None
        figure_raw = None
        figure_sha = None
        figure_dims: dict[str, Any] = {}
        link_api = None
        link_raw = None

        if not resolve_tr.success:
            err = str(resolve_tr.error_type or "")
            if err == "not_found":
                scientific = STATUS_IDENTIFIER_NOT_FOUND
            elif err == "ambiguous":
                scientific = STATUS_IDENTIFIER_AMBIGUOUS
            else:
                scientific = STATUS_SOURCE_UNAVAILABLE
        else:
            resolve_data = resolve_tr.data or {}
            string_id = resolve_data.get("string_id")
            preferred_name = str(resolve_data.get("preferred_name") or gene)
            taxon_name = str(resolve_data.get("taxon_name") or "Homo sapiens")
            resolve_bytes = _cached_bytes(transient, resolve_tr)
            resolve_api, resolve_raw, _ = _persist_string_raw(
                tool_result=resolve_tr,
                response_bytes=resolve_bytes,
                dossier_run_id=run_id,
                gene_symbol=preferred_name,
                settings=cfg,
                persist_db=persist_db,
                filename_hint=f"string-get-string-ids-{slugify(preferred_name.lower())}",
                artifact_role="get_string_ids_raw_bytes",
                existing_api_runs=api_runs,
                existing_raw=raw_artifacts,
                transient=transient,
            )
            if resolve_api is not None and all(
                getattr(a, "id", None) != getattr(resolve_api, "id", None) for a in api_runs
            ):
                api_runs.append(resolve_api)
            if resolve_raw and all(r.get("id") != resolve_raw.get("id") for r in raw_artifacts):
                raw_artifacts.append(resolve_raw)

            network_tr = sd.fetch_network(
                str(string_id),
                gene_symbol=preferred_name,
                species=section_cfg.species_taxon_id,
                add_nodes=section_cfg.add_nodes,
                required_score=section_cfg.required_score,
                network_type=section_cfg.network_type,
                settings=cfg,
                transient=transient,
                caller_identity=sd.SECTION_5A_CALLER_IDENTITY,
            )
            tool_results.append(network_tr)
            if not network_tr.success:
                scientific = STATUS_SOURCE_UNAVAILABLE
            else:
                network_bytes = _cached_bytes(transient, network_tr)
                network_api, network_raw, network_sha = _persist_string_raw(
                    tool_result=network_tr,
                    response_bytes=network_bytes,
                    dossier_run_id=run_id,
                    gene_symbol=preferred_name,
                    settings=cfg,
                    persist_db=persist_db,
                    filename_hint=f"string-network-{slugify(preferred_name.lower())}",
                    artifact_role="network_raw_bytes",
                    existing_api_runs=api_runs,
                    existing_raw=raw_artifacts,
                    transient=transient,
                )
                if network_api is not None and all(
                    getattr(a, "id", None) != getattr(network_api, "id", None) for a in api_runs
                ):
                    api_runs.append(network_api)
                if network_raw and all(r.get("id") != network_raw.get("id") for r in raw_artifacts):
                    raw_artifacts.append(network_raw)

                rows = sd.extract_network_rows(network_tr.data)
                canonical = canonicalize_network(
                    rows,
                    query_string_id=str(string_id),
                    species_taxon_id=section_cfg.species_taxon_id,
                    required_score=section_cfg.required_score,
                    raw_artifact_id=(network_raw or {}).get("id") if network_raw else None,
                    api_run_id=getattr(network_api, "id", None) if network_api else None,
                )
                warnings.extend(list(canonical.get("warnings") or []))
                if int((canonical.get("stats") or {}).get("unique_edge_count") or 0) == 0:
                    scientific = STATUS_NO_INTERACTIONS
                else:
                    scientific = STATUS_SUCCESS

            if section_cfg.attempt_network_figure and string_id:
                image_tr = sd.fetch_network_image(
                    str(string_id),
                    gene_symbol=preferred_name,
                    species=section_cfg.species_taxon_id,
                    add_color_nodes=sd.DEFAULT_ADD_COLOR_NODES,
                    add_white_nodes=sd.DEFAULT_ADD_WHITE_NODES,
                    required_score=section_cfg.required_score,
                    network_type=section_cfg.network_type,
                    network_flavor=sd.DEFAULT_NETWORK_FLAVOR,
                    hide_disconnected_nodes=1,
                    settings=cfg,
                    transient=transient,
                    caller_identity=sd.SECTION_5A_CALLER_IDENTITY,
                )
                tool_results.append(image_tr)
                try:
                    if not image_tr.success:
                        raise ValueError(image_tr.error_message or "image request failed")
                    content = (image_tr.data or {}).get("content")
                    if not isinstance(content, (bytes, bytearray)):
                        content = _cached_bytes(transient, image_tr)
                    if not isinstance(content, (bytes, bytearray)):
                        raise ValueError("missing image bytes")
                    figure_dims = sd.validate_network_image_bytes(
                        bytes(content),
                        final_url=(image_tr.data or {}).get("final_url"),
                        content_type=(image_tr.data or {}).get("content_type"),
                    )
                    figure_sha = figure_dims.get("sha256")
                    figure_api, figure_raw, figure_sha = _persist_string_raw(
                        tool_result=image_tr,
                        response_bytes=bytes(content),
                        dossier_run_id=run_id,
                        gene_symbol=preferred_name,
                        settings=cfg,
                        persist_db=persist_db,
                        filename_hint=f"string-network-figure-{slugify(preferred_name.lower())}",
                        artifact_role="network_figure_png",
                        existing_api_runs=api_runs,
                        existing_raw=raw_artifacts,
                        extension="png",
                        media_type="image/png",
                        transient=transient,
                    )
                    if figure_api is not None and all(
                        getattr(a, "id", None) != getattr(figure_api, "id", None)
                        for a in api_runs
                    ):
                        api_runs.append(figure_api)
                    if figure_raw and all(
                        r.get("id") != figure_raw.get("id") for r in raw_artifacts
                    ):
                        raw_artifacts.append(figure_raw)
                    visual = STATUS_SUCCESS
                except Exception as exc:  # noqa: BLE001
                    warnings.append(f"network figure unavailable: {exc}")
                    visual = STATUS_UNAVAILABLE

            if string_id:
                link_tr = sd.fetch_network_link(
                    str(string_id),
                    gene_symbol=preferred_name,
                    species=section_cfg.species_taxon_id,
                    add_color_nodes=sd.DEFAULT_ADD_COLOR_NODES,
                    add_white_nodes=sd.DEFAULT_ADD_WHITE_NODES,
                    required_score=section_cfg.required_score,
                    network_type=section_cfg.network_type,
                    network_flavor=sd.DEFAULT_NETWORK_FLAVOR,
                    settings=cfg,
                    transient=transient,
                    caller_identity=sd.SECTION_5A_CALLER_IDENTITY,
                )
                tool_results.append(link_tr)
                if link_tr.success:
                    string_network_url = sd.extract_link_url(link_tr.data)
                    link_bytes = _cached_bytes(transient, link_tr)
                    link_api, link_raw, _ = _persist_string_raw(
                        tool_result=link_tr,
                        response_bytes=link_bytes,
                        dossier_run_id=run_id,
                        gene_symbol=preferred_name,
                        settings=cfg,
                        persist_db=persist_db,
                        filename_hint=f"string-get-link-{slugify(preferred_name.lower())}",
                        artifact_role="get_link_raw_bytes",
                        existing_api_runs=api_runs,
                        existing_raw=raw_artifacts,
                        transient=transient,
                    )
                    if link_api is not None and all(
                        getattr(a, "id", None) != getattr(link_api, "id", None) for a in api_runs
                    ):
                        api_runs.append(link_api)
                    if link_raw and all(r.get("id") != link_raw.get("id") for r in raw_artifacts):
                        raw_artifacts.append(link_raw)
                if not string_network_url:
                    warnings.append("STRING network link unavailable")

            if section_cfg.include_interaction_partners_diagnostic and string_id:
                diag = sd.interaction_partners(
                    str(string_id),
                    gene_symbol=preferred_name,
                    species=section_cfg.species_taxon_id,
                    required_score=section_cfg.required_score,
                    network_type=section_cfg.network_type,
                    settings=cfg,
                    transient=transient,
                )
                tool_results.append(diag)
                warnings.append(
                    "interaction_partners diagnostic requested; omitting limit is not unlimited"
                )

    finally:
        reset_workflow_transient(bind_token)

    paths = paths_for(section_cfg.output_root or cfg.output_path)
    attempt_dir = paths.new_gene_attempt(preferred_name if preferred_name else gene, run_id=run_id)
    supp_dir = attempt_dir / "supplementary"
    fig_dir = attempt_dir / "figures"
    supp_dir.mkdir(parents=True, exist_ok=True)
    fig_dir.mkdir(parents=True, exist_ok=True)

    stats = dict(canonical.get("stats") or {})
    generation_ts = datetime.now(timezone.utc).isoformat()
    resolve_meta = _string_meta(resolve_tr, transient)
    network_meta = _string_meta(network_tr, transient) if network_tr else {}
    image_meta = _string_meta(image_tr, transient) if image_tr else {}
    link_meta = _string_meta(link_tr, transient) if link_tr else {}

    figure_rel = None
    figure_local = None
    if visual == STATUS_SUCCESS and image_tr is not None:
        content = (image_tr.data or {}).get("content")
        if not isinstance(content, (bytes, bytearray)):
            content = _cached_bytes(transient, image_tr)
        if isinstance(content, (bytes, bytearray)):
            figure_name = f"{preferred_name}_STRING_network.png"
            figure_local = fig_dir / figure_name
            figure_local.write_bytes(bytes(content))
            figure_rel = f"figures/{figure_name}"
            figure_sha = figure_sha or sha256_file(figure_local)

    audit_meta = {
        "generation_timestamp": generation_ts,
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
        "warnings": warnings,
        "resolve_requested_url": resolve_meta.get("requested_url") or resolve_tr.request_url,
        "resolve_final_url": resolve_meta.get("final_url"),
        "network_requested_url": network_meta.get("requested_url")
        or (network_tr.request_url if network_tr else None),
        "network_final_url": network_meta.get("final_url"),
        "network_response_sha256": network_sha or network_meta.get("response_body_sha256"),
        "network_api_run_id": getattr(network_api, "id", None) if network_api else None,
        "network_raw_artifact_id": (network_raw or {}).get("id") if network_raw else None,
        "figure_requested_url": image_meta.get("requested_url")
        or (image_tr.request_url if image_tr else None),
        "figure_final_url": image_meta.get("final_url")
        or ((image_tr.data or {}).get("final_url") if image_tr else None),
        "figure_sha256": figure_sha,
        "figure_api_run_id": getattr(figure_api, "id", None) if figure_api else None,
        "figure_raw_artifact_id": (figure_raw or {}).get("id") if figure_raw else None,
        "link_requested_url": link_meta.get("requested_url")
        or (link_tr.request_url if link_tr else None),
        "link_final_url": link_meta.get("final_url"),
        "network_configuration": {
            "add_nodes": section_cfg.add_nodes,
            "required_score": section_cfg.required_score,
            "network_type": section_cfg.network_type,
            "species": section_cfg.species_taxon_id,
        },
        "figure_configuration": {
            "add_color_nodes": sd.DEFAULT_ADD_COLOR_NODES,
            "add_white_nodes": sd.DEFAULT_ADD_WHITE_NODES,
            "required_score": section_cfg.required_score,
            "network_flavor": sd.DEFAULT_NETWORK_FLAVOR,
            "network_type": section_cfg.network_type,
            "hide_disconnected_nodes": 1,
        },
    }

    xlsx_name = f"{preferred_name}_STRING.xlsx"
    xlsx_path = supp_dir / xlsx_name
    workbook_sha = None
    if scientific in {STATUS_SUCCESS, STATUS_NO_INTERACTIONS} and string_id:
        summary_for_xlsx = {
            "string_version": sd.STRING_VERSION,
            "taxon_name": taxon_name,
            "species_taxon_id": section_cfg.species_taxon_id,
            "network_type": section_cfg.network_type,
            "required_score": section_cfg.required_score,
            "add_nodes": section_cfg.add_nodes,
            "unique_node_count": stats.get("unique_node_count"),
            "unique_edge_count": stats.get("unique_edge_count"),
            "direct_query_edge_count": stats.get("direct_query_edge_count"),
            "neighbor_neighbor_edge_count": stats.get("neighbor_neighbor_edge_count"),
            "min_combined_score": stats.get("min_combined_score"),
            "max_combined_score": stats.get("max_combined_score"),
            "string_network_url": string_network_url,
        }
        try:
            write_string_workbook(
                xlsx_path,
                gene_symbol=preferred_name,
                summary=summary_for_xlsx,
                canonical_edges=list(canonical.get("canonical_edges") or []),
                direct_partners=list(canonical.get("direct_query_partners") or []),
                nodes=list(canonical.get("nodes") or []),
                audit_meta=audit_meta,
                query_string_id=str(string_id),
            )
            workbook_sha = sha256_file(xlsx_path)
            presentation = STATUS_SUCCESS if scientific == STATUS_SUCCESS else STATUS_PARTIAL
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"workbook failed: {exc}")
            presentation = STATUS_PARTIAL if scientific == STATUS_SUCCESS else STATUS_FAILED
    elif scientific == STATUS_SUCCESS:
        presentation = STATUS_PARTIAL
    else:
        presentation = STATUS_FAILED if scientific != STATUS_NO_INTERACTIONS else STATUS_PARTIAL

    audit_meta["presentation_status"] = presentation
    audit_meta["warnings"] = warnings

    summary_payload = {
        "gene_symbol": gene,
        "official_symbol": preferred_name,
        "preferred_name": preferred_name,
        "resolved_string_id": string_id,
        "species_taxon_id": section_cfg.species_taxon_id,
        "taxon_name": taxon_name,
        "string_version": sd.STRING_VERSION,
        "string_base_url": sd.STRING_BASE_URL,
        "network_type": section_cfg.network_type,
        "required_score": section_cfg.required_score,
        "add_nodes": section_cfg.add_nodes,
        "add_color_nodes": sd.DEFAULT_ADD_COLOR_NODES,
        "add_white_nodes": sd.DEFAULT_ADD_WHITE_NODES,
        "unique_node_count": stats.get("unique_node_count"),
        "unique_edge_count": stats.get("unique_edge_count"),
        "direct_query_edge_count": stats.get("direct_query_edge_count"),
        "neighbor_neighbor_edge_count": stats.get("neighbor_neighbor_edge_count"),
        "min_combined_score": stats.get("min_combined_score"),
        "max_combined_score": stats.get("max_combined_score"),
        "string_network_url": string_network_url or STRING_HOMEPAGE,
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
        "scientific_intro": SCIENTIFIC_INTRO,
        "direct_query_association_definition": DIRECT_QUERY_ASSOCIATION_DEF,
        "supplementary_scope": SUPPLEMENTARY_SCOPE,
        "supplementary_xlsx": xlsx_name if workbook_sha else None,
        "supplementary_xlsx_sha256": workbook_sha,
        "network_response_sha256": audit_meta.get("network_response_sha256"),
        "network_figure_sha256": figure_sha,
        "network_figure_relative_path": figure_rel,
        "network_figure_local_path": str(figure_local) if figure_local else None,
        "network_figure_width": figure_dims.get("width"),
        "network_figure_height": figure_dims.get("height"),
        "warnings": warnings,
        "presentation_item_key": f"string-{slugify(preferred_name.lower())}",
        "attempt_network_figure": section_cfg.attempt_network_figure,
    }
    write_json_atomic(attempt_dir / "summary.json", summary_payload)
    write_json_atomic(
        attempt_dir / "network_associations.json",
        {"edges": canonical.get("canonical_edges") or []},
    )
    write_json_atomic(
        attempt_dir / "network_nodes.json",
        {"nodes": canonical.get("nodes") or []},
    )
    write_json_atomic(
        attempt_dir / "direct_query_partners.json",
        {"partners": canonical.get("direct_query_partners") or []},
    )

    api_run_id = audit_meta.get("network_api_run_id")
    raw_artifact_id = audit_meta.get("network_raw_artifact_id")
    summary_rec = _evidence(
        dossier_run_id=run_id,
        gene_symbol=preferred_name,
        fact_type="section_5a_summary",
        key="section-5a-summary",
        value={**summary_payload, "attempt_dir": str(attempt_dir)},
        display_text=(
            f"{preferred_name} STRING bounded network "
            f"(nodes={summary_payload['unique_node_count']}, "
            f"edges={summary_payload['unique_edge_count']})."
        ),
        api_run_id=api_run_id,
        raw_artifact_id=raw_artifact_id,
    )
    _append_evidence(evidence_records, summary_rec, persist_db=persist_db)

    for i, row in enumerate(canonical.get("direct_query_partners") or []):
        partner = (
            row.get("preferred_name_b")
            if row.get("string_id_a") == string_id
            else row.get("preferred_name_a")
        )
        rec = _evidence(
            dossier_run_id=run_id,
            gene_symbol=preferred_name,
            fact_type="section_5a_direct_partner",
            key=f"section-5a-direct-{i}",
            value={**row, "api_run_id": api_run_id, "raw_artifact_id": raw_artifact_id},
            display_text=f"{preferred_name}–{partner} functional association",
            api_run_id=api_run_id,
            raw_artifact_id=raw_artifact_id,
        )
        _append_evidence(evidence_records, rec, persist_db=persist_db)

    if figure_rel and figure_sha:
        fig_rec = _evidence(
            dossier_run_id=run_id,
            gene_symbol=preferred_name,
            fact_type="section_5a_network_figure",
            key="section-5a-network-figure",
            value={
                "relative_path": figure_rel,
                "local_artifact_path": str(figure_local),
                "sha256": figure_sha,
                "width": figure_dims.get("width"),
                "height": figure_dims.get("height"),
                "media_type": "image/png",
                "api_run_id": audit_meta.get("figure_api_run_id"),
                "raw_artifact_id": audit_meta.get("figure_raw_artifact_id"),
            },
            display_text=f"{preferred_name} STRING network figure",
            api_run_id=audit_meta.get("figure_api_run_id"),
            raw_artifact_id=audit_meta.get("figure_raw_artifact_id"),
        )
        _append_evidence(evidence_records, fig_rec, persist_db=persist_db)

    if workbook_sha:
        supp_rec = _evidence(
            dossier_run_id=run_id,
            gene_symbol=preferred_name,
            fact_type="section_5a_supplementary_workbook",
            key="section-5a-supplementary-xlsx",
            value={
                "filename": xlsx_name,
                "relative_path": f"supplementary/{xlsx_name}",
                "local_artifact_path": str(xlsx_path),
                "sha256": workbook_sha,
                "supplementary_scope": SUPPLEMENTARY_SCOPE,
                "api_run_id": api_run_id,
                "raw_artifact_id": raw_artifact_id,
                "media_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            },
            display_text=f"Supplementary Material ({xlsx_name})",
            api_run_id=api_run_id,
            raw_artifact_id=raw_artifact_id,
        )
        _append_evidence(evidence_records, supp_rec, persist_db=persist_db)

    write_json_atomic(
        attempt_dir / "evidence.json",
        {
            "summary_evidence_id": summary_rec.id,
            "supplementary_xlsx_sha256": workbook_sha,
            "network_response_sha256": audit_meta.get("network_response_sha256"),
            "network_figure_sha256": figure_sha,
            "api_run_id": api_run_id,
            "raw_artifact_id": raw_artifact_id,
        },
    )

    artifacts = {
        "summary.json": "summary.json",
        "network_associations.json": "network_associations.json",
        "network_nodes.json": "network_nodes.json",
        "direct_query_partners.json": "direct_query_partners.json",
        "evidence.json": "evidence.json",
        MANIFEST_FILENAME: MANIFEST_FILENAME,
    }
    if workbook_sha:
        artifacts[f"supplementary/{xlsx_name}"] = f"supplementary/{xlsx_name}"
    if figure_rel:
        artifacts["network_figure"] = figure_rel

    manifest = {
        "gene_symbol": gene,
        "official_symbol": preferred_name,
        "attempt_dir": str(attempt_dir),
        "scientific_status": scientific,
        "visual_status": visual,
        "presentation_status": presentation,
        "resolved_string_id": string_id,
        "supplementary_xlsx": xlsx_name if workbook_sha else None,
        "supplementary_xlsx_sha256": workbook_sha,
        "network_response_sha256": audit_meta.get("network_response_sha256"),
        "network_figure_sha256": figure_sha,
        "string_version": sd.STRING_VERSION,
        "string_base_url": sd.STRING_BASE_URL,
        "parser_version": PARSER_VERSION,
        "supplementary_scope": SUPPLEMENTARY_SCOPE,
        "created_at": generation_ts,
    }
    write_json_atomic(attempt_dir / MANIFEST_FILENAME, manifest)

    coverage = list(state.get("coverage") or [])
    if scientific == STATUS_SUCCESS:
        cov_status = SourceStatus.success
    elif scientific == STATUS_NO_INTERACTIONS:
        cov_status = SourceStatus.partial
    else:
        cov_status = SourceStatus.failed
    coverage.append(
        SourceCoverageResult(
            dossier_run_id=run_id,
            source_name=sd.SOURCE_NAME,
            status=cov_status,
            evidence_record_count=1
            + len(canonical.get("direct_query_partners") or [])
            + (1 if figure_rel else 0)
            + (1 if workbook_sha else 0),
            error_message=None
            if scientific == STATUS_SUCCESS
            else f"section_5a scientific_status={scientific}",
            notes=f"visual_status={visual}; presentation_status={presentation}",
            report_sections_supported=[SECTION_PPI],
        )
    )

    section_status = {
        "section_key": "5a",
        "summary": summary_payload,
        "rendering_status": {
            "scientific_status": scientific,
            "visual_status": visual,
            "presentation_status": presentation,
            "overall_status": (
                STATUS_SUCCESS
                if scientific == STATUS_SUCCESS and presentation == STATUS_SUCCESS
                else presentation
                if scientific == STATUS_SUCCESS
                else scientific
            ),
        },
        "audit": {
            "gene_attempt_dir": str(attempt_dir),
            "artifacts": artifacts,
            **audit_meta,
            "parser_version": PARSER_VERSION,
            "supplementary_scope": SUPPLEMENTARY_SCOPE,
            "string_version": sd.STRING_VERSION,
            "string_base_url": sd.STRING_BASE_URL,
        },
    }
    write_json_atomic(attempt_dir / "section_5a_status.json", section_status)

    return {
        **state,
        "evidence_records": evidence_records,
        "coverage": coverage,
        "api_runs": api_runs,
        "raw_artifacts": raw_artifacts,
        "section_5a_status": section_status,
        "tool_results": tool_results,
    }


__all__ = [
    "DIRECT_QUERY_ASSOCIATION_DEF",
    "PARSER_VERSION",
    "SCIENTIFIC_INTRO",
    "SECTION_PPI",
    "STRING_HOMEPAGE",
    "STATUS_FAILED",
    "STATUS_IDENTIFIER_AMBIGUOUS",
    "STATUS_IDENTIFIER_NOT_FOUND",
    "STATUS_NO_INTERACTIONS",
    "STATUS_NOT_ATTEMPTED",
    "STATUS_PARTIAL",
    "STATUS_SOURCE_UNAVAILABLE",
    "STATUS_SUCCESS",
    "STATUS_UNAVAILABLE",
    "SUBSECTION_5A",
    "SUPPLEMENTARY_SCOPE",
    "NetworkEdgeRecord",
    "Section5aConfig",
    "accept_section_5a_report",
    "canonicalize_network",
    "evaluate_section_5a_complete",
    "map_network_edge",
    "node_generate_section_5a_derived_artifacts",
    "parse_network_edge",
    "write_string_workbook",
]
